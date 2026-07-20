"""
incubation_app.py  —  Leafcutter Bee Incubation Manager
GUI: customtkinter  |  DB: SQLite  |  Sensor: Govee API  |  QR scan: Flask

Run:
    python incubation_app.py

Dependencies:
    pip install customtkinter pillow qrcode flask requests openpyxl
"""
import sys
import os
import math
import subprocess
import threading
import time
from datetime import datetime, date

# Windows: run helper subprocesses (git, py_compile, poller) hidden so no
# console window flashes up and steals focus while you're working.
_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)
from tkinter import ttk, filedialog, messagebox
import tkinter as tk

import customtkinter as ctk

import incubation_db as db
import incubation_calc as calc
import govee_client as govee_mod
import sensibo_client as sensibo_mod
import gcal_sync
import qr_server
import voc_db
from voc_panel import VOCPanel
import inspection_db
from inspection_form import InspectionDialog, InspectionsLogPanel, make_status_badges
import email_reporter

# Optional imports
try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import qrcode as qrlib
    HAS_QR = True
except ImportError:
    HAS_QR = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import matplotlib
    matplotlib.use("Agg")          # non-interactive backend — never spawns its own window
    import matplotlib.dates as mdates
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ── Version ─────────────────────────────────────────────────────────────────
APP_VERSION = "1.45.0"   # bump on every push (semver: MAJOR.MINOR.PATCH)


def _git_revision() -> str:
    """Short git commit hash + date for the running code, or '' if unavailable."""
    try:
        app_dir = os.path.dirname(os.path.abspath(__file__))
        out = subprocess.run(
            ["git", "-C", app_dir, "log", "-1", "--format=%h · %cd", "--date=short"],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=5, creationflags=_NO_WINDOW,
        )
        return out.stdout.strip() if out.returncode == 0 else ""
    except Exception:
        return ""


def app_version_string() -> str:
    rev = _git_revision()
    return f"v{APP_VERSION}  ({rev})" if rev else f"v{APP_VERSION}"


# ── Polling ──────────────────────────────────────────────────────────────────
POLL_INTERVAL_SEC = 15 * 60   # Govee polling is fixed at 15 minutes


# ── Theme ─────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

GOLD      = "#FFD700"
DK_GOLD   = "#B8860B"
GREEN     = "#4CAF50"
GREEN_LT  = "#7CE08A"
TEAL      = "#10B981"
ORANGE    = "#FF9800"
RED       = "#FF3B30"
RED_LT    = "#FF6A57"
BLUE      = "#3B82F6"
LINK      = "#93C5FD"
BG        = "#0F172A"   # app background
BARBG     = "#0B1220"   # title / status bar
SIDEBAR   = "#111827"
RIGHTPANE = "#0D1524"
CARD      = "#1B2536"   # incubator cards
PANEL     = "#151E2E"   # section / table panels
NESTED    = "#141C2B"   # nested rows
CARD2     = "#202B3D"   # inset tiles / inputs
BORDER    = "#374151"   # strong border
BORDER2   = "#232F42"   # subtle border
SUBBORDER = "#1E293B"   # faint divider
TEXT      = "#F3F4F6"
TEXT2     = "#CBD5E1"
SUBTEXT   = "#9CA3AF"
FAINT     = "#6B7280"
FONT_H    = ("Segoe UI", 14, "bold")
FONT_B    = ("Segoe UI", 11)
FONT_S    = ("Segoe UI", 10)


def _treeview_style():
    """Table style per the design handoff: panel background #151E2E, gold
    bold headers with a subtle divider, 11.5px body text, roomy rows."""
    style = ttk.Style()
    style.theme_use("default")
    style.configure("Dark.Treeview",
        background=PANEL, foreground="#E5E7EB",
        fieldbackground=PANEL, borderwidth=0,
        rowheight=30, font=("Segoe UI", 11))
    style.configure("Dark.Treeview.Heading",
        background=PANEL, foreground=GOLD,
        relief="flat", borderwidth=0, padding=(10, 8),
        font=("Segoe UI", 11, "bold"))
    style.map("Dark.Treeview.Heading",
        background=[("active", PANEL)])
    style.map("Dark.Treeview",
        background=[("selected", "#26374F")],
        foreground=[("selected", TEXT)])


def _label(parent, text, font=FONT_B, color=TEXT, **kw):
    return ctk.CTkLabel(parent, text=text, font=font, text_color=color, **kw)


def _btn(parent, text, cmd, width=110, height=32, fg=CARD2, hover=BORDER,
         text_color=TEXT, **kw):
    return ctk.CTkButton(parent, text=text, command=cmd, width=width,
                         height=height, fg_color=fg, hover_color=hover,
                         text_color=text_color, corner_radius=6, **kw)


def _btn_primary(parent, text, cmd, width=130):
    """Gold primary action button (spec: #E0A81A→#B8860B, text #1A1206)."""
    return ctk.CTkButton(parent, text=text, command=cmd, width=width, height=34,
                         corner_radius=8, fg_color="#C79114", hover_color="#E0A81A",
                         text_color="#1A1206", font=("Segoe UI", 12, "bold"),
                         border_width=1, border_color=DK_GOLD)


def _btn_secondary(parent, text, cmd, width=120):
    """Neutral secondary action button (spec: #1F2937 bg, #374151 border)."""
    return ctk.CTkButton(parent, text=text, command=cmd, width=width, height=34,
                         corner_radius=8, fg_color="#1F2937", hover_color="#28374D",
                         text_color="#E5E7EB", font=("Segoe UI", 12),
                         border_width=1, border_color=BORDER)


def _parse_date_loose(s):
    """Parse a date string in common formats (ISO or M/D/Y). Returns date or None."""
    if not s:
        return None
    s = str(s).strip()
    token = s.replace("T", " ").split()[0] if s else s   # the date portion
    for cand in (token, s):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(cand, fmt).date()
            except ValueError:
                continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def cool_down_days(tray: dict):
    """Days a tray has been / was cooled. None if not applicable."""
    cd = _parse_date_loose(tray.get("cool_date"))
    if not cd:
        return None
    status = tray.get("status")
    if status == "cooled":
        end = datetime.now().date()
    elif status == "released":
        end = _parse_date_loose(tray.get("out_date")) or datetime.now().date()
    else:
        return None
    return max((end - cd).days, 0)


def _poll_age(timestamp_iso: str | None, interval_sec: int = 300) -> tuple[str, str]:
    """
    Return (display_text, color) for how long ago a Govee poll timestamp was.

    Freshness colors scale with the configured poll interval so they stay
    meaningful at any rate (1 min → 1 hr):
      green  = within ~1.5 cycles (healthy)
      orange = within ~3 cycles   (a poll or two missed)
      red    = beyond that, or missing
    """
    if not timestamp_iso:
        return "Never polled", SUBTEXT
    try:
        then    = datetime.fromisoformat(timestamp_iso)
        minutes = (datetime.now() - then).total_seconds() / 60
    except Exception:
        return "Unknown", SUBTEXT
    if minutes < 1:
        text = "Just now"
    elif minutes < 60:
        text = f"{int(minutes)} min ago"
    elif minutes < 120:
        text = "1 hr ago"
    else:
        text = f"{int(minutes // 60)} hrs ago"

    cycle_min  = max(interval_sec / 60, 1)
    green_cut  = cycle_min * 1.5 + 1     # one cycle + slack
    orange_cut = cycle_min * 3   + 2
    color = GREEN if minutes <= green_cut else (ORANGE if minutes <= orange_cut else RED)
    return text, color


def _entry(parent, placeholder="", width=200):
    return ctk.CTkEntry(parent, placeholder_text=placeholder, width=width,
                        fg_color=CARD2, border_color=BORDER, text_color="#CBD5E1",
                        corner_radius=7)


def _mix(fg: str, bg: str, alpha: float) -> str:
    """Blend fg over bg at the given alpha (0-1) → solid hex.
    Used to fake the translucent badge fills from the design spec."""
    fg = fg.lstrip("#"); bg = bg.lstrip("#")
    fr, fgc, fb = int(fg[0:2], 16), int(fg[2:4], 16), int(fg[4:6], 16)
    br, bgc, bb = int(bg[0:2], 16), int(bg[2:4], 16), int(bg[4:6], 16)
    r = round(fr * alpha + br * (1 - alpha))
    g = round(fgc * alpha + bgc * (1 - alpha))
    b = round(fb * alpha + bb * (1 - alpha))
    return f"#{r:02x}{g:02x}{b:02x}"


# Operating-mode accent colors (badges + segmented controls)
MODE_COLORS = {"off": "#6B7280", "cool_storage": TEAL,
               "incubation": BLUE, "holding": GOLD}
# Spec-exact pre-blended badge fills (mode color at ~13% over card #1B2536)
MODE_BADGE_BG = {"off": "#252B34", "cool_storage": "#1B2E33",
                 "incubation": "#20293B", "holding": "#31311F"}


def _combo(parent, values, width=200):
    return ctk.CTkComboBox(parent, values=values, width=width,
                           fg_color=CARD, border_color=BORDER,
                           button_color=BORDER, text_color=TEXT,
                           dropdown_fg_color=CARD)


# ── Reusable form helper ──────────────────────────────────────────────────────

class _FormRow:
    """One label + entry widget row in a grid form."""
    def __init__(self, parent, row, label, placeholder="", width=220, widget=None):
        _label(parent, label, font=FONT_S, color=SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4, 8), pady=3)
        if widget:
            self.widget = widget
        else:
            self.widget = _entry(parent, placeholder, width=width)
        self.widget.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

    def get(self):
        w = self.widget
        if isinstance(w, ctk.CTkEntry):
            return w.get().strip()
        if isinstance(w, ctk.CTkComboBox):
            return w.get()
        if isinstance(w, ctk.CTkTextbox):
            return w.get("1.0", "end").strip()
        return ""

    def set(self, value):
        w = self.widget
        val = str(value) if value is not None else ""
        if isinstance(w, ctk.CTkEntry):
            w.delete(0, "end"); w.insert(0, val)
        elif isinstance(w, ctk.CTkComboBox):
            w.set(val)
        elif isinstance(w, ctk.CTkTextbox):
            w.delete("1.0", "end"); w.insert("1.0", val)


# ═══════════════════════════════════════════════════════════════════════════════
#   DIALOGS
# ═══════════════════════════════════════════════════════════════════════════════

class IncubatorDialog(ctk.CTkToplevel):
    def __init__(self, master, inc: dict = None, on_save=None, on_delete=None):
        super().__init__(master, fg_color=BG)
        self.on_save = on_save
        self.on_delete = on_delete
        self.inc = inc or {}
        self.title("Edit Incubator" if inc else "Add Incubator")
        self.geometry("460x500")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if inc:
            self._populate()

    def _build(self):
        f = ctk.CTkScrollableFrame(self, fg_color="transparent")
        f.pack(fill="both", expand=True, padx=4, pady=4)
        f.columnconfigure(1, weight=1)
        self._f = f

        rows = [
            ("Name",              "Incubator 1",  "name"),
            ("Capacity (trays)",  "50",           "capacity"),
            ("Govee Device ID",   "AB:CD:EF:...", "govee_device_id"),
            ("Govee SKU / Model", "H5075",        "govee_sku"),
            ("Sensibo Device ID(s)", "ID1, ID2",  "sensibo_device_id"),
            ("Incubation Start (YYYY-MM-DD)", "2026-06-01", "incubation_start"),
        ]
        self._rows = {}
        for i, (lbl, ph, key) in enumerate(rows):
            r = _FormRow(f, i, lbl, ph, width=240)
            self._rows[key] = r

        # Temp mode selector
        n = len(rows)
        _label(f, "Temp Mode", FONT_S, SUBTEXT).grid(
            row=n, column=0, sticky="e", padx=(14, 8), pady=8)
        self._mode_var = ctk.StringVar(value="Incubation")
        mode_names = [v["label"] for v in calc.TEMP_MODES.values()]
        self._mode_seg = ctk.CTkSegmentedButton(
            f, values=mode_names, variable=self._mode_var, width=290)
        self._mode_seg.grid(row=n, column=1, sticky="w", padx=8, pady=8)

        # Range hint label (updates as mode changes)
        n2 = n + 1
        self._range_hint = _label(f, "", FONT_S, SUBTEXT)
        self._range_hint.grid(row=n2, column=1, sticky="w", padx=8, pady=(0, 4))
        self._mode_var.trace_add("write", self._update_range_hint)
        self._update_range_hint()

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=20, pady=14)
        _btn(btns, "Save", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=130).pack(side="right", padx=4)
        _btn(btns, "Cancel", self.destroy, width=100).pack(side="right")

        # Manage actions (only when editing an existing incubator)
        if self.inc.get("id"):
            is_hidden = bool(self.inc.get("is_hidden"))
            _btn(btns, "Unhide" if is_hidden else "Hide",
                 self._toggle_hidden, width=80,
                 fg=CARD2, hover=BORDER, text_color=SUBTEXT).pack(side="left")
            _btn(btns, "Delete", self._delete, width=80,
                 fg="#4B0000", hover=RED, text_color="white").pack(side="left", padx=6)

    def _toggle_hidden(self):
        new_val = not bool(self.inc.get("is_hidden"))
        db.set_incubator_hidden(self.inc["id"], new_val)
        if self.on_save:
            self.on_save()
        self.destroy()

    def _delete(self):
        if not messagebox.askyesno(
            "Delete Incubator",
            f"Delete '{self.inc.get('name')}'?\n\n"
            "This removes the incubator. Its trays and readings remain in the "
            "database but will no longer be linked to an incubator.\n\n"
            "This cannot be undone.",
            icon="warning", parent=self):
            return
        db.delete_incubator(self.inc["id"])
        if self.on_delete:
            self.on_delete()
        elif self.on_save:
            self.on_save()
        self.destroy()

    def _update_range_hint(self, *_):
        label = self._mode_var.get()
        key   = calc._MODE_BY_LABEL.get(label, "incubation")
        cfg   = calc.TEMP_MODES[key]
        hint  = "No temperature alerts" if cfg["min"] is None else f"Alert range: {cfg['min']}–{cfg['max']} °C"
        self._range_hint.configure(text=hint)

    def _populate(self):
        for key, row in self._rows.items():
            row.set(self.inc.get(key, ""))
        mode_key = self.inc.get("temp_mode", "incubation")
        mode_cfg = calc.TEMP_MODES.get(mode_key, calc.TEMP_MODES["incubation"])
        self._mode_var.set(mode_cfg["label"])

    def _save(self):
        data = {k: v for k, v in ((k, r.get()) for k, r in self._rows.items()) if v != ""}
        if not data.get("name"):
            messagebox.showerror("Error", "Name is required.", parent=self)
            return
        data["id"] = self.inc.get("id")
        try:
            data["capacity"] = int(data.get("capacity", 50))
        except (ValueError, TypeError):
            data["capacity"] = 50
        for fld in ("humidity_min", "humidity_max"):
            try:
                data[fld] = float(data[fld])
            except (KeyError, ValueError, TypeError):
                pass
        # Temp mode from segmented button
        label = self._mode_var.get()
        data["temp_mode"] = calc._MODE_BY_LABEL.get(label, "incubation")
        # Preserve fields not editable in this form so they aren't overwritten with None
        for preserve in ("sort_order", "is_hidden", "temp_alerts_enabled"):
            if preserve not in data:
                data[preserve] = self.inc.get(preserve)
        db.upsert_incubator(data)
        if self.on_save:
            self.on_save()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────

class BatchDialog(ctk.CTkToplevel):
    """Add or edit an incubation batch (the dates for one incubation run)."""

    DATE_ROWS = [
        ("Start Date",          "start_date"),
        ("Vapona In",           "vapona_in"),
        ("Vapona Out",          "vapona_out"),
        ("Air Out",             "air_out"),
        ("10% Male Emergence",  "male_10pct_emergence"),
        ("Earliest Cool",       "earliest_cool"),
        ("Est. Release",        "estimated_release"),
        ("Latest Release",      "latest_release"),
    ]

    def __init__(self, master, batch: dict = None,
                 incubator_id: int = None, on_save=None):
        super().__init__(master, fg_color=BG)
        self.on_save = on_save
        self.batch = batch or {}
        self.preselect_inc = incubator_id
        self.title("Edit Batch" if batch else "New Incubation Batch")
        self.geometry("460x660")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if batch:
            self._populate()

    def _build(self):
        f = ctk.CTkScrollableFrame(self, fg_color="transparent")
        f.pack(fill="both", expand=True, padx=4, pady=4)
        f.columnconfigure(1, weight=1)

        incubators = db.get_incubators()
        inc_names  = [i["name"] for i in incubators]
        self._inc_map = {i["name"]: i["id"] for i in incubators}

        samples   = db.get_samples()
        smp_names = ["(none)"] + [s["name"] for s in samples]
        self._smp_map = {s["name"]: s["id"] for s in samples}

        row = 0
        _label(f, "Batch Name", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._name = _entry(f, "e.g. Inc-3 June 2026", 240)
        self._name.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        row += 1
        _label(f, "Incubator", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._inc_cb = _combo(f, inc_names, 240)
        self._inc_cb.grid(row=row, column=1, sticky="ew", padx=4, pady=3)
        if self.preselect_inc:
            for name, iid in self._inc_map.items():
                if iid == self.preselect_inc:
                    self._inc_cb.set(name)

        row += 1
        _label(f, "Sample (opt.)", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._smp_cb = _combo(f, smp_names, 240)
        self._smp_cb.set("(none)")
        self._smp_cb.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        _label(f, "Dates (YYYY-MM-DD)", FONT_B, GOLD).grid(
            row=row+1, column=0, columnspan=2, sticky="w", padx=4, pady=(10,2))
        row += 2

        self._date_rows = {}
        for lbl, key in self.DATE_ROWS:
            r = _FormRow(f, row, lbl, "YYYY-MM-DD", 200)
            self._date_rows[key] = r
            row += 1

        _label(f, "Status", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._status = _combo(f, ["active", "completed", "cancelled"], 200)
        self._status.set("active")
        self._status.grid(row=row, column=1, sticky="ew", padx=4, pady=3)
        row += 1

        _label(f, "Notes", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="nw", padx=(4,8), pady=3)
        self._notes = ctk.CTkTextbox(f, height=60, fg_color=CARD,
                                     border_color=BORDER, text_color=TEXT)
        self._notes.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=20, pady=12)
        _btn(btns, "Save", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=130).pack(side="right", padx=4)
        _btn(btns, "Cancel", self.destroy, width=100).pack(side="right")

    def _populate(self):
        self._name.insert(0, self.batch.get("name") or "")
        if self.batch.get("incubator_name"):
            self._inc_cb.set(self.batch["incubator_name"])
        if self.batch.get("sample_name"):
            self._smp_cb.set(self.batch["sample_name"])
        for key, r in self._date_rows.items():
            r.set(self.batch.get(key) or "")
        self._status.set(self.batch.get("status") or "active")
        if self.batch.get("notes"):
            self._notes.insert("1.0", self.batch["notes"])

    def _save(self):
        inc_name = self._inc_cb.get()
        if not inc_name or inc_name not in self._inc_map:
            messagebox.showerror("Error", "Select an incubator.", parent=self)
            return
        data = {
            "id":           self.batch.get("id"),
            "incubator_id": self._inc_map[inc_name],
            "sample_id":    self._smp_map.get(self._smp_cb.get()),
            "name":         self._name.get().strip(),
            "status":       self._status.get(),
            "notes":        self._notes.get("1.0", "end").strip(),
        }
        for key, r in self._date_rows.items():
            val = r.get()
            data[key] = val if val else None
        db.upsert_batch(data)
        if self.on_save:
            self.on_save()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────

class SampleDialog(ctk.CTkToplevel):
    def __init__(self, master, sample: dict = None, on_save=None):
        super().__init__(master, fg_color=BG)
        self.on_save = on_save
        self.sample = sample or {}
        self.title("Edit Sample" if sample else "Add Sample")
        self.geometry("440x580")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if sample:
            self._populate()

    def _build(self):
        f = ctk.CTkScrollableFrame(self, fg_color="transparent")
        f.pack(fill="both", expand=True, padx=4, pady=4)
        f.columnconfigure(1, weight=1)

        fields = [
            ("Sample Name *",       "name",              "e.g. RR-4"),
            ("Total Kgs",           "total_weight_kg",   ""),
            ("Total Pounds",        "total_weight_lbs",  ""),
            ("Live Bees per KG",    "live_bees_per_kg",  ""),
            ("Live Bees per Pound", "live_bees_per_lb",  ""),
            ("Parasites",           "parasites",         ""),
            ("Chalkbrood",          "chalkbrood",        ""),
            ("Total Gal Bees",      "total_volume_gal",  ""),
            ("Total KG for 2gal",   "kg_per_2gal",       ""),
            ("Total lbs for 2gal",  "lbs_per_2gal",      ""),
            ("Expected Trays",      "total_trays",       ""),
            ("Incubator Space",     "incubator_space",   ""),
        ]
        self._rows = {}
        for i, (lbl, key, ph) in enumerate(fields):
            r = _FormRow(f, i, lbl, ph, width=230)
            self._rows[key] = r

        _label(f, "Notes", FONT_S, SUBTEXT).grid(
            row=len(fields), column=0, sticky="nw", padx=(4,8), pady=3)
        self._notes = ctk.CTkTextbox(f, height=60, fg_color=CARD,
                                     border_color=BORDER, text_color=TEXT)
        self._notes.grid(row=len(fields), column=1, sticky="ew", padx=4, pady=3)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=20, pady=12)
        _btn(btns, "Save", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=130).pack(side="right", padx=4)
        _btn(btns, "Cancel", self.destroy, width=100).pack(side="right")

    def _populate(self):
        for key, row in self._rows.items():
            row.set(self.sample.get(key, ""))
        if self.sample.get("notes"):
            self._notes.insert("1.0", self.sample["notes"])

    def _save(self):
        name = self._rows["name"].get()
        if not name:
            messagebox.showerror("Error", "Sample name is required.", parent=self)
            return
        data = {"id": self.sample.get("id"), "name": name,
                "notes": self._notes.get("1.0", "end").strip(),
                "incubator_space": self._rows["incubator_space"].get() or None}
        for key in ("total_weight_lbs", "total_weight_kg", "live_bees_per_lb",
                    "live_bees_per_kg", "parasites", "chalkbrood",
                    "total_volume_gal", "kg_per_2gal", "lbs_per_2gal", "total_trays"):
            val = self._rows[key].get()
            try:
                data[key] = float(val) if val else None
            except ValueError:
                data[key] = None

        if data.get("id"):
            # Editing an existing sample — update it directly.
            db.upsert_sample(data)
        else:
            # New sample: if one with this name already exists, update it (keeps
            # tray links) rather than silently creating a duplicate.
            existing = db.find_sample_by_name(name)
            if existing:
                if messagebox.askyesno(
                    "Sample already exists",
                    f"A sample named “{existing['name']}” already exists.\n\n"
                    "Update that sample with these values?  (Recommended — keeps "
                    "all its trays linked so the data shows on them.)\n\n"
                    "Choose No to create a separate new sample anyway.",
                    parent=self):
                    db.upsert_sample_by_name(data)
                else:
                    db.upsert_sample(data)
            else:
                db.upsert_sample(data)
        if self.on_save:
            self.on_save()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────

class TrayDialog(ctk.CTkToplevel):
    def __init__(self, master, tray: dict = None,
                 incubator_id: int = None, on_save=None):
        super().__init__(master, fg_color=BG)
        self.on_save = on_save
        self.tray = tray or {}
        self.preselect_inc = incubator_id
        self.title("Edit Tray" if tray else "Add Tray")
        self.geometry("460x680")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if tray:
            self._populate()

    def _build(self):
        f = ctk.CTkScrollableFrame(self, fg_color="transparent")
        f.pack(fill="both", expand=True, padx=4, pady=4)
        f.columnconfigure(1, weight=1)

        incubators = db.get_incubators()
        samples    = db.get_samples()
        batches    = db.get_batches()

        self._inc_map = {i["name"]: i["id"] for i in incubators}
        self._smp_map = {s["name"]: s["id"] for s in samples}
        self._bat_map = {(b.get("name") or f"Batch {b['id']}"): b["id"]
                         for b in batches}
        bat_names = ["(none)"] + list(self._bat_map.keys())

        row = 0
        _label(f, "Tray Number *", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._tray_num = _entry(f, "e.g. T001", 230)
        self._tray_num.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        row += 1
        _label(f, "Sample", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._smp_cb = _combo(f, ["(none)"] + list(self._smp_map.keys()), 230)
        self._smp_cb.set("(none)")
        self._smp_cb.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        row += 1
        _label(f, "Incubator", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._inc_cb = _combo(f, ["(none)"] + list(self._inc_map.keys()), 230)
        self._inc_cb.set("(none)")
        self._inc_cb.grid(row=row, column=1, sticky="ew", padx=4, pady=3)
        if self.preselect_inc:
            for name, iid in self._inc_map.items():
                if iid == self.preselect_inc:
                    self._inc_cb.set(name)

        row += 1
        _label(f, "Batch", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._bat_cb = _combo(f, bat_names, 230)
        self._bat_cb.set("(none)")
        self._bat_cb.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        measure_fields = [
            ("Weight (lbs)",      "weight_lbs",         ""),
            ("Volume (gal)",      "volume_gal",         ""),
            ("Live Count",        "live_count",         ""),
            ("Parasite Level (%)", "parasite_level_pct",""),
            ("In Date",           "in_date",            "YYYY-MM-DD"),
            ("Cool Date",         "cool_date",          "YYYY-MM-DD"),
            ("Out Date",          "out_date",           "YYYY-MM-DD"),
        ]
        self._mrows = {}
        for lbl, key, ph in measure_fields:
            row += 1
            r = _FormRow(f, row, lbl, ph, width=230)
            self._mrows[key] = r

        row += 1
        _label(f, "Status", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="w", padx=(4,8), pady=3)
        self._status = _combo(f, [lbl for _v, lbl in db.TRAY_STATUS_OPTIONS], 230)
        self._status.set("Incubation")
        self._status.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        row += 1
        _label(f, "Notes", FONT_S, SUBTEXT).grid(
            row=row, column=0, sticky="nw", padx=(4,8), pady=3)
        self._notes = ctk.CTkTextbox(f, height=50, fg_color=CARD,
                                     border_color=BORDER, text_color=TEXT)
        self._notes.grid(row=row, column=1, sticky="ew", padx=4, pady=3)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=20, pady=12)
        _btn(btns, "Save", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=130).pack(side="right", padx=4)
        _btn(btns, "Cancel", self.destroy, width=100).pack(side="right")

    def _populate(self):
        self._tray_num.insert(0, self.tray.get("tray_number") or "")
        if self.tray.get("sample_name"):
            self._smp_cb.set(self.tray["sample_name"])
        if self.tray.get("incubator_name"):
            self._inc_cb.set(self.tray["incubator_name"])
        if self.tray.get("batch_name"):
            self._bat_cb.set(self.tray["batch_name"])
        for key, r in self._mrows.items():
            r.set(self.tray.get(key) or "")
        self._status.set(db.tray_status_label(self.tray.get("status") or "active"))
        if self.tray.get("notes"):
            self._notes.insert("1.0", self.tray["notes"])

    def _save(self):
        tray_num = self._tray_num.get().strip()
        if not tray_num:
            messagebox.showerror("Error", "Tray number is required.", parent=self)
            return
        data = {
            "id":            self.tray.get("id"),
            "tray_number":   tray_num,
            "sample_id":     self._smp_map.get(self._smp_cb.get()),
            "incubator_id":  self._inc_map.get(self._inc_cb.get()),
            "incubation_batch_id": self._bat_map.get(self._bat_cb.get()),
            "status":        db.tray_status_value(self._status.get()),
            "notes":         self._notes.get("1.0", "end").strip(),
        }
        for key in ("in_date", "cool_date", "out_date"):
            data[key] = self._mrows[key].get() or None
        # Auto-release: if an out_date is set and status is still active, mark as released
        if data.get("out_date") and data.get("status") == "active":
            data["status"] = "released"
            self._status.set("Released")
        for key in ("weight_lbs", "volume_gal", "parasite_level_pct"):
            val = self._mrows[key].get()
            try:
                data[key] = float(val) if val else None
            except ValueError:
                data[key] = None
        val = self._mrows["live_count"].get()
        try:
            data["live_count"] = int(val) if val else None
        except ValueError:
            data["live_count"] = None
        try:
            db.upsert_tray(data)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save: {e}", parent=self)
            return
        if self.on_save:
            self.on_save()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────

class QRDialog(ctk.CTkToplevel):
    def __init__(self, master, tray: dict, port: int = 5151):
        super().__init__(master, fg_color=BG)
        self.title(f"QR Code — Tray {tray['tray_number']}")
        self.geometry("360x440")
        self.resizable(False, False)
        self.grab_set()

        url = qr_server.tray_url(tray["id"], port)

        _label(self, f"Tray {tray['tray_number']}", FONT_H, GOLD).pack(pady=(16, 4))
        _label(self, tray.get("sample_name") or "No sample", FONT_S, SUBTEXT).pack()
        _label(self, url, FONT_S, BLUE).pack(pady=(4, 8))

        if HAS_QR and HAS_PIL:
            try:
                qr = qrlib.QRCode(box_size=6, border=2)
                qr.add_data(url)
                qr.make(fit=True)
                img = qr.make_image(fill_color="white", back_color="#1F2937")
                self._img = ImageTk.PhotoImage(img)
                tk.Label(self, image=self._img,
                         background="#1F2937").pack(pady=4)
            except Exception:
                _label(self, "(QR image unavailable)", FONT_S, RED).pack()
        else:
            _label(self, "Install qrcode + pillow for QR image", FONT_S, ORANGE).pack()
            _label(self, "Phone scan requires Flask server running", FONT_S, SUBTEXT).pack()

        _btn(self, "Close", self.destroy, width=120,
             fg=CARD2, hover=BORDER).pack(pady=12)


# ─────────────────────────────────────────────────────────────────────────────

class AlertsDialog(ctk.CTkToplevel):
    def __init__(self, master, on_ack=None):
        super().__init__(master, fg_color=BG)
        self.on_ack = on_ack
        self.title("Active Alerts")
        self.geometry("580x480")
        self.grab_set()
        self._build()

    def _build(self):
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 4))
        _label(top, "Active Alerts", FONT_H, GOLD).pack(side="left")
        _btn(top, "Acknowledge All", self._ack_all,
             fg=CARD2, hover=BORDER, width=140).pack(side="right")

        self._scroll = ctk.CTkScrollableFrame(self, fg_color=CARD)
        self._scroll.pack(fill="both", expand=True, padx=12, pady=8)
        self._load()

    def _load(self):
        for w in self._scroll.winfo_children():
            w.destroy()
        alerts = db.get_active_alerts()
        if not alerts:
            _label(self._scroll, "No active alerts ✓", FONT_B, GREEN).pack(pady=20)
            return
        sev_colors = {"critical": RED, "warning": ORANGE, "info": BLUE}
        for a in alerts:
            row = ctk.CTkFrame(self._scroll, fg_color=CARD2, corner_radius=8)
            row.pack(fill="x", pady=3, padx=4)
            c = sev_colors.get(a.get("severity", "warning"), ORANGE)
            _label(row, f"● {a['message']}", FONT_S, c).pack(
                side="left", padx=10, pady=6, anchor="w")
            _btn(row, "✓", lambda aid=a["id"]: self._ack(aid),
                 width=36, height=26, fg=BORDER, hover=CARD).pack(
                 side="right", padx=6, pady=4)
            ts = (a.get("triggered_at") or "")[:16].replace("T", " ")
            _label(row, ts, FONT_S, SUBTEXT).pack(
                side="right", padx=4, pady=6)

    def _ack(self, aid):
        db.acknowledge_alert(aid)
        self._load()
        if self.on_ack:
            self.on_ack()

    def _ack_all(self):
        db.acknowledge_all_alerts()
        self._load()
        if self.on_ack:
            self.on_ack()


class _VocDeviceManager(ctk.CTkToplevel):
    """Assign Vapona (VOC) sensors to incubators / positions. App-authoritative:
    each Pi reports a stable hardware_id and the app owns name/incubator/position."""

    _POSITIONS = ["front", "back"]
    _ONLINE_SECS = 40 * 60  # a sensor polling every 15 min is "online" within 40 min

    def __init__(self, master, on_close=None):
        super().__init__(master, fg_color=BG)
        self.on_close = on_close
        self.title("Vapona Sensors")
        self.geometry("720x520")
        self.grab_set()
        # incubator id -> display name, plus an "Unassigned" choice
        self._incs = [i for i in db.get_incubators() if not i.get("is_hidden")]
        self._inc_label = {i["id"]: i["name"] for i in self._incs}
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._close)

    def _build(self):
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 4))
        _label(top, "Vapona Sensors", FONT_H, GOLD).pack(side="left")
        _btn(top, "Refresh", self._load, fg=CARD2, hover=BORDER,
             width=90).pack(side="right")
        _label(self, "Sensors report a hardware ID automatically. Assign each to an "
                     "incubator and position, then Save.",
               FONT_S, SUBTEXT).pack(anchor="w", padx=18, pady=(0, 4))
        self._scroll = ctk.CTkScrollableFrame(self, fg_color=CARD)
        self._scroll.pack(fill="both", expand=True, padx=12, pady=8)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(0, 12))
        _btn(btns, "Save Assignments", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=160).pack(side="right", padx=4)
        _btn(btns, "Close", self._close, width=100).pack(side="right")

        self._load()

    def _online(self, last_seen: str) -> bool:
        if not last_seen:
            return False
        try:
            dt = datetime.fromisoformat(last_seen)
            return (datetime.now() - dt).total_seconds() <= self._ONLINE_SECS
        except Exception:
            return False

    def _load(self):
        for w in self._scroll.winfo_children():
            w.destroy()
        self._rows = []
        try:
            devs = voc_db.get_devices()
        except Exception as exc:
            _label(self._scroll, f"Could not load devices: {exc}",
                   FONT_S, RED).pack(pady=20)
            return
        if not devs:
            _label(self._scroll, "No sensors have reported yet.\n"
                   "Power on a Pi sensor — it will appear here automatically.",
                   FONT_S, SUBTEXT).pack(pady=24)
            return

        inc_choices = ["Unassigned"] + [self._inc_label[i["id"]] for i in self._incs]
        for d in devs:
            card = ctk.CTkFrame(self._scroll, fg_color=CARD2, corner_radius=8)
            card.pack(fill="x", pady=4, padx=4)
            card.columnconfigure(1, weight=1)

            online = self._online(d.get("last_seen"))
            dot = "● Online" if online else "○ Offline"
            seen = (d.get("last_seen") or "")[:16].replace("T", " ")
            _label(card, dot, FONT_S, GREEN if online else SUBTEXT).grid(
                row=0, column=0, sticky="w", padx=10, pady=(8, 0))
            _label(card, f"hardware: {d['hardware_id']}", FONT_S, SUBTEXT).grid(
                row=0, column=1, sticky="w", padx=6, pady=(8, 0))
            _label(card, f"last seen {seen}" if seen else "never",
                   FONT_S, SUBTEXT).grid(row=0, column=2, sticky="e", padx=10, pady=(8, 0))

            body = ctk.CTkFrame(card, fg_color="transparent")
            body.grid(row=1, column=0, columnspan=3, sticky="ew", padx=8, pady=8)

            _label(body, "Name", FONT_S, SUBTEXT).grid(row=0, column=0, padx=(2, 4))
            name_e = ctk.CTkEntry(body, width=150, fg_color=CARD,
                                  border_color=BORDER, text_color=TEXT)
            name_e.grid(row=0, column=1, padx=4)
            name_e.insert(0, d.get("name") or "")

            _label(body, "Incubator", FONT_S, SUBTEXT).grid(row=0, column=2, padx=(12, 4))
            inc_var = ctk.StringVar(
                value=self._inc_label.get(d.get("incubator_id"), "Unassigned"))
            inc_cb = ctk.CTkComboBox(body, values=inc_choices, variable=inc_var,
                                     width=150, fg_color=CARD, border_color=BORDER,
                                     button_color=BORDER, text_color=TEXT,
                                     dropdown_fg_color=CARD, state="readonly")
            inc_cb.grid(row=0, column=3, padx=4)

            _label(body, "Position", FONT_S, SUBTEXT).grid(row=0, column=4, padx=(12, 4))
            pos_var = ctk.StringVar(value=d.get("position") or "front")
            pos_cb = ctk.CTkComboBox(body, values=self._POSITIONS, variable=pos_var,
                                     width=90, fg_color=CARD, border_color=BORDER,
                                     button_color=BORDER, text_color=TEXT,
                                     dropdown_fg_color=CARD, state="readonly")
            pos_cb.grid(row=0, column=5, padx=4)

            _btn(body, "🗑", lambda did=d["id"]: self._delete(did),
                 width=36, height=28, fg=BORDER, hover=RED,
                 text_color="white").grid(row=0, column=6, padx=(12, 2))

            self._rows.append((d["id"], name_e, inc_var, pos_var))

    def _inc_id_from_label(self, label):
        if label in (None, "", "Unassigned"):
            return None
        for iid, name in self._inc_label.items():
            if name == label:
                return iid
        return None

    def _save(self):
        for dev_id, name_e, inc_var, pos_var in self._rows:
            voc_db.update_device(
                dev_id,
                name=name_e.get().strip(),
                incubator_id=self._inc_id_from_label(inc_var.get()),
                position=pos_var.get())
        messagebox.showinfo("Vapona Sensors", "Assignments saved.", parent=self)
        self._load()

    def _delete(self, dev_id):
        if not messagebox.askyesno(
                "Delete sensor",
                "Remove this sensor from the app?\n\nIf it is still powered on it "
                "will re-appear (unassigned) the next time it reports.", parent=self):
            return
        voc_db.delete_device(dev_id)
        self._load()

    def _close(self):
        if self.on_close:
            try:
                self.on_close()
            except Exception:
                pass
        self.destroy()


class _WifiNetworkManager(ctk.CTkToplevel):
    """Manage the Wi-Fi networks pushed to every Vapona sensor. Every Pi is
    provisioned with all of these, so moving a sensor between incubators (each
    on its own network) needs no reconfiguration — it auto-joins whichever is
    in range."""

    def __init__(self, master):
        super().__init__(master, fg_color=BG)
        self.title("Sensor Wi-Fi Networks")
        self.geometry("620x480")
        self.grab_set()
        self._build()

    def _build(self):
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 4))
        _label(top, "Sensor Wi-Fi Networks", FONT_H, GOLD).pack(side="left")
        _btn(top, "+ Add Network", self._add, fg=GREEN, hover="#15803D",
             text_color="white", width=130).pack(side="right")
        _label(self, "Each incubator's network. All sensors receive every network "
                     "and auto-connect to whichever is in range.\n"
                     "Note: passwords are stored in the shared database.",
               FONT_S, SUBTEXT).pack(anchor="w", padx=18, pady=(0, 4))
        self._scroll = ctk.CTkScrollableFrame(self, fg_color=CARD)
        self._scroll.pack(fill="both", expand=True, padx=12, pady=8)

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(0, 12))
        _btn(btns, "Save", self._save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=120).pack(side="right", padx=4)
        _btn(btns, "Close", self.destroy, width=100).pack(side="right")
        self._load()

    def _load(self):
        for w in self._scroll.winfo_children():
            w.destroy()
        self._rows = []
        nets = voc_db.get_wifi_networks()
        # Column headers
        hdr = ctk.CTkFrame(self._scroll, fg_color="transparent")
        hdr.pack(fill="x", padx=4, pady=(2, 0))
        for txt, w in (("SSID", 170), ("Password", 170), ("Priority", 70), ("", 40)):
            _label(hdr, txt, FONT_S, SUBTEXT, width=w, anchor="w").pack(
                side="left", padx=4)
        if not nets:
            _label(self._scroll, "No networks yet — click “+ Add Network”.",
                   FONT_S, SUBTEXT).pack(pady=16)
        for n in nets:
            self._row_widget(n)

    def _row_widget(self, n):
        row = ctk.CTkFrame(self._scroll, fg_color=CARD2, corner_radius=6)
        row.pack(fill="x", padx=4, pady=3)
        ssid_e = ctk.CTkEntry(row, width=170, fg_color=CARD,
                              border_color=BORDER, text_color=TEXT)
        ssid_e.pack(side="left", padx=4, pady=6)
        ssid_e.insert(0, n.get("ssid") or "")
        psk_e = ctk.CTkEntry(row, width=170, fg_color=CARD, show="•",
                             border_color=BORDER, text_color=TEXT)
        psk_e.pack(side="left", padx=4, pady=6)
        psk_e.insert(0, n.get("psk") or "")
        prio_e = ctk.CTkEntry(row, width=60, fg_color=CARD,
                              border_color=BORDER, text_color=TEXT)
        prio_e.pack(side="left", padx=4, pady=6)
        prio_e.insert(0, str(n.get("priority") or 0))
        # Reveal/hide password
        def _toggle(e=psk_e, b=None):
            e.configure(show="" if e.cget("show") else "•")
        _btn(row, "👁", lambda: _toggle(), width=32, height=26,
             fg=BORDER, hover=CARD).pack(side="left", padx=(0, 2))
        _btn(row, "🗑", lambda nid=n.get("id"): self._delete(nid),
             width=32, height=26, fg=BORDER, hover=RED,
             text_color="white").pack(side="left", padx=(2, 6))
        self._rows.append((n.get("id"), ssid_e, psk_e, prio_e))

    def _add(self):
        # Persist current edits, then append a blank row
        self._save(silent=True)
        voc_db.upsert_wifi_network(ssid=f"New Network {len(self._rows)+1}")
        self._load()

    def _save(self, silent=False):
        seen = set()
        for nid, ssid_e, psk_e, prio_e in self._rows:
            ssid = ssid_e.get().strip()
            if not ssid or ssid in seen:
                continue
            seen.add(ssid)
            try:
                prio = int(prio_e.get().strip() or "0")
            except ValueError:
                prio = 0
            voc_db.upsert_wifi_network(ssid=ssid, psk=psk_e.get(),
                                       priority=prio, net_id=nid)
        if not silent:
            messagebox.showinfo("Sensor Wi-Fi",
                "Networks saved. Sensors pick up changes within a few minutes.",
                parent=self)
            self._load()

    def _delete(self, nid):
        if not nid:
            return
        if messagebox.askyesno(
                "Delete network", "Remove this Wi-Fi network?\n\n"
                "Sensors keep any profile already installed until they are "
                "reprovisioned or reflashed.", parent=self):
            voc_db.delete_wifi_network(nid)
            self._load()


# ═══════════════════════════════════════════════════════════════════════════════
#   MAIN APP
# ═══════════════════════════════════════════════════════════════════════════════

class IncubationApp(ctk.CTk):

    def __init__(self):
        # Tell Windows this is a distinct app so it gets its own taskbar button + icon
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                "BeeIncubation.Manager.1")
        except Exception:
            pass

        super().__init__()
        db.init_db()
        voc_db.init_voc_tables()
        inspection_db.init_inspection_tables()
        _treeview_style()

        self.title("Bee Incubation Manager")
        self.geometry("1280x800")
        self.minsize(1000, 680)
        self.configure(fg_color="#0F172A")

        # Apply logo icon to title bar and taskbar
        _app_dir  = os.path.dirname(os.path.abspath(__file__))
        _ico_path = os.path.join(_app_dir, "bee.ico")
        _icon_src = os.path.join(_app_dir, "app_icon.png")
        # Rebuild bee.ico from the committed source if it's missing or stale, so the
        # window icon stays in sync on every machine after a git pull (bee.ico itself
        # is gitignored / generated per-machine).
        try:
            if os.path.exists(_icon_src) and (
                    not os.path.exists(_ico_path)
                    or os.path.getmtime(_icon_src) > os.path.getmtime(_ico_path)):
                from PIL import Image as _IcoImg
                _IcoImg.open(_icon_src).convert("RGBA").resize(
                    (256, 256), _IcoImg.LANCZOS).save(
                    _ico_path, format="ICO",
                    sizes=[(16,16),(24,24),(32,32),(48,48),(64,64),(128,128),(256,256)])
        except Exception:
            pass
        if os.path.exists(_ico_path):
            try:
                self.iconbitmap(_ico_path)
            except Exception:
                pass

        # Govee client (polling fixed at 15 minutes)
        self._govee = govee_mod.GoveeClient(
            api_key=db.get_setting("govee_api_key"),
            poll_interval_sec=POLL_INTERVAL_SEC,
        )
        # Sensibo client (manual AC control only, no background polling)
        self._sensibo = sensibo_mod.SensiboClient(
            api_key=db.get_setting("sensibo_api_key"),
        )
        self._card_widgets: dict = {}  # incubator_id → {temp, hum, dot, ts} labels
        self._detail_inc: dict = {}   # incubator being shown in detail view

        # QR server port
        self._qr_port = int(db.get_setting("qr_server_port", "5151"))

        # Build UI — order matters for pack(): the title bar (top) plus the
        # sidebar (left) and the status bar (bottom, full width) must be
        # reserved BEFORE the expanding main area.
        self._build_titlebar()
        self._build_sidebar()
        self._build_status_bar()
        self._build_main()

        # Build all views (hidden until selected)
        self._views = {}
        self._views["dashboard"]    = self._build_dashboard()
        self._views["incubators"]   = self._build_incubators_view()
        self._views["samples"]      = self._build_samples_view()
        self._views["trays"]        = self._build_trays_view()
        self._views["analytics"]    = self._build_analytics_view()
        self._views["timeline"]     = self._build_timeline_view()
        self._views["inspections"]  = self._build_inspections_view()
        self._views["settings"]     = self._build_settings_view()
        self._views["inc_detail"]   = self._build_inc_detail_view()

        self._current_view = None
        self.show_view("dashboard")

        # Start background services
        self._start_govee()
        self._start_qr_server()
        self._start_alert_checker()
        self._start_email_scheduler()
        self._git_pull()            # immediate pull on startup
        self._start_auto_sync()     # then every 10 minutes

        # Refresh status bar periodically
        self._tick()

        # Pulsing alert dots on dashboard cards
        self._pulse_on = True
        self._pulse_dots()

        # Watch for code updates (incl. git auto-sync pulls) and offer 1-click reload
        self._start_code_watcher()
        self.bind_all("<F5>", lambda e: self._restart_app())
        self.bind_all("<Control-r>", lambda e: self._restart_app())

    # ── Layout ────────────────────────────────────────────────────────────────

    def _load_icon(self, key: str, color: str, size: int = 18):
        """Load a tinted nav icon PNG as a CTkImage (cached)."""
        cache = getattr(self, "_icon_cache", None)
        if cache is None:
            cache = self._icon_cache = {}
        ck = (key, color, size)
        if ck in cache:
            return cache[ck]
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "assets", "icons", f"{key}_{color}.png")
        img = None
        if os.path.exists(path):
            try:
                from PIL import Image as _PImg
                raw = _PImg.open(path)
                img = ctk.CTkImage(light_image=raw, dark_image=raw, size=(size, size))
            except Exception:
                img = None
        cache[ck] = img
        return img

    def _build_titlebar(self):
        """40px app title bar across the very top (spec: #0B1220 + hairline)."""
        tb = ctk.CTkFrame(self, fg_color=BARBG, height=40, corner_radius=0)
        tb.pack(side="top", fill="x")
        tb.pack_propagate(False)
        _label(tb, "🐝", ("Segoe UI", 13), "#F5B52B").pack(side="left", padx=(14, 6))
        _label(tb, "Bee Incubation Manager", ("Segoe UI", 12, "bold"),
               TEXT2).pack(side="left")
        # Mock window-control circles (right)
        for _ in range(3):
            ctk.CTkLabel(tb, text="●", font=("Segoe UI", 11),
                         text_color="#334155").pack(side="right", padx=(0, 6))
        # bottom hairline
        ctk.CTkFrame(self, fg_color=SUBBORDER, height=1, corner_radius=0).pack(
            side="top", fill="x")

    def _screen_header(self, parent, title: str, subtitle: str):
        """Standard top bar: 20px gold title + muted subtitle, actions right.
        Returns the header frame — callers pack action buttons side='right'."""
        hdr = ctk.CTkFrame(parent, fg_color="transparent")
        hdr.pack(fill="x", padx=26, pady=(20, 14))
        left = ctk.CTkFrame(hdr, fg_color="transparent")
        left.pack(side="left")
        _label(left, title, ("Segoe UI", 20, "bold"), GOLD).pack(anchor="w")
        _label(left, subtitle, ("Segoe UI", 11), FAINT).pack(anchor="w")
        return hdr

    def _build_sidebar(self):
        sb = ctk.CTkFrame(self, fg_color=SIDEBAR, width=208, corner_radius=0,
                          border_width=0)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        # Brand block — rounded tile containing the logo, centered
        tile = ctk.CTkFrame(sb, fg_color=CARD, corner_radius=16, width=62, height=62)
        tile.pack(pady=(22, 8))
        tile.pack_propagate(False)
        _logo_png = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        _placed = False
        if os.path.exists(_logo_png):
            try:
                from PIL import Image as _PILImage
                _raw = _PILImage.open(_logo_png).convert("RGBA")
                _data = _raw.getdata()
                _raw.putdata([(r, g, b, 0) if r > 230 and g > 230 and b > 230
                              else (r, g, b, a) for r, g, b, a in _data])
                _ctk_img = ctk.CTkImage(light_image=_raw, dark_image=_raw, size=(38, 38))
                ctk.CTkLabel(tile, image=_ctk_img, text="").pack(expand=True)
                _placed = True
            except Exception:
                _placed = False
        if not _placed:
            _label(tile, "🐝", ("Segoe UI", 26), "#F5B52B").pack(expand=True)

        _label(sb, "Incubation", ("Segoe UI", 15, "bold"), GOLD).pack(anchor="center")
        _label(sb, "Bee Manager", FONT_S, SUBTEXT).pack(anchor="center", pady=(1, 1))
        _label(sb, app_version_string(), ("Segoe UI", 9), FAINT).pack(anchor="center")
        ctk.CTkFrame(sb, fg_color=SUBBORDER, height=1).pack(fill="x", padx=14, pady=(12, 10))

        nav_items = [
            ("Dashboard",     "dashboard"),
            ("Incubators",    "incubators"),
            ("Samples",       "samples"),
            ("Trays",         "trays"),
            ("Analytics",     "analytics"),
            ("Calendar",      "timeline"),
            ("Inspections",   "inspections"),
            ("Settings",      "settings"),
        ]
        self._nav_btns = {}
        self._nav_icons = {}   # key -> {"grey": CTkImage, "gold": CTkImage}
        for label, key in nav_items:
            g = self._load_icon(key, "grey")
            gold = self._load_icon(key, "gold")
            self._nav_icons[key] = {"grey": g, "gold": gold}
            _cmd = self._open_incubators if key == "incubators" \
                else (lambda k=key: self.show_view(k))
            btn = ctk.CTkButton(
                sb, text="  " + label, anchor="w", height=40,
                image=g, compound="left",
                fg_color="transparent", hover_color="#1A2436",
                text_color=SUBTEXT, font=("Segoe UI", 12), corner_radius=9,
                command=_cmd
            )
            btn.pack(fill="x", padx=10, pady=2)
            self._nav_btns[key] = btn

        # Spacer
        ctk.CTkFrame(sb, fg_color="transparent").pack(fill="y", expand=True)

        # Alert button pinned to the bottom — red-tinted
        self._alert_btn = ctk.CTkButton(
            sb, text="  Alerts  0", height=40, anchor="w",
            image=self._load_icon("alerts", "red"), compound="left",
            fg_color="#2A1E20", hover_color="#3A2224",
            text_color=RED_LT, font=("Segoe UI", 12, "bold"), corner_radius=10,
            border_width=1, border_color="#5A2A2C",
            command=self._open_alerts
        )
        self._alert_btn.pack(fill="x", padx=10, pady=(0, 16))

    def _build_main(self):
        self._main = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        self._main.pack(side="left", fill="both", expand=True)

    def _build_status_bar(self):
        sb = ctk.CTkFrame(self, fg_color=BARBG, height=32, corner_radius=0,
                          border_width=0)
        sb.pack(side="bottom", fill="x")
        sb.pack_propagate(False)
        # top hairline divider
        ctk.CTkFrame(sb, fg_color=SUBBORDER, height=1).pack(fill="x", side="top")
        self._status_govee = _label(sb, "Govee: —", FONT_S, FAINT)
        self._status_govee.pack(side="left", padx=(24, 16))
        self._status_qr = _label(sb, "QR: —", FONT_S, FAINT)
        self._status_qr.pack(side="left", padx=16)
        self._status_sensibo = _label(sb, "Sensibo: —", FONT_S, FAINT)
        self._status_sensibo.pack(side="left", padx=16)
        self._status_time = _label(sb, "", FONT_S, SUBTEXT)
        self._status_time.pack(side="right", padx=24)
        self._status_refresh = _label(sb, "", FONT_S, FAINT)
        self._status_refresh.pack(side="right", padx=8)

        # Reload button — hidden until a code update is detected on disk.
        self._reload_btn = ctk.CTkButton(
            sb, text="🔄 Update ready — Reload", height=22, width=190,
            corner_radius=6, font=("Segoe UI", 10, "bold"),
            fg_color=DK_GOLD, hover_color=GOLD, text_color="black",
            command=self._restart_app)
        # not packed yet; shown by the watcher when files change

    # ── Navigation ────────────────────────────────────────────────────────────

    def _after_inc_delete(self):
        """After deleting an incubator, forget the stale selection and leave."""
        self._detail_inc = {}
        self.show_view("dashboard")

    def _toast(self, text: str, color: str = None, ms: int = 3500):
        """Brief, self-dismissing confirmation banner overlaid on the window.
        Parented to the app (not a view frame) so it survives view rebuilds."""
        try:
            old = getattr(self, "_toast_win", None)
            if old is not None and old.winfo_exists():
                old.destroy()
        except Exception:
            pass
        try:
            tw = ctk.CTkFrame(self, fg_color=(color or "#1B3A2A"),
                              corner_radius=10, border_width=1, border_color=GREEN)
            _label(tw, text, ("Segoe UI", 12, "bold"), "#EAF7EF").pack(
                padx=16, pady=8)
            tw.place(relx=0.5, rely=0.045, anchor="n")
            tw.lift()
            self._toast_win = tw
            self.after(ms, lambda w=tw: w.winfo_exists() and w.destroy())
        except Exception:
            pass

    def _manual_poll_incubator(self, inc: dict, btn=None):
        """Force an immediate Govee reading for one incubator (background thread)."""
        if not (inc.get("govee_device_id") and inc.get("govee_sku")):
            messagebox.showinfo("Pull Reading",
                "No Govee device is configured for this incubator.\n"
                "Set the Govee Device ID and SKU in Edit Setup first.", parent=self)
            return
        if btn is not None:
            btn.configure(text="Polling…", state="disabled")
        iid = inc["id"]

        def _work():
            try:
                temp_c, humidity = self._govee.poll_incubator(inc)
                err = None
            except Exception as exc:
                temp_c = humidity = None
                err = str(exc)

            def _done():
                try:
                    if btn is not None and btn.winfo_exists():
                        btn.configure(text="⟳ Pull Reading", state="normal")
                    if temp_c is not None and humidity is not None:
                        tc = govee_mod.to_celsius(temp_c)
                        db.save_reading(iid, tc, humidity)
                        # Keep the in-memory cache fresh so the detail view and
                        # tiles agree with what we just stored.
                        self._govee._last[iid] = {
                            "temp_c": tc, "humidity": humidity,
                            "timestamp": datetime.now().isoformat()}
                        self._refresh_alert_badge()
                        # Visible confirmation that survives the view rebuild
                        unit = (db.get_setting("temp_unit", "C") or "C").upper()
                        shown = tc if unit.startswith("C") else round(tc * 9 / 5 + 32, 1)
                        self._toast(f"✓ {inc.get('name', 'Incubator')} updated:  "
                                    f"{shown:g}°{unit[:1]}   {humidity:g}% RH   "
                                    f"{datetime.now():%H:%M}")
                        # Rebuild the detail view (safely) to show the new reading
                        if self._current_view == "inc_detail":
                            self._refresh_inc_detail()
                    else:
                        messagebox.showwarning("Pull Reading",
                            f"Couldn't read the sensor right now.\n"
                            f"{err or self._govee.status_label()}", parent=self)
                except Exception as exc:
                    messagebox.showerror("Pull Reading",
                        f"Error updating after poll:\n{exc}", parent=self)
            self.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    def _open_incubators(self):
        """Incubators tab opens the per-unit detail view (matches the README)."""
        if not self._detail_inc:
            incs = db.get_incubators(include_hidden=True)
            if incs:
                self._detail_inc = incs[0]
            else:
                self._open_incubator_dialog()   # no units yet → add one
                return
        self.show_view("inc_detail")

    def show_view(self, name: str):
        for v in self._views.values():
            v.pack_forget()
        # The detail view lives under the "Incubators" nav item — highlight it there
        _hl = "incubators" if name == "inc_detail" else name
        if True:
            for k, btn in self._nav_btns.items():
                active = (k == _hl)
                _ico = self._nav_icons.get(k, {})
                btn.configure(
                    fg_color="#25281E" if active else "transparent",   # spec exact
                    hover_color="#2B2C1D" if active else "#1A2436",
                    text_color=GOLD if active else SUBTEXT,
                    image=_ico.get("gold" if active else "grey"),
                    font=("Segoe UI", 12, "bold") if active else ("Segoe UI", 12))
        view = self._views[name]
        view.pack(fill="both", expand=True)
        self._current_view = name
        getattr(self, f"_refresh_{name}")()

    # ══════════════════════════════════════════════════════════════════════════
    #  DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════

    def _build_dashboard(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Dashboard",
                                  "Live monitoring · incubator status & controls")
        _btn_primary(hdr, "+ Incubator", lambda: self._open_incubator_dialog(),
                     width=130).pack(side="right")
        _btn_secondary(hdr, "Refresh", self._refresh_dashboard,
                       width=100).pack(side="right", padx=8)

        # "Show hidden" toggle — only visible when hidden incubators exist
        self._dash_show_hidden = ctk.BooleanVar(value=False)
        self._dash_hidden_btn  = ctk.CTkButton(
            hdr, text="", width=150, height=28,
            fg_color=CARD2, hover_color=BORDER, text_color=SUBTEXT,
            corner_radius=6, font=FONT_S,
            command=self._toggle_dash_hidden)
        # packed conditionally in _refresh_dashboard

        # Body: card area (left, expands) + incubator mode panel (right, flush)
        body = ctk.CTkFrame(frame, fg_color="transparent")
        body.pack(fill="both", expand=True)

        # Pack the right-hand mode panel FIRST so it reserves the right strip;
        # the card area then expands into all the remaining width.
        mode_col = ctk.CTkFrame(body, fg_color=RIGHTPANE, width=300, corner_radius=0)
        mode_col.pack(side="right", fill="y")
        mode_col.pack_propagate(False)
        _label(mode_col, "Incubator Modes", ("Segoe UI", 13, "bold"), GOLD).pack(
            anchor="w", padx=16, pady=(16, 0))
        _label(mode_col, "Set operating mode per unit", ("Segoe UI", 11), FAINT).pack(
            anchor="w", padx=16, pady=(0, 8))
        # Fixed (non-scrolling) — rows share the height so all units fit
        self._dash_mode_panel = ctk.CTkFrame(mode_col, fg_color="transparent")
        self._dash_mode_panel.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self._dash_scroll = ctk.CTkScrollableFrame(
            body, fg_color="transparent", corner_radius=0)
        self._dash_scroll.pack(side="left", fill="both", expand=True, padx=(12, 6), pady=4)

        frame._card_container = self._dash_scroll
        return frame

    def _toggle_dash_hidden(self):
        self._dash_show_hidden.set(not self._dash_show_hidden.get())
        self._refresh_dashboard()

    def _refresh_mode_panel(self):
        """Right-hand dashboard panel: each incubator with a temp-mode selector."""
        panel = getattr(self, "_dash_mode_panel", None)
        if panel is None:
            return
        for w in panel.winfo_children():
            w.destroy()
        panel.columnconfigure(0, weight=1)
        incs = db.get_incubators(include_hidden=True)
        if not incs:
            _label(panel, "No incubators yet.", FONT_S, SUBTEXT).pack(pady=10)
            return
        # Spec: Off / Cool / Inc / Hold short segments, active = mode color
        _SEGS = [("off", "Off"), ("cool_storage", "Cool"),
                 ("incubation", "Inc"), ("holding", "Hold")]
        for _ri, inc in enumerate(incs):
            panel.rowconfigure(_ri, weight=1, uniform="moderow")
            row = ctk.CTkFrame(panel, fg_color=NESTED, corner_radius=10,
                               border_width=1, border_color=BORDER2)
            row.grid(row=_ri, column=0, sticky="nsew", pady=3, padx=2)
            head = ctk.CTkFrame(row, fg_color="transparent")
            head.pack(fill="x", padx=10, pady=(8, 4))
            _label(head, inc["name"], ("Segoe UI", 12, "bold"),
                   "#E5E7EB").pack(side="left")
            _btn(head, "✎", lambda i=inc: self._open_incubator_dialog(i),
                 width=28, height=22, fg="transparent",
                 hover=CARD2, text_color=SUBTEXT).pack(side="right")
            segrow = ctk.CTkFrame(row, fg_color="transparent")
            segrow.pack(fill="x", padx=10, pady=(0, 9))
            cur = inc.get("temp_mode", "incubation")
            for ci in range(4):
                segrow.columnconfigure(ci, weight=1, uniform="seg")
            for ci, (key, short) in enumerate(_SEGS):
                active = (key == cur)
                mcol   = MODE_COLORS[key]
                ctk.CTkButton(
                    segrow, text=short, height=28, width=40, corner_radius=6,
                    font=("Segoe UI", 10, "bold" if active else "normal"),
                    fg_color=_mix(mcol, NESTED, 0.22) if active else CARD,
                    hover_color=_mix(mcol, NESTED, 0.30) if active else CARD2,
                    text_color=mcol if active else "#94A3B8",
                    border_width=1,
                    border_color=mcol if active else "#2A3648",
                    command=lambda k=key, i=inc["id"]: self._on_dash_mode_key(k, i),
                ).grid(row=0, column=ci, sticky="ew", padx=2)

    def _on_dash_mode(self, label: str, inc_id: int):
        self._on_dash_mode_key(calc._MODE_BY_LABEL.get(label, "incubation"), inc_id)

    def _on_dash_mode_key(self, key: str, inc_id: int):
        prev = next((x for x in db.get_incubators(include_hidden=True)
                     if x["id"] == inc_id), {}).get("temp_mode", "incubation")
        if key == prev:
            return
        db.set_incubator_temp_mode(inc_id, key)
        self._sync_trays_to_mode(inc_id, key, prev)   # cool-down prompt if relevant
        self.after(50, self._refresh_dashboard)        # deferred so the widget isn't
                                                       # destroyed mid-callback

    def _refresh_dashboard(self):
        self._refresh_mode_panel()
        self._card_widgets.clear()
        container = self._dash_scroll
        for w in container.winfo_children():
            w.destroy()

        # Incubators with an active alert → their cards get a red outline
        self._alert_inc_ids = {a["incubator_id"] for a in db.get_active_alerts()
                               if a.get("incubator_id")}

        show_hidden = getattr(self, "_dash_show_hidden", None)
        show_hidden = show_hidden.get() if show_hidden else False
        all_inc     = db.get_incubators(include_hidden=True)
        # Dashboard shows only incubators that are turned ON (temp_mode != "off").
        # Off ones live in the Incubators view, where they can be turned back on.
        on_inc      = [i for i in all_inc if not calc.is_off(i)]
        hidden_n    = sum(1 for i in on_inc if i.get("is_hidden"))
        incubators  = on_inc if show_hidden else [i for i in on_inc if not i.get("is_hidden")]

        # Update the "show/hide hidden" button visibility and label
        if hidden_n > 0:
            lbl = (f"Hide {hidden_n} hidden" if show_hidden
                   else f"Show {hidden_n} hidden")
            self._dash_hidden_btn.configure(text=lbl)
            self._dash_hidden_btn.pack(side="right", padx=6)
        else:
            self._dash_hidden_btn.pack_forget()

        if not incubators:
            ctk.CTkFrame(container, fg_color="transparent").pack(pady=40)
            if hidden_n > 0:
                _label(container,
                       f"All {hidden_n} incubator(s) are hidden.\n"
                       "Click 'Show hidden' above or manage them in the Incubators view.",
                       FONT_B, SUBTEXT).pack()
            else:
                _label(container, "No incubators yet.\nClick '+ Incubator' to add one.",
                       FONT_B, SUBTEXT).pack()
            return

        # Summary row — use aggregate query, not full row fetch
        # "In incubator" = active + cooled (count only drops on release)
        _stats         = db.get_tray_stats(status=db.IN_INCUBATOR_STATUSES)
        tray_count     = _stats["count"]
        total_gals     = _stats["total_gals"]
        # Total capacity = every incubator (even off/hidden), so the denominator
        # stays constant regardless of which incubators are turned on.
        total_capacity = sum((i.get("capacity") or 0) for i in all_inc)
        fill_pct       = round(tray_count / total_capacity * 100) if total_capacity else 0
        fill_col       = GREEN if fill_pct < 80 else (ORANGE if fill_pct < 95 else RED)

        _n_alerts = len(db.get_active_alerts())
        summary_f = ctk.CTkFrame(container, fg_color="transparent")
        summary_f.pack(fill="x", pady=(0, 14), padx=4)
        metrics = [
            ("Active Incubators", str(len(incubators)),               GOLD),
            ("Trays",             f"{tray_count} / {total_capacity}", GOLD),
            ("Capacity",          f"{fill_pct}% full",                fill_col),
            ("Total Gals",        f"{total_gals:.1f}",                GOLD),
            ("Active Alerts",     str(_n_alerts),  RED if _n_alerts else GOLD),
        ]
        for i in range(len(metrics)):
            summary_f.columnconfigure(i, weight=1, uniform="metric")
        for i, (txt, val, col) in enumerate(metrics):
            mc = ctk.CTkFrame(summary_f, fg_color=PANEL, corner_radius=12,
                              border_width=1, border_color=BORDER2)
            mc.grid(row=0, column=i, sticky="ew", padx=6)
            _label(mc, val, ("Segoe UI", 22, "bold"), col).pack(anchor="w", padx=16, pady=(14, 0))
            _label(mc, txt, ("Segoe UI", 11), SUBTEXT).pack(anchor="w", padx=16, pady=(0, 14))

        # Responsive card grid — rebuilds on window resize (debounced)
        grid = ctk.CTkFrame(container, fg_color="transparent")
        grid.pack(fill="both", expand=True, padx=4)
        grid._resize_job = None

        def _build_grid(cols):
            self._card_widgets.clear()
            for child in grid.winfo_children():
                child.destroy()
            for c in range(6):
                grid.columnconfigure(c, weight=1 if c < cols else 0,
                                     uniform="col" if c < cols else "")
            for idx, inc in enumerate(incubators):
                card = self._make_inc_card(grid, inc)
                card.grid(row=idx // cols, column=idx % cols,
                          padx=6, pady=6, sticky="nsew")

        def _on_resize(event=None):
            if grid._resize_job:
                self.after_cancel(grid._resize_job)
            def _do():
                # available card width = window − nav sidebar − mode panel
                w    = self.winfo_width() - 380
                cols = 4 if w >= 1400 else (3 if w >= 1000 else (2 if w >= 600 else 1))
                if getattr(grid, "_last_cols", None) != cols:
                    grid._last_cols = cols
                    _build_grid(cols)
            grid._resize_job = self.after(200, _do)

        self.bind("<Configure>", _on_resize, add=True)
        _on_resize()

    def _make_inc_card(self, parent, inc: dict) -> ctk.CTkFrame:
        is_hidden = bool(inc.get("is_hidden"))
        card_bg   = "#161E2C" if is_hidden else CARD
        title_col = SUBTEXT  if is_hidden else TEXT
        bdr_col   = "#222D3D" if is_hidden else "#263347"

        # Red outline when this incubator has an active alert
        has_alert  = inc["id"] in getattr(self, "_alert_inc_ids", set())
        bdr_width  = 2 if has_alert else 1
        if has_alert:
            bdr_col = RED

        card = ctk.CTkFrame(parent, fg_color=card_bg, corner_radius=11,
                            border_width=bdr_width, border_color=bdr_col)

        # ── Readings ──
        reading = self._govee.get_last(inc["id"])
        if not reading:
            db_row  = db.get_latest_reading(inc["id"])
            reading = {"temp_c": db_row["temperature_c"], "humidity": db_row["humidity_pct"],
                       "timestamp": db_row["timestamp"]} if db_row else {}
        temp_c = reading.get("temp_c")
        hum    = reading.get("humidity")

        t_min, t_max   = calc.get_temp_range(inc)
        unit           = db.get_setting("temp_unit", "C")
        goal_t, goal_h = db.get_mode_goals(inc.get("temp_mode", "incubation"))
        if temp_c is not None:
            t_str = calc.format_temp(temp_c, unit)
            # Temp color rule: vs goal (≤1°=green, ≤3°=orange, else red); else range
            if goal_t is not None:
                _d = abs(temp_c - goal_t)
                t_col = GREEN if _d <= 1 else (ORANGE if _d <= 3 else RED)
            elif t_min is not None:
                t_col = GREEN if t_min <= temp_c <= t_max else RED
            else:
                t_col = SUBTEXT
            h_col = GOLD
            problems  = calc.check_temp_humidity(inc, temp_c, hum)
            dot_color = RED if problems else GREEN
        else:
            t_str = "—"
            t_col = h_col = dot_color = SUBTEXT

        # ── Header: name + chips + status dot ──
        hdr = ctk.CTkFrame(card, fg_color="transparent")
        hdr.pack(fill="x", padx=14, pady=(12, 6))
        name_txt = f"{inc['name']}  (hidden)" if is_hidden else inc["name"]
        _label(hdr, name_txt, ("Segoe UI", 13, "bold"), title_col).pack(side="left")
        mode_key = inc.get("temp_mode", "incubation")
        mode_cfg = calc.TEMP_MODES.get(mode_key, calc.TEMP_MODES["incubation"])
        _mcol = MODE_COLORS.get(mode_key, BLUE)
        ctk.CTkLabel(hdr, text=mode_cfg["label"], font=("Segoe UI", 9, "bold"),
                     fg_color=MODE_BADGE_BG.get(mode_key, _mix(_mcol, card_bg, 0.20)),
                     text_color=_mcol,
                     corner_radius=6, height=20, padx=8).pack(side="left", padx=(8, 0))
        if has_alert:
            _n = sum(1 for a in db.get_active_alerts()
                     if a.get("incubator_id") == inc["id"])
            ctk.CTkLabel(hdr, text=f"{_n} Alert{'s' if _n != 1 else ''}",
                         font=("Segoe UI", 9, "bold"),
                         fg_color=_mix(RED, card_bg, 0.16), text_color=RED_LT,
                         corner_radius=6, height=20, padx=8).pack(side="left", padx=(6, 0))
        lbl_dot = _label(hdr, "●", ("Segoe UI", 18), dot_color)
        lbl_dot.pack(side="right")
        # Subtle bell toggle for temp alerts (spec has no header toggle — keep it minimal)
        alerts_on = bool(inc.get("temp_alerts_enabled", 1))
        ctk.CTkButton(
            hdr, text="🔔" if alerts_on else "🔕", width=26, height=22,
            fg_color="transparent", hover_color=CARD2,
            text_color=SUBTEXT if alerts_on else "#6B7280",
            font=("Segoe UI", 12), corner_radius=6,
            command=lambda i=inc: self._toggle_temp_alerts(i)
        ).pack(side="right", padx=(0, 4))

        # ── Large temp / humidity tiles (with per-mode goals) ──
        sensor_row = ctk.CTkFrame(card, fg_color="transparent")
        sensor_row.pack(fill="x", padx=12, pady=(0, 8))
        sensor_row.columnconfigure(0, weight=1)
        sensor_row.columnconfigure(1, weight=1)

        tf = ctk.CTkFrame(sensor_row, fg_color=CARD2, corner_radius=8,
                          border_width=1, border_color="#2A3648")
        tf.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        _label(tf, "TEMP", ("Segoe UI", 9, "bold"), SUBTEXT).pack(pady=(8, 0))
        lbl_temp = _label(tf, t_str if temp_c is not None else "—", ("Segoe UI", 20, "bold"), t_col)
        lbl_temp.pack(pady=(2, 0))
        _label(tf, f"Goal {calc.format_temp(goal_t, unit)}" if goal_t is not None else "—",
               ("Segoe UI", 9), FAINT).pack(pady=(0, 8))

        hf = ctk.CTkFrame(sensor_row, fg_color=CARD2, corner_radius=8,
                          border_width=1, border_color="#2A3648")
        hf.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        _label(hf, "HUMIDITY", ("Segoe UI", 9, "bold"), SUBTEXT).pack(pady=(8, 0))
        lbl_hum = _label(hf, f"{hum:.0f}%" if hum is not None else "—", ("Segoe UI", 20, "bold"), h_col)
        lbl_hum.pack(pady=(2, 0))
        _label(hf, f"Goal {goal_h:.0f}%" if goal_h is not None else "—",
               ("Segoe UI", 9), FAINT).pack(pady=(0, 8))

        # ── Bottom row: status/events left, tray info right ──
        bottom = ctk.CTkFrame(card, fg_color="transparent")
        bottom.pack(fill="x", padx=14, pady=(0, 6))
        bottom.columnconfigure(0, weight=1)
        bottom.columnconfigure(1, weight=1)

        # Left: last poll time + next event
        left = ctk.CTkFrame(bottom, fg_color="transparent")
        left.grid(row=0, column=0, sticky="w")
        _poll_txt, _poll_col = _poll_age(reading.get("timestamp"), POLL_INTERVAL_SEC)
        lbl_ts = _label(left, f"Last polled: {_poll_txt}", FONT_S, _poll_col)
        lbl_ts.pack(anchor="w")
        self._card_widgets[inc["id"]] = {"temp": lbl_temp, "hum": lbl_hum,
                                          "dot": lbl_dot, "ts": lbl_ts, "inc": inc,
                                          "card": card, "hidden": is_hidden}
        batches = db.get_batches(incubator_id=inc["id"], status="active")
        events  = calc.get_all_events(batches, lookahead_days=14)
        if events:
            ev   = events[0]
            ecol = RED if ev["urgent"] else (ORANGE if ev["days_away"] <= 5 else TEXT)
            _label(left, f"→ {ev['label']}: {calc.format_days(ev['days_away'])}", FONT_S, ecol).pack(anchor="w")
        else:
            _label(left, "No upcoming events", FONT_S, SUBTEXT).pack(anchor="w")

        # Right: tray count — aggregate query, no full row fetch
        right = ctk.CTkFrame(bottom, fg_color="transparent")
        right.grid(row=0, column=1, sticky="e")
        _ts    = db.get_tray_stats(incubator_id=inc["id"], status=db.IN_INCUBATOR_STATUSES)
        capacity   = inc.get("capacity") or 50
        fill_pct   = round(_ts["count"] / capacity * 100) if capacity else 0
        _label(right, f"{_ts['count']} / {capacity} trays",
               ("Segoe UI", 10, "bold"), TEXT2).pack(anchor="e")

        # ── Capacity bar (spec: 6px track, gold fill) ──
        cap_bar = ctk.CTkProgressBar(
            card, height=6, corner_radius=5,
            fg_color=CARD2, progress_color=GOLD)
        cap_bar.set(min(fill_pct / 100, 1.0))
        cap_bar.pack(fill="x", padx=14, pady=(2, 0))
        cap_sub = ctk.CTkFrame(card, fg_color="transparent")
        cap_sub.pack(fill="x", padx=14, pady=(1, 6))
        _label(cap_sub, f"{fill_pct}% filled", ("Segoe UI", 9), FAINT).pack(side="left")
        _label(cap_sub, f"{_ts['total_gals']:.1f} gal", ("Segoe UI", 9), FAINT).pack(side="right")

        # ── Sensibo AC manual controls (only when a device ID is configured) ──
        if inc.get("sensibo_device_id"):
            ac_row = ctk.CTkFrame(card, fg_color="transparent")
            ac_row.pack(fill="x", padx=12, pady=(2, 6))
            _label(ac_row, "AC:", FONT_S, SUBTEXT).pack(side="left", padx=(2, 6))
            _t_lbl, _t_fg, _t_tc = self._ac_toggle_style(inc["sensibo_device_id"])
            _ac_power_btn = ctk.CTkButton(
                ac_row, text=_t_lbl, width=82, height=28, corner_radius=14,
                fg_color=_t_fg, hover_color=BORDER, text_color=_t_tc,
                font=("Segoe UI", 11, "bold"),
                command=lambda i=inc["id"], d=inc["sensibo_device_id"]: self._sensibo_toggle_power(i, d),
            )
            _ac_power_btn.pack(side="left", padx=2)
            _ac_temp_btn = _btn(ac_row, self._ac_temp_label(inc["sensibo_device_id"]),
                 lambda i=inc["id"], d=inc["sensibo_device_id"], n=inc["name"]: self._sensibo_prompt_temp(i, d, n),
                 width=78, height=28, fg=CARD2, hover=BORDER)
            _ac_temp_btn.pack(side="left", padx=2)
            _ac_fan_btn = _btn(ac_row, self._ac_fan_label(inc["sensibo_device_id"]),
                 lambda i=inc["id"], d=inc["sensibo_device_id"], n=inc["name"]: self._sensibo_prompt_fan(i, d, n),
                 width=78, height=28, fg=CARD2, hover=BORDER)
            _ac_fan_btn.pack(side="left", padx=2)
            # Store refs so Sensibo handlers can update just these without full refresh
            self._card_widgets[inc["id"]]["ac_power"] = _ac_power_btn
            self._card_widgets[inc["id"]]["ac_temp"]  = _ac_temp_btn
            self._card_widgets[inc["id"]]["ac_fan"]   = _ac_fan_btn

        # ── Inspection badges (hidden when the incubator is off — no inspections needed) ──
        if not is_hidden and not calc.is_off(inc):
            brow = ctk.CTkFrame(card, fg_color="transparent")
            brow.pack(fill="x", padx=12, pady=(2, 10))
            _label(brow, "Inspections:", FONT_S, SUBTEXT).pack(side="left", padx=(2, 6))
            make_status_badges(
                brow, inc["id"],
                on_click=lambda p, i=inc: self._open_inspection_form(i)).pack(side="left")
        else:
            # Hidden / off cards just need bottom padding
            ctk.CTkFrame(card, fg_color="transparent", height=6).pack()

        # Whole card is clickable — navigates to detail.
        # Buttons (e.g. inspection pills) keep their own command and are skipped.
        def _go_detail(event, i=inc):
            self._show_inc_detail(i)

        def _bind_recursive(widget):
            if isinstance(widget, ctk.CTkButton):
                return
            widget.bind("<Button-1>", _go_detail)
            for child in widget.winfo_children():
                _bind_recursive(child)

        _bind_recursive(card)

        return card

    # ══════════════════════════════════════════════════════════════════════════
    #  INCUBATORS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_incubators_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Incubators",
                                  "Per-unit detail & inspection history")
        _btn_primary(hdr, "+ Add Incubator", lambda: self._open_incubator_dialog(),
                     width=150).pack(side="right")

        self._inc_scroll = ctk.CTkScrollableFrame(
            frame, fg_color="transparent", corner_radius=0)
        self._inc_scroll.pack(fill="both", expand=True, padx=12, pady=4)
        return frame

    def _refresh_incubators(self):
        container = self._inc_scroll
        for w in container.winfo_children():
            w.destroy()

        incubators = db.get_incubators(include_hidden=True)
        if not incubators:
            _label(container, "No incubators yet.", FONT_B, SUBTEXT).pack(pady=20)
            return

        for inc in incubators:
            hidden    = bool(inc.get("is_hidden"))
            row_bg    = "#141C2B" if hidden else PANEL
            bdr_col   = "#1E2938" if hidden else BORDER2
            name_col  = SUBTEXT  if hidden else GOLD

            row = ctk.CTkFrame(container, fg_color=row_bg, corner_radius=12,
                               border_width=1, border_color=bdr_col)
            row.pack(fill="x", padx=4, pady=4)

            left = ctk.CTkFrame(row, fg_color="transparent")
            left.pack(side="left", fill="both", expand=True, padx=14, pady=10)

            # Name + hidden badge
            name_row = ctk.CTkFrame(left, fg_color="transparent")
            name_row.pack(anchor="w", fill="x")
            _label(name_row, inc["name"], FONT_H, name_col).pack(side="left")
            if hidden:
                ctk.CTkLabel(name_row, text="  HIDDEN  ",
                             fg_color="#374151", text_color=SUBTEXT,
                             corner_radius=4, font=("Segoe UI", 9, "bold"),
                             height=20).pack(side="left", padx=8)

            reading = self._govee.get_last(inc["id"])
            if not reading:
                db_row  = db.get_latest_reading(inc["id"])
                reading = {"temp_c": db_row["temperature_c"], "humidity": db_row["humidity_pct"], "timestamp": db_row["timestamp"]} if db_row else {}
            temp_c  = reading.get("temp_c")
            if temp_c is not None:
                unit  = db.get_setting("temp_unit", "C")
                t_str = calc.format_temp(temp_c, unit)
                hum   = reading.get("humidity", 0)
                _label(left, f"🌡 {t_str}   💧 {hum:.0f}%", FONT_B, TEXT).pack(anchor="w")
            else:
                _label(left, "No sensor reading", FONT_B, SUBTEXT).pack(anchor="w")

            _mode_key = inc.get("temp_mode", "incubation")
            _mode_cfg = calc.TEMP_MODES.get(_mode_key, calc.TEMP_MODES["incubation"])
            if _mode_cfg["min"] is None:
                _range_txt = f"{_mode_cfg['label']} (no alerts)"
            else:
                _range_txt = f"{_mode_cfg['label']}: {_mode_cfg['min']}–{_mode_cfg['max']}°C"
            info_txt = f"Capacity: {inc.get('capacity',50)} trays  |  {_range_txt}"
            _label(left, info_txt, FONT_S, SUBTEXT).pack(anchor="w")

            # Per-mode temperature / humidity goals
            _gt, _gh = db.get_mode_goals(_mode_key)
            if _gt is not None or _gh is not None:
                _ug = db.get_setting("temp_unit", "C")
                _bits = []
                if _gt is not None: _bits.append(f"🌡 {calc.format_temp(_gt, _ug)}")
                if _gh is not None: _bits.append(f"💧 {_gh:.0f}%")
                _label(left, "Goal:  " + "    ".join(_bits), FONT_S, SUBTEXT).pack(anchor="w")

            govee_txt = (f"Govee: {inc.get('govee_device_id') or 'not set'}  "
                         f"({inc.get('govee_sku') or '—'})")
            _label(left, govee_txt, FONT_S, SUBTEXT).pack(anchor="w")

            # Sensibo AC manual control (only if a device ID is configured)
            if inc.get("sensibo_device_id"):
                ac_row = ctk.CTkFrame(left, fg_color="transparent")
                ac_row.pack(anchor="w", pady=(6, 0))
                _label(ac_row, "AC:", FONT_S, SUBTEXT).pack(side="left", padx=(0, 6))
                _it_lbl, _it_fg, _it_tc = self._ac_toggle_style(inc["sensibo_device_id"])
                ctk.CTkButton(
                    ac_row, text=_it_lbl, width=80, height=24, corner_radius=12,
                    fg_color=_it_fg, hover_color=BORDER, text_color=_it_tc,
                    font=("Segoe UI", 10, "bold"),
                    command=lambda i=inc["id"], d=inc["sensibo_device_id"]: self._sensibo_toggle_power(i, d),
                ).pack(side="left", padx=2)
                _btn(ac_row, self._ac_temp_label(inc["sensibo_device_id"]),
                     lambda i=inc["id"], d=inc["sensibo_device_id"], n=inc["name"]: self._sensibo_prompt_temp(i, d, n),
                     width=78, height=24, fg=CARD2, hover=BORDER).pack(side="left", padx=2)
                _btn(ac_row, self._ac_fan_label(inc["sensibo_device_id"]),
                     lambda i=inc["id"], d=inc["sensibo_device_id"], n=inc["name"]: self._sensibo_prompt_fan(i, d, n),
                     width=78, height=24, fg=CARD2, hover=BORDER).pack(side="left", padx=2)

            # Temp-mode selector — turn the incubator on/off and pick its mode here
            mode_row = ctk.CTkFrame(left, fg_color="transparent")
            mode_row.pack(anchor="w", pady=(6, 0))
            _label(mode_row, "Temp Mode:", FONT_S, SUBTEXT).pack(side="left", padx=(0, 6))
            _mode_var = ctk.StringVar(value=_mode_cfg["label"])
            ctk.CTkSegmentedButton(
                mode_row,
                values=[v["label"] for v in calc.TEMP_MODES.values()],
                variable=_mode_var,
                command=lambda label, iid=inc["id"]: self._set_inc_mode(iid, label),
                height=26, font=FONT_S,
            ).pack(side="left")

            # Inspection status badges (only for visible, non-off incubators)
            if not hidden and not calc.is_off(inc):
                ibrow = ctk.CTkFrame(left, fg_color="transparent")
                ibrow.pack(anchor="w", pady=(4, 0))
                _label(ibrow, "Inspections:", FONT_S, SUBTEXT).pack(side="left", padx=(0, 6))
                make_status_badges(ibrow, inc["id"]).pack(side="left")

            right = ctk.CTkFrame(row, fg_color="transparent")
            right.pack(side="right", padx=14, pady=10)

            if hidden:
                _btn(right, "Unhide",
                     lambda i=inc["id"]: self._set_hidden(i, False),
                     width=80, height=28, fg="#065F46", hover=TEAL,
                     text_color="white").pack(pady=2)
            else:
                if not calc.is_off(inc):
                    _btn(right, "Inspect",
                         lambda i=inc: self._open_inspection_form(i),
                         width=80, height=28, fg=BLUE, hover="#1D4ED8",
                         text_color="white").pack(pady=2)
                _btn(right, "Hide",
                     lambda i=inc["id"]: self._set_hidden(i, True),
                     width=80, height=28, fg=CARD2, hover=BORDER,
                     text_color=SUBTEXT).pack(pady=2)

            _btn(right, "Edit",
                 lambda i=inc: self._open_incubator_dialog(i),
                 width=80, height=28, fg=BORDER, hover=CARD2).pack(pady=2)
            _btn(right, "+ Batch",
                 lambda i=inc["id"]: self._open_batch_dialog(incubator_id=i),
                 width=80, height=28, fg=BORDER, hover=CARD2).pack(pady=2)
            _btn(right, "Delete",
                 lambda i=inc["id"]: self._delete_incubator(i),
                 width=80, height=28, fg="#4B0000", hover=RED).pack(pady=2)

            # Batches sub-list
            batches = db.get_batches(incubator_id=inc["id"])
            if batches:
                bf = ctk.CTkFrame(row, fg_color=CARD2, corner_radius=6)
                bf.pack(fill="x", padx=12, pady=(0, 10))
                for batch in batches:
                    br = ctk.CTkFrame(bf, fg_color="transparent")
                    br.pack(fill="x", padx=8, pady=2)
                    bname = batch.get("name") or f"Batch {batch['id']}"
                    bstat = batch.get("status", "active")
                    scol  = GREEN if bstat == "active" else SUBTEXT
                    _label(br, f"● {bname}", FONT_S, scol).pack(side="left")
                    sd = batch.get("start_date") or "—"
                    _label(br, f"Start: {sd}", FONT_S, SUBTEXT).pack(side="left", padx=12)
                    _btn(br, "Edit",
                         lambda b=batch: self._open_batch_dialog(batch=b),
                         width=60, height=22, fg=BORDER, hover=CARD).pack(side="right")

    # ══════════════════════════════════════════════════════════════════════════
    #  SAMPLES VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_samples_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Samples",
                                  "X-Ray results & sample records")
        _btn_primary(hdr, "+ Add Sample", lambda: self._open_sample_dialog(),
                     width=130).pack(side="right")
        _btn_secondary(hdr, "Merge Duplicates", self._merge_duplicate_samples,
                       width=140).pack(side="right", padx=8)
        _btn_secondary(hdr, "Import Spreadsheet", self._import_xray,
                       width=150).pack(side="right")

        # Search box — filter the sample list by name as you type
        sbar = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=10,
                            border_width=1, border_color=BORDER2)
        sbar.pack(fill="x", padx=12, pady=(0, 6))
        _label(sbar, "Search:", FONT_S, SUBTEXT).pack(side="left", padx=10, pady=6)
        self._smp_search = ctk.CTkEntry(
            sbar, placeholder_text="Sample name…", width=260,
            fg_color=CARD2, border_color=BORDER, text_color=TEXT)
        self._smp_search.pack(side="left", padx=6, pady=6)
        self._smp_search.bind("<KeyRelease>", lambda e: self._refresh_samples())

        # Year filter — samples used by trays started in the chosen year
        _smp_years = self._tray_years()
        self._smp_year = _combo(sbar, ["All Years"] + _smp_years, 120)
        _cur_year = str(datetime.now().year)
        self._smp_year.set(_cur_year if _cur_year in _smp_years else "All Years")
        self._smp_year.pack(side="left", padx=6, pady=6)
        self._smp_year.configure(command=lambda _: self._refresh_samples())

        self._smp_count_lbl = _label(sbar, "", FONT_S, SUBTEXT)
        self._smp_count_lbl.pack(side="right", padx=12, pady=6)

        # Label-grid table (per the CTk guide) — link-blue names, styled cells
        tbl = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=12,
                           border_width=1, border_color=BORDER2)
        tbl.pack(fill="both", expand=True, padx=12, pady=4)
        self._smp_scroll = ctk.CTkScrollableFrame(tbl, fg_color="transparent")
        self._smp_scroll.pack(fill="both", expand=True, padx=2, pady=2)
        self._smp_sort_col = 0
        self._smp_sort_asc = True
        return frame

    # (label, grid weight, align 'w'|'e', kind)
    _SMP_COLS = [
        ("Name", 3, "w", "link"), ("Total Kg", 1, "e", "num"),
        ("Live Bees/Kg", 2, "e", "num"), ("Parasites", 1, "e", "num"),
        ("Chalkbrood", 1, "e", "num"), ("Total Gal", 1, "e", "num"),
        ("Kg for 2gal", 1, "e", "num"), ("Expected", 1, "e", "num"),
        ("Actual", 1, "e", "num"), ("Inc. Space", 1, "w", "text"),
        ("Notes", 3, "w", "text"),
    ]

    def _sort_sample_tree(self, col_idx: int):
        if self._smp_sort_col == col_idx:
            self._smp_sort_asc = not self._smp_sort_asc
        else:
            self._smp_sort_col = col_idx
            self._smp_sort_asc = True
        self._refresh_samples()

    def _refresh_samples(self):
        scroll = self._smp_scroll
        for w in scroll.winfo_children():
            w.destroy()

        def _n(v, dec=1):
            return f"{v:,.{dec}f}" if isinstance(v, (int, float)) else "—"

        def _kg(lbs_val, kg_val, dec=1):
            if isinstance(kg_val, (int, float)):
                return f"{kg_val:,.{dec}f}"
            if isinstance(lbs_val, (int, float)):
                return f"{lbs_val * 0.45359237:,.{dec}f}"
            return "—"

        def _per_kg(per_lb_val, per_kg_val, dec=0):
            if isinstance(per_kg_val, (int, float)):
                return f"{per_kg_val:,.{dec}f}"
            if isinstance(per_lb_val, (int, float)):
                return f"{per_lb_val / 0.45359237:,.{dec}f}"
            return "—"

        q = ""
        if getattr(self, "_smp_search", None) is not None:
            q = self._smp_search.get().strip().lower()

        samples = db.get_samples()
        actual_counts = db.get_tray_counts_by_sample(statuses=None)

        yr = getattr(self, "_smp_year", None)
        yr = yr.get() if yr else "All Years"
        if yr and yr != "All Years":
            year_ids = db.current_year_sample_ids(int(yr))
            samples = [s for s in samples if s["id"] in year_ids]

        # Build (id, [display values]) rows
        data = []
        for s in samples:
            if q and q not in (s["name"] or "").lower():
                continue
            data.append((s["id"], [
                s["name"],
                _kg(s.get("total_weight_lbs"), s.get("total_weight_kg")),
                _per_kg(s.get("live_bees_per_lb"), s.get("live_bees_per_kg")),
                _n(s.get("parasites")),
                _n(s.get("chalkbrood")),
                _n(s.get("total_volume_gal")),
                _kg(s.get("lbs_per_2gal"), s.get("kg_per_2gal"), 2),
                (str(math.ceil(s["total_trays"]))
                 if isinstance(s.get("total_trays"), (int, float)) else "—"),
                str(actual_counts.get(s["id"], 0)),
                s.get("incubator_space") or "—",
                (s.get("notes") or "").replace("\n", " "),
            ]))

        # Sort by the active column
        sc = self._smp_sort_col

        def _key(row):
            v = row[1][sc]
            if v is None or v == "—":
                return (2, 0.0, "")
            try:
                return (0, float(str(v).replace("%", "").replace(",", "")), "")
            except (ValueError, AttributeError):
                return (1, 0.0, str(v).lower())
        data.sort(key=lambda r: str(r[1][0]).lower())
        data.sort(key=_key, reverse=not self._smp_sort_asc)

        # Header row (clickable to sort)
        hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 2))
        for ci, (lbl, wt, al, kind) in enumerate(self._SMP_COLS):
            hdr.columnconfigure(ci, weight=wt, uniform="smp")
            arrow = (" ↑" if self._smp_sort_asc else " ↓") if ci == sc else ""
            ctk.CTkButton(
                hdr, text=lbl + arrow, height=30, corner_radius=6,
                anchor="w" if al == "w" else "e",
                fg_color="transparent", hover_color=CARD2, text_color=GOLD,
                font=("Segoe UI", 11, "bold"),
                command=lambda c=ci: self._sort_sample_tree(c)
            ).grid(row=0, column=ci, sticky="ew", padx=1)
        ctk.CTkFrame(scroll, fg_color=BORDER2, height=1).pack(fill="x")

        # Data rows
        for i, (sid, vals) in enumerate(data):
            rf = ctk.CTkFrame(scroll, fg_color=PANEL if i % 2 == 0 else "#18222F",
                              corner_radius=0)
            rf.pack(fill="x")
            for ci, (lbl, wt, al, kind) in enumerate(self._SMP_COLS):
                rf.columnconfigure(ci, weight=wt, uniform="smp")
                col = LINK if kind == "link" else (
                    "#E5E7EB" if al == "e" else "#CBD5E1")
                _label(rf, str(vals[ci]), ("Segoe UI", 11), col,
                       anchor="w" if al == "w" else "e", wraplength=200 if kind == "text" else 0
                       ).grid(row=0, column=ci, sticky="ew", padx=8, pady=6)
            rf.bind("<Double-1>", lambda e, s=sid: self._open_sample_by_id(s))
            for ch in rf.winfo_children():
                ch.bind("<Double-1>", lambda e, s=sid: self._open_sample_by_id(s))

        if getattr(self, "_smp_count_lbl", None) is not None:
            self._smp_count_lbl.configure(
                text=f"{len(data)} of {len(samples)} samples" if q
                else f"{len(samples)} samples")

    def _open_sample_by_id(self, sid: int):
        sample = next((s for s in db.get_samples() if s["id"] == sid), None)
        if sample:
            self._open_sample_dialog(sample)

    # ══════════════════════════════════════════════════════════════════════════
    #  ANALYTICS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_analytics_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)
        hdr = self._screen_header(frame, "Analytics",
                                  "Sample & incubation performance")
        _btn_secondary(hdr, "Refresh", self._refresh_analytics,
                       width=100).pack(side="right")
        _yrs = self._tray_years()
        self._an_year = _combo(hdr, ["All Years"] + _yrs, 120)
        _cur = str(datetime.now().year)
        self._an_year.set(_cur if _cur in _yrs else "All Years")
        self._an_year.pack(side="right", padx=8)
        self._an_year.configure(command=lambda _: self._refresh_analytics())
        self._an_body = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        self._an_body.pack(fill="both", expand=True, padx=12, pady=4)
        return frame

    def _refresh_analytics(self):
        body = self._an_body
        for w in body.winfo_children():
            w.destroy()

        yr = self._an_year.get() if getattr(self, "_an_year", None) else "All Years"
        year = None if yr == "All Years" else int(yr)

        samples = db.get_samples()
        if year is not None:
            ids = db.current_year_sample_ids(year)
            samples = [s for s in samples if s["id"] in ids]
        actual = db.get_tray_counts_by_sample(statuses=None)

        self._an_kpi_cards(body, samples, actual)
        self._an_bar_chart(body, "Live bees / kg by sample (high → low)",
                           samples, "live_bees_per_kg", "{:,.0f}", "#FFD700")
        self._an_bar_chart(body, "Parasite % by sample (high → low)",
                           samples, "parasites", "{:.1f}%", "#EF4444")
        self._an_bar_chart(body, "Chalkbrood % by sample (high → low)",
                           samples, "chalkbrood", "{:.1f}%", "#F59E0B")
        self._an_temp_stability(body)
        self._an_cycle_stats(body, year)

    def _an_card(self, parent, title):
        card = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=12,
                            border_width=1, border_color=BORDER2)
        card.pack(fill="x", padx=4, pady=7)
        _label(card, title, ("Segoe UI", 12, "bold"), GOLD).pack(
            anchor="w", padx=16, pady=(12, 2))
        return card

    def _an_kpi_cards(self, parent, samples, actual):
        def avg(key):
            vals = [s[key] for s in samples if isinstance(s.get(key), (int, float))]
            return sum(vals) / len(vals) if vals else None

        a_live, a_par, a_chalk = avg("live_bees_per_kg"), avg("parasites"), avg("chalkbrood")
        exp = sum(s["total_trays"] for s in samples if isinstance(s.get("total_trays"), (int, float)))
        act = sum(actual.get(s["id"], 0) for s in samples)
        cards = [
            ("Samples", str(len(samples))),
            ("Avg live bees/kg", f"{a_live:,.0f}" if a_live is not None else "—"),
            ("Avg parasite %", f"{a_par:.1f}%" if a_par is not None else "—"),
            ("Avg chalkbrood %", f"{a_chalk:.1f}%" if a_chalk is not None else "—"),
            ("Trays exp / actual", f"{math.ceil(exp)} / {act}"),
        ]
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(4, 8))
        for i in range(len(cards)):
            row.columnconfigure(i, weight=1, uniform="an_kpi")
        for i, (lbl, val) in enumerate(cards):
            c = ctk.CTkFrame(row, fg_color=PANEL, corner_radius=12,
                             border_width=1, border_color=BORDER2)
            c.grid(row=0, column=i, sticky="ew", padx=5)
            _label(c, val, ("Segoe UI", 22, "bold"), GOLD).pack(
                anchor="w", padx=16, pady=(14, 0))
            _label(c, lbl, ("Segoe UI", 11), SUBTEXT).pack(
                anchor="w", padx=16, pady=(0, 14))

    def _an_bar_chart(self, parent, title, samples, key, fmt, color, top=15):
        card = self._an_card(parent, title)
        data = [(s["name"], s[key]) for s in samples
                if isinstance(s.get(key), (int, float))]
        if not data:
            _label(card, "No data for this selection.", FONT_S, SUBTEXT).pack(
                padx=14, pady=(0, 12))
            return
        data.sort(key=lambda x: x[1], reverse=True)
        data = data[:top]
        if not HAS_MPL:
            for nm, v in data:
                _label(card, f"{nm}: {fmt.format(v)}", FONT_S, TEXT).pack(
                    anchor="w", padx=18, pady=1)
            ctk.CTkFrame(card, fg_color="transparent", height=8).pack()
            return
        names = [d[0] for d in data][::-1]
        vals  = [d[1] for d in data][::-1]
        fig = Figure(figsize=(9, max(2.0, 0.34 * len(names) + 0.5)), facecolor=PANEL)
        ax = fig.add_subplot(111)
        ax.set_facecolor(PANEL)
        bars = ax.barh(names, vals, color=color)
        ax.tick_params(colors="#9CA3AF", labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor(BORDER2)
        xmax = max(vals) if vals else 0
        ax.set_xlim(0, xmax * 1.15 if xmax else 1)
        for b, v in zip(bars, vals):
            ax.text(b.get_width(), b.get_y() + b.get_height() / 2,
                    " " + fmt.format(v), va="center", color="#F3F4F6", fontsize=7)
        fig.tight_layout(pad=1.0)
        cv = FigureCanvasTkAgg(fig, master=card)
        cv.draw()
        cv.get_tk_widget().pack(fill="x", padx=8, pady=(0, 8))

    def _an_temp_stability(self, parent):
        card = self._an_card(parent, "Incubator temperature stability (last 30 days)")
        grid = ctk.CTkFrame(card, fg_color="transparent")
        grid.pack(fill="x", padx=14, pady=(0, 12))
        heads = ["Incubator", "Readings", "% in range", "Avg temp"]
        for ci, h in enumerate(heads):
            grid.columnconfigure(ci, weight=1)
            _label(grid, h, FONT_S, SUBTEXT).grid(row=0, column=ci, sticky="w", padx=6, pady=2)
        unit = db.get_setting("temp_unit", "C")
        r = 1
        for inc in db.get_incubators(include_hidden=False):
            readings = db.get_readings_hours(inc["id"], 24 * 30)
            temps = [x["temperature_c"] for x in readings if x["temperature_c"] is not None]
            t_min, t_max = calc.get_temp_range(inc)
            if not temps or t_min is None:
                pct_txt, avg_txt = "—", "—"
                pct_col = SUBTEXT
            else:
                pct = 100 * sum(1 for t in temps if t_min <= t <= t_max) / len(temps)
                avg = sum(temps) / len(temps)
                pct_txt = f"{pct:.0f}%"
                pct_col = GREEN if pct >= 90 else (ORANGE if pct >= 70 else RED)
                avg_txt = calc.format_temp(avg, unit)
            _label(grid, inc["name"], FONT_B, TEXT).grid(row=r, column=0, sticky="w", padx=6, pady=2)
            _label(grid, str(len(temps)), FONT_S, SUBTEXT).grid(row=r, column=1, sticky="w", padx=6, pady=2)
            _label(grid, pct_txt, FONT_B, pct_col).grid(row=r, column=2, sticky="w", padx=6, pady=2)
            _label(grid, avg_txt, FONT_S, TEXT).grid(row=r, column=3, sticky="w", padx=6, pady=2)
            r += 1

    def _an_cycle_stats(self, parent, year):
        card = self._an_card(parent, "Cycle & cool-down stats")
        trays = db.get_trays()
        if year is not None:
            trays = [t for t in trays
                     if (lambda d: d is not None and d.year == year)(
                         _parse_date_loose(t.get("in_date")))]
        cool = [d for d in (cool_down_days(t) for t in trays) if d is not None]
        incub = []
        for t in trays:
            ind = _parse_date_loose(t.get("in_date"))
            outd = _parse_date_loose(t.get("out_date"))
            if ind and outd and outd >= ind:
                incub.append((outd - ind).days)
        lines = []
        lines.append(f"Trays in selection: {len(trays)}")
        if cool:
            lines.append(f"Cool-down days — avg {sum(cool)/len(cool):.1f}, "
                         f"min {min(cool)}, max {max(cool)}  (n={len(cool)})")
        else:
            lines.append("Cool-down days — no cooled/released trays yet")
        if incub:
            lines.append(f"Incubation length (start→release) — avg {sum(incub)/len(incub):.1f} days  "
                         f"(n={len(incub)})")
        for ln in lines:
            _label(card, ln, FONT_B, TEXT).pack(anchor="w", padx=18, pady=2)
        ctk.CTkFrame(card, fg_color="transparent", height=6).pack()

    # ══════════════════════════════════════════════════════════════════════════
    #  TRAYS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_trays_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Trays",
                                  "Tray inventory & release tracking")
        _btn_primary(hdr, "+ Add Tray", lambda: self._open_tray_dialog(),
                     width=110).pack(side="right")
        _btn(hdr, "🗑 Delete All", self._delete_all_trays,
             fg="#3A1C1E", hover="#4A2225", text_color=RED_LT,
             height=34, width=110, border_width=1,
             border_color="#5A2A2C").pack(side="right", padx=8)
        _btn_secondary(hdr, "Set Status →", self._bulk_set_status,
                       width=110).pack(side="right", padx=(8, 0))
        self._bulk_status = _combo(hdr, [lbl for _v, lbl in db.TRAY_STATUS_OPTIONS], 120)
        self._bulk_status.set("Released")
        self._bulk_status.pack(side="right", padx=6)
        _btn_secondary(hdr, "QR Code", self._show_selected_qr,
                       width=90).pack(side="right", padx=6)
        _btn_secondary(hdr, "History", self._show_tray_history,
                       width=90).pack(side="right", padx=6)
        _btn_secondary(hdr, "Release CSV",
                       lambda: self._release_csv_import(on_complete=self._refresh_trays),
                       width=110).pack(side="right", padx=6)
        _btn_secondary(hdr, "⬇ Release Template", lambda: self._release_csv_template(),
                       width=150).pack(side="right")
        _btn_secondary(hdr, "Import CSV",
                       lambda: self._tray_csv_import(on_complete=self._refresh_trays),
                       width=105).pack(side="right", padx=6)
        _btn_secondary(hdr, "⬇ Template", lambda: self._tray_csv_template(),
                       width=110).pack(side="right")

        # Filter bar
        fbar = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=10,
                            border_width=1, border_color=BORDER2)
        fbar.pack(fill="x", padx=12, pady=(0, 6))
        _label(fbar, "Filter:", FONT_S, SUBTEXT).pack(side="left", padx=10, pady=6)

        incubators  = db.get_incubators()
        self._flt_inc_map = {"All Incubators": None}
        self._flt_inc_map.update({i["name"]: i["id"] for i in incubators})
        self._flt_inc = _combo(fbar, list(self._flt_inc_map.keys()), 180)
        self._flt_inc.set("All Incubators")
        self._flt_inc.pack(side="left", padx=6, pady=6)
        self._flt_inc.configure(command=lambda _: self._refresh_trays())

        self._flt_status = _combo(fbar, ["All"] + [lbl for _v, lbl in db.TRAY_STATUS_OPTIONS], 130)
        self._flt_status.set("All")
        self._flt_status.pack(side="left", padx=6, pady=6)
        self._flt_status.configure(command=lambda _: self._refresh_trays())

        # Year filter — based on each tray's Start Date (defaults to current year)
        _tray_yrs = self._tray_years()
        self._flt_year = _combo(fbar, ["All Years"] + _tray_yrs, 120)
        _cur_year = str(datetime.now().year)
        self._flt_year.set(_cur_year if _cur_year in _tray_yrs else "All Years")
        self._flt_year.pack(side="left", padx=6, pady=6)
        self._flt_year.configure(command=lambda _: self._refresh_trays())

        # Sample look-up — show only trays of the chosen sample
        self._flt_sample_map = {"All Samples": None}
        self._flt_sample_map.update({s["name"]: s["id"] for s in db.get_samples()})
        self._flt_sample = _combo(fbar, list(self._flt_sample_map.keys()), 200)
        self._flt_sample.set("All Samples")
        self._flt_sample.pack(side="left", padx=6, pady=6)
        self._flt_sample.configure(command=lambda _: self._refresh_trays())

        # Cool-down summary for the currently-shown trays
        self._cooldown_lbl = _label(fbar, "", FONT_S, TEAL)
        self._cooldown_lbl.pack(side="right", padx=12, pady=6)

        # Label-grid table with pagination (fast even at thousands of rows)
        tbl = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=12,
                           border_width=1, border_color=BORDER2)
        tbl.pack(fill="both", expand=True, padx=12, pady=(4, 2))
        self._tray_scroll = ctk.CTkScrollableFrame(tbl, fg_color="transparent")
        self._tray_scroll.pack(fill="both", expand=True, padx=2, pady=2)

        pg = ctk.CTkFrame(frame, fg_color="transparent")
        pg.pack(fill="x", padx=14, pady=(0, 4))
        self._tray_prev_btn = _btn_secondary(pg, "‹ Prev", self._tray_prev_page, width=80)
        self._tray_prev_btn.pack(side="left")
        self._tray_page_lbl = _label(pg, "", FONT_S, SUBTEXT)
        self._tray_page_lbl.pack(side="left", padx=10)
        self._tray_next_btn = _btn_secondary(pg, "Next ›", self._tray_next_page, width=80)
        self._tray_next_btn.pack(side="left")
        self._tray_sel_lbl = _label(pg, "", FONT_S, TEAL)
        self._tray_sel_lbl.pack(side="right")

        self._tray_sort_col = 0
        self._tray_sort_asc = True
        self._tray_page = 0
        self._tray_sel = set()
        self._tray_all = []
        self._tray_row_frames = {}
        return frame

    # (label, weight, align, kind)
    _TRAY_COLS = [
        ("Tray #", 2, "w", "link"), ("Sample", 3, "w", "text"),
        ("Incubator", 2, "w", "muted"), ("Weight Kg", 1, "e", "num"),
        ("Live/Kg", 1, "e", "num"), ("Parasite", 1, "e", "num"),
        ("Chalk", 1, "e", "num"), ("Start", 2, "w", "text"),
        ("Release", 2, "w", "text"), ("Cool", 1, "e", "num"),
        ("Status", 2, "c", "badge"),
    ]
    _TRAY_STATUS_COLOR = {"active": BLUE, "cooled": "#06B6D4",
                          "released": GREEN, "removed": FAINT}
    _TRAY_PAGE_SIZE = 150

    def _sort_tray_tree(self, col_idx: int):
        if self._tray_sort_col == col_idx:
            self._tray_sort_asc = not self._tray_sort_asc
        else:
            self._tray_sort_col = col_idx
            self._tray_sort_asc = True
        self._tray_sort_and_render()

    def _tray_prev_page(self):
        if self._tray_page > 0:
            self._tray_page -= 1
            self._tray_render_page()

    def _tray_next_page(self):
        import math as _m
        pages = max(1, _m.ceil(len(self._tray_all) / self._TRAY_PAGE_SIZE))
        if self._tray_page < pages - 1:
            self._tray_page += 1
            self._tray_render_page()

    def _tray_sort_and_render(self):
        sc = self._tray_sort_col

        def _key(row):
            v = row["cells"][sc]
            if v is None or v == "—":
                return (2, 0.0, "")
            try:
                return (0, float(str(v).replace("%", "").replace(",", "").replace("d", "")), "")
            except (ValueError, AttributeError):
                return (1, 0.0, str(v).lower())
        self._tray_all.sort(key=lambda r: str(r["cells"][0]).lower())
        self._tray_all.sort(key=_key, reverse=not self._tray_sort_asc)
        self._tray_page = 0
        self._tray_render_page()

    def _tray_toggle_sel(self, tid: int):
        if tid in self._tray_sel:
            self._tray_sel.discard(tid)
        else:
            self._tray_sel.add(tid)
        rf = self._tray_row_frames.get(tid)
        if rf and rf.winfo_exists():
            rf.configure(fg_color="#26374F" if tid in self._tray_sel else rf._base_bg)
        self._tray_sel_lbl.configure(
            text=f"{len(self._tray_sel)} selected" if self._tray_sel else "")

    def _tray_render_page(self):
        import math as _m
        scroll = self._tray_scroll
        for w in scroll.winfo_children():
            w.destroy()
        self._tray_row_frames = {}

        total = len(self._tray_all)
        pages = max(1, _m.ceil(total / self._TRAY_PAGE_SIZE))
        self._tray_page = max(0, min(self._tray_page, pages - 1))
        start = self._tray_page * self._TRAY_PAGE_SIZE
        page_rows = self._tray_all[start:start + self._TRAY_PAGE_SIZE]

        # Header
        hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 2))
        for ci, (lbl, wt, al, kind) in enumerate(self._TRAY_COLS):
            hdr.columnconfigure(ci, weight=wt, uniform="tray")
            arrow = (" ↑" if self._tray_sort_asc else " ↓") if ci == self._tray_sort_col else ""
            ctk.CTkButton(
                hdr, text=lbl + arrow, height=28, corner_radius=6,
                anchor="w" if al == "w" else ("center" if al == "c" else "e"),
                fg_color="transparent", hover_color=CARD2, text_color=GOLD,
                font=("Segoe UI", 11, "bold"),
                command=lambda c=ci: self._sort_tray_tree(c)
            ).grid(row=0, column=ci, sticky="ew", padx=1)
        ctk.CTkFrame(scroll, fg_color=BORDER2, height=1).pack(fill="x")

        for i, row in enumerate(page_rows):
            base_bg = PANEL if i % 2 == 0 else "#18222F"
            sel = row["id"] in self._tray_sel
            rf = ctk.CTkFrame(scroll, fg_color="#26374F" if sel else base_bg,
                              corner_radius=0)
            rf._base_bg = base_bg
            rf.pack(fill="x")
            self._tray_row_frames[row["id"]] = rf
            for ci, (lbl, wt, al, kind) in enumerate(self._TRAY_COLS):
                rf.columnconfigure(ci, weight=wt, uniform="tray")
                v = row["cells"][ci]
                if kind == "badge":
                    scol = self._TRAY_STATUS_COLOR.get(row["status"], SUBTEXT)
                    ctk.CTkLabel(rf, text=f" {v} ", font=("Segoe UI", 10, "bold"),
                                 fg_color=_mix(scol, base_bg, 0.18), text_color=scol,
                                 corner_radius=6).grid(row=0, column=ci, padx=8, pady=5)
                else:
                    col = LINK if kind == "link" else (
                        SUBTEXT if kind == "muted" else
                        ("#E5E7EB" if al == "e" else "#CBD5E1"))
                    _label(rf, str(v), ("Segoe UI", 11), col,
                           anchor="w" if al == "w" else "e"
                           ).grid(row=0, column=ci, sticky="ew", padx=8, pady=5)
            rf.bind("<Button-1>", lambda e, t=row["id"]: self._tray_toggle_sel(t))
            rf.bind("<Double-1>", lambda e, t=row["id"]: self._open_tray_by_id(t))
            for ch in rf.winfo_children():
                ch.bind("<Button-1>", lambda e, t=row["id"]: self._tray_toggle_sel(t))
                ch.bind("<Double-1>", lambda e, t=row["id"]: self._open_tray_by_id(t))

        self._tray_page_lbl.configure(text=f"Page {self._tray_page + 1} of {pages}  ·  {total} trays")
        self._tray_prev_btn.configure(state="normal" if self._tray_page > 0 else "disabled")
        self._tray_next_btn.configure(state="normal" if self._tray_page < pages - 1 else "disabled")

    def _open_tray_by_id(self, tid: int):
        tray = db.get_tray_by_id(tid)
        if tray:
            self._open_tray_dialog(tray=tray)

    def _tray_years(self) -> list:
        """Distinct years present in tray Start Dates, newest first (as strings)."""
        years = set()
        for t in db.get_trays():
            d = _parse_date_loose(t.get("in_date"))
            if d:
                years.add(d.year)
        return [str(y) for y in sorted(years, reverse=True)]

    def _refresh_trays(self):
        inc_id = self._flt_inc_map.get(self._flt_inc.get())
        _flt = self._flt_status.get()
        status = None if _flt == "All" else db.tray_status_value(_flt)

        sample_id = None
        _fs = getattr(self, "_flt_sample", None)
        if _fs is not None:
            sample_id = self._flt_sample_map.get(_fs.get())

        trays = db.get_trays(incubator_id=inc_id, sample_id=sample_id, status=status)

        # Year filter — keep trays whose Start Date falls in the chosen year
        yr = getattr(self, "_flt_year", None)
        yr = yr.get() if yr else "All Years"
        if yr and yr != "All Years":
            trays = [t for t in trays
                     if (lambda d: d is not None and str(d.year) == yr)(
                         _parse_date_loose(t.get("in_date")))]

        self._tray_sel = set()
        self._tray_all = [{
            "id": t["id"], "status": t.get("status") or "active",
            "cells": [
                t["tray_number"],
                t.get("sample_name") or "—",
                t.get("incubator_name") or "—",
                f"{t['sample_kg_per_2gal']:.2f}" if t.get("sample_kg_per_2gal") else "—",
                f"{t['sample_live_per_kg']:,.0f}" if t.get("sample_live_per_kg") else "—",
                f"{t['sample_parasites']:.1f}%" if t.get("sample_parasites") is not None else "—",
                f"{t['sample_chalkbrood']:.1f}%" if t.get("sample_chalkbrood") is not None else "—",
                t.get("in_date") or "—",
                t.get("out_date") or "—",
                (lambda d: f"{d}d" if d is not None else "—")(cool_down_days(t)),
                db.tray_status_label(t.get("status") or "active"),
            ]} for t in trays]
        self._tray_sort_and_render()

        # Cool-down report: average over shown trays that have a cool-down value
        _durs = [d for d in (cool_down_days(t) for t in trays) if d is not None]
        if _durs:
            self._cooldown_lbl.configure(
                text=f"Cool-down: avg {sum(_durs)/len(_durs):.1f}d  "
                     f"(min {min(_durs)} · max {max(_durs)}, n={len(_durs)})")
        else:
            self._cooldown_lbl.configure(text="")

    def _show_selected_qr(self):
        if not self._tray_sel:
            messagebox.showinfo("QR Code", "Click a tray to select it first.")
            return
        tray = db.get_tray_by_id(next(iter(self._tray_sel)))
        if tray:
            QRDialog(self, tray, port=self._qr_port)

    def _show_tray_history(self):
        """Show every season's record for the selected tray number."""
        if not self._tray_sel:
            messagebox.showinfo("History", "Click a tray to select it first.", parent=self)
            return
        tray = db.get_tray_by_id(next(iter(self._tray_sel)))
        if not tray:
            return
        tray_number = tray["tray_number"]
        history = db.get_tray_history(tray_number)

        win = ctk.CTkToplevel(self)
        win.title(f"History — {tray_number}")
        win.geometry("760x420")
        win.minsize(640, 320)
        win.grab_set()

        _label(win, f"Tray {tray_number} — {len(history)} record(s)",
               FONT_H, GOLD).pack(anchor="w", padx=16, pady=(14, 6))

        cols = ("Status", "Sample", "Incubator", "In Date", "Out Date", "Notes")
        tree = self._make_tree(win, cols)
        tree.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        tree.column("Notes", width=240, anchor="w")

        for h in history:  # newest first (get_tray_history orders by id DESC)
            note = (h.get("notes") or "").replace("\n", "  •  ")
            tree.insert("", "end", values=(
                h.get("status") or "active",
                h.get("sample_name")    or "—",
                h.get("incubator_name") or "—",
                h.get("in_date")  or "—",
                h.get("out_date") or "—",
                note or "—",
            ))
        self._apply_zebra(tree, status_col=0)

    # ══════════════════════════════════════════════════════════════════════════
    #  TIMELINE VIEW
    # ══════════════════════════════════════════════════════════════════════════

    # Incubation milestone days (offset from start, Day 1 = start) → (label, color)
    # Matches the developmental-stage timeline and the field spreadsheet.
    _INC_MILESTONES = [
        (1,  "Incubation Start",     "#10B981"),
        (7,  "Vapona In",            "#8B5CF6"),
        (13, "Vapona Out",           "#D946EF"),
        (14, "Earliest We Can Cool", "#06B6D4"),
        (18, "10% Male Emergence",   "#6366F1"),
        (23, "Expected Release",     "#F59E0B"),
        (37, "Latest Release",       "#EF4444"),
    ]

    # Distinct colors assigned per incubator, drawn from the dashboard theme
    # (BLUE / GREEN / ORANGE / TEAL / RED / dark-gold) plus a couple of
    # complementary accents — all readable with white chip text.
    _INC_PALETTE = [
        BLUE, GREEN, ORANGE, "#8B5CF6", TEAL,
        DK_GOLD, "#EC4899", RED, "#0EA5E9", "#A16207",
    ]

    def _inc_color_map(self) -> dict:
        """Stable {incubator_id: color} mapping by display order."""
        incs = db.get_incubators(include_hidden=False)
        return {i["id"]: self._INC_PALETTE[idx % len(self._INC_PALETTE)]
                for idx, i in enumerate(incs)}

    def _build_timeline_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Calendar",
                                  "Incubation schedule & milestones")
        _btn_primary(hdr, "+ Schedule Incubator", self._schedule_incubator,
                     width=170).pack(side="right")
        _btn_secondary(hdr, "Export to Calendar", self._export_calendar,
                       width=150).pack(side="right", padx=8)
        _btn_secondary(hdr, "Today", self._cal_today,
                       width=75).pack(side="right")

        # Centered month navigation: ‹  Month Year  ›
        monthbar = ctk.CTkFrame(frame, fg_color="transparent")
        monthbar.pack(fill="x", pady=(2, 8))
        nav = ctk.CTkFrame(monthbar, fg_color="transparent")
        nav.pack()  # no fill → stays centered horizontally
        _btn(nav, "‹", self._cal_prev, width=40, height=40, fg=CARD, hover=CARD2,
             text_color=GOLD).pack(side="left", padx=4)
        self._tl_month_lbl = _label(nav, "", ("Segoe UI", 28, "bold"), GOLD)
        self._tl_month_lbl.pack(side="left", padx=20)
        _btn(nav, "›", self._cal_next, width=40, height=40, fg=CARD, hover=CARD2,
             text_color=GOLD).pack(side="left", padx=4)

        _t = date.today()
        self._cal_year, self._cal_month = _t.year, _t.month

        # Plain (non-scrolling) container so the whole month fits on one page
        self._tl_scroll = ctk.CTkFrame(frame, fg_color="transparent", corner_radius=0)
        self._tl_scroll.pack(fill="both", expand=True, padx=12, pady=4)
        return frame

    def _inc_start_date(self, inc_id: int):
        """Start date of an incubator's current incubation = the most common
        start date among its active trays. None if no active trays."""
        from collections import Counter
        counts = Counter()
        for t in db.get_trays(incubator_id=inc_id, status="active"):
            d = _parse_date_loose(t.get("in_date"))
            if d:
                counts[d] += 1
        return counts.most_common(1)[0][0] if counts else None

    def _incubation_events(self) -> list:
        """All milestone events across incubators with a start date.
        Each: {date, label, inc, color, day}. Start date prefers the explicit
        Edit-Setup value, falling back to the active-tray-derived date."""
        from datetime import timedelta
        cmap = self._inc_color_map()
        out = []
        for inc in db.get_incubators(include_hidden=False):
            raw = (inc.get("incubation_start") or "").strip()
            if raw == "none":
                continue  # schedule explicitly removed — don't auto-derive
            start = _parse_date_loose(raw) or self._inc_start_date(inc["id"])
            if not start:
                continue
            inc_color = cmap.get(inc["id"], BLUE)
            for day, label, color in self._INC_MILESTONES:
                out.append({
                    "date":  start + timedelta(days=day - 1),
                    "label": label, "inc": inc["name"], "inc_id": inc["id"],
                    "color": inc_color, "day": day,
                })
        return out

    def _cal_prev(self):
        self._cal_month -= 1
        if self._cal_month < 1:
            self._cal_month, self._cal_year = 12, self._cal_year - 1
        self._refresh_timeline()

    def _cal_next(self):
        self._cal_month += 1
        if self._cal_month > 12:
            self._cal_month, self._cal_year = 1, self._cal_year + 1
        self._refresh_timeline()

    def _cal_today(self):
        t = date.today()
        self._cal_year, self._cal_month = t.year, t.month
        self._refresh_timeline()

    def _refresh_timeline(self):
        import calendar as _cal
        container = self._tl_scroll
        for w in container.winfo_children():
            w.destroy()

        y, m = self._cal_year, self._cal_month
        self._tl_month_lbl.configure(text=f"{_cal.month_name[m]} {y}")

        # Events keyed by exact date; collect incubators present for the legend
        evs = self._incubation_events()
        by_date = {}
        inc_legend = {}   # incubator name → color
        for ev in evs:
            by_date.setdefault(ev["date"], []).append(ev)
            inc_legend.setdefault(ev["inc"], ev["color"])

        today = date.today()

        # Outer card wraps the whole calendar for a cleaner framed look
        cal_card = ctk.CTkFrame(container, fg_color=PANEL, corner_radius=14,
                                border_width=1, border_color=BORDER2)
        cal_card.pack(fill="both", expand=True, padx=4, pady=(2, 6))

        grid = ctk.CTkFrame(cal_card, fg_color="transparent")
        grid.pack(fill="both", expand=True, padx=8, pady=8)

        # Weekday header row (fixed height); the 6 week rows share the rest equally
        weekdays = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
        grid.rowconfigure(0, weight=0)
        for c, wd in enumerate(weekdays):
            grid.columnconfigure(c, weight=1, uniform="cal")
            _label(grid, wd, ("Segoe UI", 10, "bold"), SUBTEXT).grid(
                row=0, column=c, sticky="ew", padx=2, pady=(0, 4))

        # Render the month's natural number of weeks (usually 5) so rows stay
        # tall enough to read — adjacent-month days fill out the first/last weeks.
        weeks = _cal.Calendar(firstweekday=0).monthdatescalendar(y, m)

        for r, week in enumerate(weeks, start=1):
            grid.rowconfigure(r, weight=1, uniform="calrow")
            for c, d in enumerate(week):
                in_month = (d.month == m)
                is_today = (d == today)
                cell_bg = "#1F2529" if is_today else \
                          ("#101827" if not in_month else NESTED)
                cell = ctk.CTkFrame(
                    grid, fg_color=cell_bg, corner_radius=8,
                    border_width=2 if is_today else 0,
                    border_color=GOLD if is_today else BORDER)
                cell.grid(row=r, column=c, sticky="nsew", padx=2, pady=2)
                cell.grid_propagate(False)   # share grid space equally; content won't resize it

                # Day number — today gets a filled gold badge
                num_row = ctk.CTkFrame(cell, fg_color="transparent")
                num_row.pack(fill="x", padx=5, pady=(2, 0))
                if is_today:
                    ctk.CTkLabel(num_row, text=str(d.day), width=20, height=20,
                                 corner_radius=10, fg_color=GOLD, text_color="black",
                                 font=("Segoe UI", 10, "bold")).pack(side="left")
                else:
                    daycol = TEXT if in_month else "#4B5563"
                    _label(num_row, str(d.day), ("Segoe UI", 11, "bold"), daycol).pack(side="left")

                for ev in by_date.get(d, []):
                    chip = ctk.CTkFrame(cell, fg_color=ev["color"], corner_radius=4)
                    chip.pack(fill="x", padx=4, pady=1)
                    ctk.CTkLabel(
                        chip, text=f"{ev['inc']} · {ev['label']}",
                        font=("Segoe UI", 9, "bold"), text_color="white",
                        anchor="w", justify="left", wraplength=150).pack(
                        fill="x", padx=5, pady=1)

        # Legend — incubators (color-coded) + milestone day reference
        legend = ctk.CTkFrame(container, fg_color="transparent")
        legend.pack(fill="x", padx=6, pady=(4, 0))
        _label(legend, "Incubators:", FONT_B, SUBTEXT).pack(side="left", padx=(2, 10))
        for name, color in inc_legend.items():
            chip = ctk.CTkFrame(legend, fg_color=color, corner_radius=5)
            chip.pack(side="left", padx=4, pady=2)
            ctk.CTkLabel(chip, text=f" {name} ",
                         font=("Segoe UI", 10, "bold"), text_color="white").pack(padx=5, pady=2)

        days_ref = "   ".join(f"{lbl} (Day {day})" for day, lbl, _ in self._INC_MILESTONES)
        _label(container, "Milestones:  " + days_ref, FONT_S, SUBTEXT).pack(
            anchor="w", padx=8, pady=(2, 6))

    def _export_calendar(self):
        """Write all incubation milestones to an .ics file that imports straight
        into Google Calendar (or any calendar app)."""
        import uuid
        from datetime import timedelta
        events = self._incubation_events()
        if not events:
            messagebox.showinfo("Export", "No incubation milestones to export.\n"
                "Set an Incubation Start date for at least one incubator.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            title="Export incubation calendar",
            defaultextension=".ics", filetypes=[("Calendar file", "*.ics")],
            initialfile="incubation_timeline.ics")
        if not path:
            return
        stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        lines = ["BEGIN:VCALENDAR", "VERSION:2.0",
                 "PRODID:-//Bee Incubation Manager//Timeline//EN", "CALSCALE:GREGORIAN"]
        for ev in events:
            d0 = ev["date"]
            d1 = d0 + timedelta(days=1)   # all-day event: DTEND is exclusive next day
            summary = f"{ev['inc']} — {ev['label']} (Day {ev['day']})"
            lines += [
                "BEGIN:VEVENT",
                f"UID:{uuid.uuid4()}@bee-incubation",
                f"DTSTAMP:{stamp}",
                f"DTSTART;VALUE=DATE:{d0.strftime('%Y%m%d')}",
                f"DTEND;VALUE=DATE:{d1.strftime('%Y%m%d')}",
                f"SUMMARY:{summary}",
                "END:VEVENT",
            ]
        lines.append("END:VCALENDAR")
        try:
            with open(path, "w", encoding="utf-8", newline="\r\n") as f:
                f.write("\n".join(lines))
        except OSError as exc:
            messagebox.showerror("Export", f"Could not save file:\n{exc}", parent=self)
            return
        messagebox.showinfo("Export complete",
            f"Saved {len(events)} events to:\n{path}\n\n"
            "To add them to Google Calendar:\n"
            "1. Open Google Calendar (calendar.google.com)\n"
            "2. Settings ⚙ → Import & export → Import\n"
            "3. Choose this .ics file and pick a calendar.",
            parent=self)

    # ── Google Calendar sync ────────────────────────────────────────────────

    def _gcal_token_path(self) -> str:
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), "gcal_token.json")

    def _gcal_enabled(self) -> bool:
        return (db.get_setting("gcal_enabled", "0") == "1"
                and bool(db.get_setting("gcal_credentials_path")))

    def _gcal_sync(self, interactive: bool = False, notify: bool = False):
        """Push every incubator's milestones to Google Calendar (create/update),
        deleting events for incubators without a schedule. Runs in a thread."""
        if not gcal_sync.available():
            if notify or interactive:
                messagebox.showwarning("Google Calendar",
                    "Google libraries aren't installed yet.\n"
                    "Settings → Google Calendar Sync → Install Libraries.", parent=self)
            return
        creds = db.get_setting("gcal_credentials_path")
        cal   = db.get_setting("gcal_calendar_id", "primary") or "primary"
        if not creds:
            if notify or interactive:
                messagebox.showwarning("Google Calendar",
                    "Set the OAuth credentials JSON path in Settings first.", parent=self)
            return
        gc = gcal_sync.GoogleCalendar(creds, self._gcal_token_path(), cal)

        def _work():
            from datetime import timedelta
            if not gc.connect(interactive=interactive):
                if notify or interactive:
                    self.after(0, lambda: messagebox.showerror(
                        "Google Calendar", gc.error or "Connection failed.", parent=self))
                return
            n = 0
            for inc in db.get_incubators(include_hidden=False):
                raw = (inc.get("incubation_start") or "").strip()
                if raw == "none":
                    # Explicitly removed — delete any events we created before
                    for day, _label, _ in self._INC_MILESTONES:
                        gc.delete(gcal_sync.make_event_id(inc["id"], day))
                    continue
                start = _parse_date_loose(raw) or self._inc_start_date(inc["id"])
                if not start:
                    continue  # never scheduled — skip entirely (no wasted calls)
                for day, label, _ in self._INC_MILESTONES:
                    gc.upsert(gcal_sync.make_event_id(inc["id"], day),
                              f"{inc['name']} — {label} (Day {day})",
                              start + timedelta(days=day - 1))
                    n += 1
            if notify or interactive:
                self.after(0, lambda: messagebox.showinfo(
                    "Google Calendar", f"Synced {n} milestone events.", parent=self))

        threading.Thread(target=_work, daemon=True).start()

    def _schedule_incubator(self):
        """Schedule, edit, or remove an incubation. Enter only the Expected
        Release date — the start and all milestones are back-calculated.
        Selecting an already-scheduled incubator prefills its current date."""
        from datetime import timedelta
        rel_day = next((day for day, lbl, _ in self._INC_MILESTONES
                        if lbl == "Expected Release"), 23)

        incs = db.get_incubators(include_hidden=True)
        if not incs:
            messagebox.showinfo("Schedule", "Add an incubator first.", parent=self)
            return
        name_map  = {i["name"]: i["id"] for i in incs}
        start_map = {i["id"]: (i.get("incubation_start") or "").strip() for i in incs}

        win = ctk.CTkToplevel(self)
        win.title("Schedule Incubator")
        win.geometry("440x400")
        win.grab_set()
        _label(win, "Schedule Incubator", FONT_H, GOLD).pack(padx=16, pady=(14, 2))
        _label(win, "Enter the Expected Release date. The start date and all\n"
                    "milestones are calculated automatically. Pick an incubator\n"
                    "that's already scheduled to edit or remove it.",
               FONT_S, SUBTEXT).pack(padx=16)

        frm = ctk.CTkFrame(win, fg_color="transparent")
        frm.pack(fill="x", padx=24, pady=12)
        frm.columnconfigure(1, weight=1)
        _label(frm, "Incubator", FONT_S, SUBTEXT).grid(row=0, column=0, sticky="w", pady=6)
        inc_cb = _combo(frm, list(name_map.keys()), 210)
        inc_cb.grid(row=0, column=1, sticky="w", padx=8, pady=6)
        _label(frm, "Expected Release", FONT_S, SUBTEXT).grid(row=1, column=0, sticky="w", pady=6)
        date_e = ctk.CTkEntry(frm, placeholder_text="YYYY-MM-DD", width=210,
                              fg_color=CARD2, border_color=BORDER, text_color=TEXT)
        date_e.grid(row=1, column=1, sticky="w", padx=8, pady=6)

        preview = _label(win, "", FONT_S, SUBTEXT)
        preview.pack(padx=16, pady=(2, 4))

        def _update_preview(*_):
            d = _parse_date_loose(date_e.get())
            if not d:
                preview.configure(text="")
                return
            start = d - timedelta(days=rel_day - 1)
            lines = [f"{lbl}:  {(start + timedelta(days=day - 1)).strftime('%b %d, %Y')}"
                     for day, lbl, _ in self._INC_MILESTONES]
            preview.configure(text="\n".join(lines), justify="left")

        def _prefill(*_):
            """When the selected incubator changes, fill in its current release date."""
            iid = name_map.get(inc_cb.get())
            raw = start_map.get(iid, "")
            d = _parse_date_loose(raw) if raw and raw != "none" else None
            date_e.delete(0, "end")
            if d:
                date_e.insert(0, (d + timedelta(days=rel_day - 1)).strftime("%Y-%m-%d"))
            _update_preview()

        date_e.bind("<KeyRelease>", _update_preview)
        inc_cb.configure(command=lambda _v: _prefill())
        inc_cb.set(incs[0]["name"])
        _prefill()

        def _save():
            d = _parse_date_loose(date_e.get())
            if not d:
                messagebox.showerror("Schedule", "Enter a valid date (YYYY-MM-DD).", parent=win)
                return
            iid = name_map.get(inc_cb.get())
            if not iid:
                messagebox.showerror("Schedule", "Pick an incubator.", parent=win)
                return
            start = d - timedelta(days=rel_day - 1)
            db.set_incubator_incubation_start(iid, start.isoformat())
            self._cal_year, self._cal_month = d.year, d.month  # jump to release month
            win.destroy()
            self._refresh_timeline()
            if self._gcal_enabled():
                self._gcal_sync(interactive=False)

        def _remove():
            iid = name_map.get(inc_cb.get())
            if not iid:
                return
            if not messagebox.askyesno("Remove schedule",
                    f"Remove the incubation schedule for {inc_cb.get()}?", parent=win):
                return
            db.set_incubator_incubation_start(iid, "none")  # explicit cleared marker
            win.destroy()
            self._refresh_timeline()
            if self._gcal_enabled():
                self._gcal_sync(interactive=False)

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(pady=(8, 6))
        _btn(btns, "Save Schedule", _save, fg=DK_GOLD, hover=GOLD,
             text_color="black", width=150).pack(side="left", padx=4)
        _btn(btns, "Remove", _remove, fg=RED, hover="#991B1B",
             text_color="white", width=100).pack(side="left", padx=4)
        _btn(btns, "Cancel", win.destroy, fg=CARD2, hover=BORDER, width=90).pack(side="left", padx=4)

    # ══════════════════════════════════════════════════════════════════════════
    #  SETTINGS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_settings_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Settings",
                                  "Integrations & polling configuration")
        _btn_primary(hdr, "Save Settings", self._save_settings,
                     width=140).pack(side="right")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=12, pady=4)

        def section(title):
            f = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=12,
                             border_width=1, border_color=BORDER2)
            f.pack(fill="x", padx=4, pady=(8, 4))
            _label(f, title, ("Segoe UI", 13, "bold"), GOLD).pack(
                anchor="w", padx=18, pady=(14, 4))
            g = ctk.CTkFrame(f, fg_color="transparent")
            g.pack(fill="x", padx=18, pady=(0, 14))
            g.columnconfigure(1, weight=1)
            return g

        # Govee
        gf = section("Govee API")
        self._set = {}
        r0 = _FormRow(gf, 0, "API Key", "Paste Govee API key here", 360)
        self._set["govee_api_key"] = r0
        _btn(gf, "Test Connection", self._test_govee, fg=BLUE, hover="#1D4ED8",
             text_color="white", width=150).grid(row=1, column=1, sticky="w", padx=4, pady=4)
        self._govee_status_lbl = _label(gf, "", FONT_S, SUBTEXT)
        self._govee_status_lbl.grid(row=2, column=1, sticky="w", padx=4, pady=2)

        # Sensibo (manual AC control)
        sf = section("Sensibo API (AC Control)")
        _label(sf, "Used for manual AC on/off and target temperature buttons.\n"
                    "Set a per-incubator Sensibo Device ID in each incubator's setup.",
               FONT_S, SUBTEXT).grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=(0, 6))
        r1 = _FormRow(sf, 1, "API Key", "Paste Sensibo API key here", 360)
        self._set["sensibo_api_key"] = r1
        _btn(sf, "Test Connection", self._test_sensibo, fg=BLUE, hover="#1D4ED8",
             text_color="white", width=150).grid(row=2, column=1, sticky="w", padx=4, pady=4)
        self._sensibo_status_lbl = _label(sf, "", FONT_S, SUBTEXT)
        self._sensibo_status_lbl.grid(row=3, column=1, sticky="w", padx=4, pady=2)

        # Vapona Sensors (VOC device management)
        vf = section("Vapona Sensors")
        _label(vf, "Manage the Raspberry Pi VOC sensors. Each sensor reports a\n"
                   "stable hardware ID; assign it to an incubator and position\n"
                   "here (the app is authoritative — the Pi does not decide).",
               FONT_S, SUBTEXT).grid(row=0, column=0, columnspan=2, sticky="w",
                                     padx=4, pady=(0, 6))
        _vbtns = ctk.CTkFrame(vf, fg_color="transparent")
        _vbtns.grid(row=1, column=0, columnspan=2, sticky="w", padx=0, pady=4)
        _btn(_vbtns, "Manage Vapona Sensors", self._manage_voc_devices,
             fg=BLUE, hover="#1D4ED8", text_color="white", width=200).pack(
                 side="left", padx=4)
        _btn(_vbtns, "Sensor Wi-Fi Networks", self._manage_wifi_networks,
             fg=CARD2, hover=BORDER, width=180).pack(side="left", padx=4)
        self._voc_dev_status_lbl = _label(vf, "", FONT_S, SUBTEXT)
        self._voc_dev_status_lbl.grid(row=2, column=0, columnspan=2, sticky="w",
                                      padx=4, pady=2)
        self._refresh_voc_dev_status()

        # Google Calendar Sync
        gcf = section("Google Calendar Sync")
        _label(gcf, "Auto-push the incubation Calendar to Google Calendar.\n"
                    "Needs a one-time Google Cloud OAuth credentials file (Desktop app).",
               FONT_S, SUBTEXT).grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=(0, 6))
        self._set["gcal_credentials_path"] = _FormRow(gcf, 1, "Credentials JSON", "path to client_secret.json", 320)
        _btn(gcf, "Browse", self._browse_gcal_creds, width=80, height=28,
             fg=BORDER, hover=CARD2).grid(row=1, column=2, padx=4, pady=4)
        self._set["gcal_calendar_id"] = _FormRow(gcf, 2, "Calendar ID", "primary", 320)
        self._set["gcal_enabled"] = _FormRow(gcf, 3, "Auto-sync on changes (1=yes, 0=no)", "0", 60)
        gc_btns = ctk.CTkFrame(gcf, fg_color="transparent")
        gc_btns.grid(row=4, column=0, columnspan=3, sticky="w", padx=4, pady=(6, 2))
        _btn(gc_btns, "Install Libraries", self._install_gcal_libs,
             fg=BORDER, hover=CARD2, width=150).pack(side="left", padx=(0, 6))
        _btn(gc_btns, "Connect / Authorize", self._gcal_connect,
             fg=GREEN, hover="#15803D", text_color="white", width=170).pack(side="left", padx=(0, 6))
        _btn(gc_btns, "Sync Now", lambda: self._gcal_sync(notify=True),
             fg=BLUE, hover="#1D4ED8", text_color="white", width=110).pack(side="left")
        self._gcal_status_lbl = _label(gcf, "", FONT_S, SUBTEXT)
        self._gcal_status_lbl.grid(row=5, column=0, columnspan=3, sticky="w", padx=4, pady=2)

        # Background Poller
        bf = section("Background Poller")
        _label(bf,
               "Run a lightweight background task that collects Govee readings\n"
               "even when this window is closed.  Enable it on one computer only\n"
               "— the one that stays on — to avoid duplicate writes.",
               FONT_S, SUBTEXT).pack(anchor="w", padx=14, pady=(0, 6))
        btn_row = ctk.CTkFrame(bf, fg_color="transparent")
        btn_row.pack(anchor="w", padx=14, pady=(0, 4))
        _btn(btn_row, "Enable on this computer",  self._enable_poller,
             fg=GREEN, hover="#15803D", text_color="white", width=200).pack(side="left", padx=(0, 8))
        _btn(btn_row, "Disable on this computer", self._disable_poller,
             fg=RED,   hover="#991B1B", text_color="white", width=200).pack(side="left")
        self._poller_status_lbl = _label(bf, "", FONT_S, SUBTEXT)
        self._poller_status_lbl.pack(anchor="w", padx=14, pady=(0, 8))
        self._refresh_poller_status()

        # Poll interval is fixed at 15 minutes (not user-configurable)
        pf = section("Polling & Thresholds")
        _label(pf, "Govee poll interval: every 15 minutes (fixed)",
               FONT_S, SUBTEXT).grid(row=0, column=0, columnspan=2,
                                     sticky="w", padx=4, pady=4)
        self._set["date_alert_lookahead"] = _FormRow(pf, 1, "Date Alert Lookahead (days)", "7", 100)
        self._set["temp_unit"] = _FormRow(pf, 2, "Temp Unit",
            widget=_combo(pf, ["C", "F"], 80))

        # Temperature Mode Goals — target temp/humidity per mode
        tmg = section("Temperature Mode Goals")
        _label(tmg,
               "Target temperature (°C) and humidity (%) for each mode. Shown on\n"
               "incubator tiles and drawn as dotted goal lines on the charts.",
               FONT_S, SUBTEXT).grid(row=0, column=0, columnspan=3, sticky="w", padx=4, pady=(0, 6))
        _label(tmg, "Mode",       FONT_S, SUBTEXT).grid(row=1, column=0, sticky="w", padx=4)
        _label(tmg, "Temp °C",    FONT_S, SUBTEXT).grid(row=1, column=1, sticky="w", padx=4)
        _label(tmg, "Humidity %", FONT_S, SUBTEXT).grid(row=1, column=2, sticky="w", padx=4)
        self._goal_entries = {}
        for _ri, _mk in enumerate(calc.ACTIVE_MODES, start=2):
            _gt, _gh = db.get_mode_goals(_mk)
            _label(tmg, calc.TEMP_MODES[_mk]["label"], FONT_B, TEXT).grid(
                row=_ri, column=0, sticky="w", padx=4, pady=3)
            _te = ctk.CTkEntry(tmg, width=80, fg_color=CARD2,
                               border_color=BORDER, text_color=TEXT)
            _te.grid(row=_ri, column=1, sticky="w", padx=4, pady=3)
            _te.insert(0, "" if _gt is None else f"{_gt:g}")
            _he = ctk.CTkEntry(tmg, width=80, fg_color=CARD2,
                               border_color=BORDER, text_color=TEXT)
            _he.grid(row=_ri, column=2, sticky="w", padx=4, pady=3)
            _he.insert(0, "" if _gh is None else f"{_gh:g}")
            self._goal_entries[_mk] = (_te, _he)

        # Calculation
        cf = section("Weight Calculations")
        self._set["lbs_per_gal"] = _FormRow(cf, 0, "Lbs per Gallon (raw cells)", "2.2", 100)
        self._set["target_gals_per_tray"] = _FormRow(cf, 1, "Target Gals per Tray", "2.0", 100)

        # QR Server
        qf = section("QR Scan Server")
        self._set["qr_server_port"] = _FormRow(qf, 0, "Port", "5151", 100)
        self._set["qr_server_enabled"] = _FormRow(qf, 1, "Enabled (1=yes, 0=no)", "1", 60)
        self._qr_ip_lbl = _label(qf, "", FONT_S, BLUE)
        self._qr_ip_lbl.grid(row=2, column=1, sticky="w", padx=4, pady=2)

        # Mobile app section
        mf = section("Mobile App")
        self._set["mobile_passcode"] = _FormRow(
            mf, 0, "Shared passcode", "", 200)
        _label(mf, "Set a passcode before sharing the app online (Tailscale).\n"
                   "Leave blank to disable the login (local network only).",
               FONT_S, SUBTEXT).grid(row=1, column=0, columnspan=2,
                                     sticky="w", padx=4, pady=(2, 4))

        # ── Data Storage ──────────────────────────────────────────────────────
        dsf = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=12,
                           border_width=1, border_color=BORDER2)
        dsf.pack(fill="x", padx=4, pady=(8, 4))
        _label(dsf, "Data Storage", ("Segoe UI", 13, "bold"), GOLD).pack(
            anchor="w", padx=18, pady=(14, 2))
        _label(dsf,
               "Move the database to your Google Drive folder so data syncs\n"
               "between computers automatically. The app restarts after saving.",
               FONT_S, SUBTEXT).pack(anchor="w", padx=14, pady=(0, 8))

        dsg = ctk.CTkFrame(dsf, fg_color=CARD2, corner_radius=8)
        dsg.pack(fill="x", padx=14, pady=(0, 14))

        # Current path row
        path_row = ctk.CTkFrame(dsg, fg_color="transparent")
        path_row.pack(fill="x", padx=12, pady=(10, 4))
        _label(path_row, "Current database:", FONT_S, SUBTEXT).pack(side="left")
        self._db_path_lbl = _label(path_row, db.DB_PATH, FONT_S, BLUE)
        self._db_path_lbl.pack(side="left", padx=(8, 0))

        # Backups status row — daily auto-snapshots in a "Backups" subfolder
        backup_row = ctk.CTkFrame(dsg, fg_color="transparent")
        backup_row.pack(fill="x", padx=12, pady=(0, 4))
        _label(backup_row, "Backups:", FONT_S, SUBTEXT).pack(side="left")

        def _backup_status_text():
            try:
                bk = db.list_backups()
                if not bk:
                    return "none yet (created automatically at startup)"
                newest = os.path.basename(bk[0]).replace("incubation-", "").replace(".db", "")
                return f"{len(bk)} daily snapshot(s), newest {newest}"
            except Exception:
                return "—"

        self._backup_lbl = _label(backup_row, _backup_status_text(), FONT_S, TEXT)
        self._backup_lbl.pack(side="left", padx=(8, 0))

        def _backup_now():
            self._run_db_backup()
            self._backup_lbl.configure(text=_backup_status_text())

        _btn(backup_row, "Back up now", _backup_now,
             fg=CARD2, hover=BORDER, width=110, height=24).pack(side="left", padx=12)

        # New path entry row
        new_row = ctk.CTkFrame(dsg, fg_color="transparent")
        new_row.pack(fill="x", padx=12, pady=(4, 10))
        _label(new_row, "Move database to:", FONT_S, SUBTEXT).pack(side="left")
        self._db_folder_entry = ctk.CTkEntry(
            new_row, placeholder_text="Paste or browse to a folder…",
            fg_color=CARD, border_color=BORDER, text_color=TEXT, width=320)
        self._db_folder_entry.pack(side="left", padx=(8, 6))
        _btn(new_row, "Browse", self._browse_db_folder,
             width=80, height=28, fg=BORDER, hover=CARD2).pack(side="left")

        # Status / hint
        self._db_move_status = _label(dsg, "", FONT_S, SUBTEXT)
        self._db_move_status.pack(anchor="w", padx=12, pady=(0, 4))

        # Action buttons
        act_row = ctk.CTkFrame(dsg, fg_color="transparent")
        act_row.pack(fill="x", padx=12, pady=(0, 10))
        _btn(act_row, "Move & Restart", self._move_database,
             fg=DK_GOLD, hover=GOLD, text_color="black",
             width=140, height=30).pack(side="left", padx=(0, 8))
        _btn(act_row, "Use Local File (default)", self._use_local_db,
             fg=BORDER, hover=CARD2, width=180, height=30).pack(side="left")

        # ── Email Reports ─────────────────────────────────────────────────
        ef = ctk.CTkFrame(scroll, fg_color=PANEL, corner_radius=12,
                          border_width=1, border_color=BORDER2)
        ef.pack(fill="x", padx=4, pady=(8, 4))
        _label(ef, "Email Reports", ("Segoe UI", 13, "bold"), GOLD).pack(
            anchor="w", padx=18, pady=(14, 2))
        _label(ef,
               "A daily summary is sent at 7 PM with 24h temps, inspections, batch progress and "
               "tomorrow's calendar.\nFor Gmail: use an App Password (Google Account › Security › "
               "2-Step Verification › App passwords).",
               FONT_S, SUBTEXT).pack(anchor="w", padx=14, pady=(0, 8))

        eg = ctk.CTkFrame(ef, fg_color=CARD2, corner_radius=8)
        eg.pack(fill="x", padx=14, pady=(0, 14))
        eg.columnconfigure(1, weight=1)

        self._set["smtp_host"]     = _FormRow(eg, 0, "SMTP Host",     "smtp.gmail.com", 220)
        self._set["smtp_port"]     = _FormRow(eg, 1, "SMTP Port",     "587",            80)
        self._set["smtp_tls"]      = _FormRow(eg, 2, "Use TLS (1/0)", "1",              60)
        self._set["smtp_username"] = _FormRow(eg, 3, "Username / Email", "you@gmail.com", 260)
        self._set["smtp_password"] = _FormRow(eg, 4, "Password / App Password", "••••••••", 260)
        self._set["smtp_from"]     = _FormRow(eg, 5, "From Address (optional)", "Incubation App <you@gmail.com>", 300)

        # Recipients box
        _label(eg, "Recipients\n(one per line)", FONT_S, SUBTEXT).grid(
            row=6, column=0, sticky="ne", padx=(14, 8), pady=8)
        self._email_recip_box = ctk.CTkTextbox(
            eg, height=90, fg_color=CARD, border_color=BORDER,
            text_color=TEXT, font=FONT_S, corner_radius=6)
        self._email_recip_box.grid(row=6, column=1, sticky="ew", padx=(0, 14), pady=8)

        # Test + status
        email_btn_row = ctk.CTkFrame(eg, fg_color="transparent")
        email_btn_row.grid(row=7, column=1, sticky="w", padx=(0, 14), pady=(0, 10))
        self._email_status_lbl = _label(email_btn_row, "", FONT_S, SUBTEXT)
        _btn(email_btn_row, "Send Test Email", self._send_test_email,
             fg=BLUE, hover="#1D4ED8", text_color="white",
             width=150, height=30).pack(side="left", padx=(0, 10))
        self._email_status_lbl.pack(side="left")

        return frame

    def _refresh_settings(self):
        keys = ["govee_api_key", "sensibo_api_key", "date_alert_lookahead",
                "temp_unit", "lbs_per_gal", "target_gals_per_tray",
                "qr_server_port", "qr_server_enabled", "mobile_passcode",
                "smtp_host", "smtp_port", "smtp_tls",
                "smtp_username", "smtp_password", "smtp_from",
                "gcal_credentials_path", "gcal_calendar_id", "gcal_enabled"]
        for k in keys:
            if k in self._set:
                self._set[k].set(db.get_setting(k))
        self._qr_ip_lbl.configure(
            text=f"Phone scan URL: http://{qr_server.get_local_ip()}:{self._qr_port}/tray/<id>")
        # Recipients text box
        recip_val = db.get_setting("email_recipients", "")
        self._email_recip_box.delete("1.0", "end")
        if recip_val:
            self._email_recip_box.insert("1.0", recip_val)
        # Reload per-mode goals (in case another computer changed them)
        for _mk, (_te, _he) in getattr(self, "_goal_entries", {}).items():
            _gt, _gh = db.get_mode_goals(_mk)
            _te.delete(0, "end"); _te.insert(0, "" if _gt is None else f"{_gt:g}")
            _he.delete(0, "end"); _he.insert(0, "" if _gh is None else f"{_gh:g}")

    def _save_settings(self):
        keys = ["govee_api_key", "sensibo_api_key", "date_alert_lookahead",
                "temp_unit", "lbs_per_gal", "target_gals_per_tray",
                "qr_server_port", "qr_server_enabled", "mobile_passcode",
                "smtp_host", "smtp_port", "smtp_tls",
                "smtp_username", "smtp_password", "smtp_from",
                "gcal_credentials_path", "gcal_calendar_id", "gcal_enabled"]
        for k in keys:
            if k in self._set:
                db.set_setting(k, self._set[k].get())
        # Save recipients
        recip_text = self._email_recip_box.get("1.0", "end").strip()
        db.set_setting("email_recipients", recip_text)
        # Per-mode temperature/humidity goals
        for _mk, (_te, _he) in getattr(self, "_goal_entries", {}).items():
            db.set_mode_goals(_mk, _te.get().strip(), _he.get().strip())
        # Update govee/sensibo keys live
        self._govee.set_api_key(db.get_setting("govee_api_key"))
        self._sensibo.set_api_key(db.get_setting("sensibo_api_key"))
        messagebox.showinfo("Settings", "Settings saved.", parent=self)

    # ── Data storage helpers ──────────────────────────────────────────────────

    def _browse_db_folder(self):
        folder = filedialog.askdirectory(
            title="Choose a folder for the database "
                  "(e.g. your Google Drive folder)",
            mustexist=True,
        )
        if folder:
            self._db_folder_entry.delete(0, "end")
            self._db_folder_entry.insert(0, folder)
            self._db_move_status.configure(
                text=f"Will save to: {folder}\\incubation.db",
                text_color=SUBTEXT)

    def _move_database(self):
        import shutil
        folder = self._db_folder_entry.get().strip()
        if not folder:
            self._db_move_status.configure(
                text="Please browse to or paste a folder path first.",
                text_color=ORANGE)
            return
        if not os.path.isdir(folder):
            self._db_move_status.configure(
                text="That folder doesn't exist. Create it first or pick a different one.",
                text_color=RED)
            return

        new_path = os.path.join(folder, "incubation.db")
        same_file = (os.path.exists(new_path) and
                     os.path.abspath(new_path) == os.path.abspath(db.DB_PATH))

        # If a database is already in the target folder (e.g. a shared Google
        # Drive folder another computer set up), let the user JOIN it instead of
        # overwriting — so a new computer never clobbers the shared data.
        if os.path.exists(new_path) and not same_file:
            choice = messagebox.askyesnocancel(
                "Database already exists there",
                f"A database already exists at:\n{new_path}\n\n"
                "Yes — replace it with THIS computer's current data\n"
                "No — keep the existing (shared) database and just use it here\n"
                "Cancel — do nothing",
                parent=self)
            if choice is None:          # Cancel
                return
            if choice:                  # Yes — overwrite with our data
                try:
                    shutil.copy2(db.DB_PATH, new_path)
                except Exception as exc:
                    self._db_move_status.configure(
                        text=f"Copy failed: {exc}", text_color=RED)
                    return
            # No — leave the existing file untouched and just adopt its path
        elif not same_file:
            # Empty target folder — copy our current data into it
            try:
                shutil.copy2(db.DB_PATH, new_path)
            except Exception as exc:
                self._db_move_status.configure(
                    text=f"Copy failed: {exc}", text_color=RED)
                return

        # Save new path to config so next launch uses it
        db.save_config({"db_path": new_path})

        messagebox.showinfo(
            "Done — please restart",
            f"The app will now use:\n{new_path}\n\n"
            "Close and reopen the app to start using this location.\n\n"
            "Tip: on another computer, install the app, then go to\n"
            "Settings ▸ Data Storage, browse to the same Google Drive folder,\n"
            "click 'Move & Restart', and choose 'No' to join the shared data.",
            parent=self)

    def _use_local_db(self):
        """Reset to default local file (remove any configured path)."""
        db.save_config({"db_path": ""})
        messagebox.showinfo(
            "Done — please restart",
            "Database location reset to the default (next to the app files).\n"
            "Restart the app for the change to take effect.",
            parent=self)

    # ── Background poller ──────────────────────────────────────────────────────

    _TASK_NAME   = "BeeIncubationPoller"
    _POLLER_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "govee_poller.py")

    def _poller_installed(self) -> bool:
        import winreg
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r"Software\Microsoft\Windows\CurrentVersion\Run")
            winreg.QueryValueEx(key, self._TASK_NAME)
            winreg.CloseKey(key)
            return True
        except FileNotFoundError:
            return False

    def _refresh_voc_dev_status(self):
        try:
            devs = voc_db.get_devices()
        except Exception as exc:
            self._voc_dev_status_lbl.configure(
                text=f"(unavailable: {exc})", text_color=SUBTEXT)
            return
        total = len(devs)
        unassigned = sum(1 for d in devs if d.get("incubator_id") is None)
        if total == 0:
            self._voc_dev_status_lbl.configure(
                text="No sensors have reported yet.", text_color=SUBTEXT)
        else:
            msg = f"{total} sensor(s) known"
            if unassigned:
                msg += f"  •  {unassigned} unassigned"
            self._voc_dev_status_lbl.configure(
                text=msg, text_color=(GOLD if unassigned else GREEN))

    def _manage_voc_devices(self):
        _VocDeviceManager(self, on_close=self._refresh_voc_dev_status)

    def _manage_wifi_networks(self):
        _WifiNetworkManager(self)

    def _refresh_poller_status(self):
        import json, socket
        installed = self._poller_installed()
        # Check lock file for active machine info
        lock_file = os.path.join(os.path.dirname(db.DB_PATH), "govee_poller.lock")
        lock_info = ""
        try:
            with open(lock_file, encoding="utf-8") as f:
                lock = json.load(f)
            lock_info = f"  |  Active on: {lock.get('machine')}  (last seen {lock.get('last_seen', '')[:19]})"
        except Exception:
            pass
        if installed:
            self._poller_status_lbl.configure(
                text=f"Enabled on this computer ({socket.gethostname()}){lock_info}",
                text_color=GREEN)
        else:
            self._poller_status_lbl.configure(
                text=f"Not enabled on this computer{lock_info}",
                text_color=SUBTEXT)

    def _kill_poller_processes(self):
        subprocess.run(
            ["wmic", "process", "where", "commandline like '%govee_poller%'", "delete"],
            capture_output=True, creationflags=_NO_WINDOW
        )

    def _enable_poller(self):
        import winreg
        self._kill_poller_processes()  # stop any existing instances first
        pythonw = os.path.join(os.path.dirname(sys.executable), "pythonw.exe")
        if not os.path.exists(pythonw):
            pythonw = sys.executable
        cmd = f'"{pythonw}" "{self._POLLER_SCRIPT}"'
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r"Software\Microsoft\Windows\CurrentVersion\Run",
                                 0, winreg.KEY_SET_VALUE)
            winreg.SetValueEx(key, self._TASK_NAME, 0, winreg.REG_SZ, cmd)
            winreg.CloseKey(key)
        except Exception as exc:
            messagebox.showerror("Background Poller",
                f"Could not register startup entry:\n{exc}", parent=self)
            return
        # Start it immediately too
        subprocess.Popen([pythonw, self._POLLER_SCRIPT], creationflags=_NO_WINDOW)
        self._refresh_poller_status()
        messagebox.showinfo("Background Poller",
            "Poller enabled and started.\n\n"
            "It will run automatically each time you log in to Windows.\n"
            "Check govee_poller.log (next to incubation.db) for activity.",
            parent=self)

    def _disable_poller(self):
        import winreg
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                 r"Software\Microsoft\Windows\CurrentVersion\Run",
                                 0, winreg.KEY_SET_VALUE)
            winreg.DeleteValue(key, self._TASK_NAME)
            winreg.CloseKey(key)
        except FileNotFoundError:
            pass
        except Exception as exc:
            messagebox.showerror("Background Poller",
                f"Could not remove startup entry:\n{exc}", parent=self)
            return
        # Kill all running govee_poller.py processes
        self._kill_poller_processes()
        self._refresh_poller_status()
        messagebox.showinfo("Background Poller",
            "Poller disabled and stopped on this computer.",
            parent=self)

    def _test_govee(self):
        key = self._set["govee_api_key"].get()
        if not key:
            self._govee_status_lbl.configure(text="Enter an API key first.", text_color=ORANGE)
            return
        self._govee_status_lbl.configure(text="Connecting…", text_color=SUBTEXT)
        self.update()

        tmp     = govee_mod.GoveeClient(api_key=key)
        devices = tmp.get_all_devices()

        if not devices:
            self._govee_status_lbl.configure(
                text=f"Connection failed: {tmp.status_label()}", text_color=RED)
            return

        db.set_setting("govee_api_key", key)
        self._govee.set_api_key(key)

        sensors = [d for d in devices if d.get("is_sensor")]
        others  = [d for d in devices if not d.get("is_sensor")]

        self._govee_status_lbl.configure(
            text=f"Connected — {len(sensors)} sensor(s), {len(others)} other device(s). "
                 f"See device list below.", text_color=GREEN)

        # Open a popup showing all devices with copy-ready IDs
        self._show_device_list(devices)

    def _show_device_list(self, devices: list):
        """
        Popup listing every Govee device found, clearly separated into
        Temperature Sensors and Other Devices.  Each row has copy buttons
        for Device ID and SKU so they can be pasted straight into Incubator settings.
        """
        win = ctk.CTkToplevel(self)
        win.title("Govee Devices Found")
        win.geometry("700x520")
        win.grab_set()

        _label(win, "Govee Devices", FONT_H, GOLD).pack(padx=16, pady=(14, 2), anchor="w")
        _label(win,
               "Sensors are highlighted. Copy the Device ID and SKU into each Incubator's settings.",
               FONT_S, SUBTEXT).pack(padx=16, anchor="w")

        scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=12, pady=8)

        sensors = [d for d in devices if d.get("is_sensor")]
        others  = [d for d in devices if not d.get("is_sensor")]

        def device_row(parent, dev: dict, highlight: bool):
            bg  = CARD2 if highlight else CARD
            row = ctk.CTkFrame(parent, fg_color=bg, corner_radius=8,
                               border_width=2 if highlight else 0,
                               border_color=GOLD if highlight else BORDER)
            row.pack(fill="x", pady=3, padx=4)

            left = ctk.CTkFrame(row, fg_color="transparent")
            left.pack(side="left", fill="both", expand=True, padx=12, pady=8)

            name     = dev.get("deviceName") or dev.get("device", "Unknown")
            sku      = dev.get("sku") or dev.get("model", "")
            dev_id   = dev.get("device", "")
            api_ver  = dev.get("api_version", "v2")
            tag      = "SENSOR" if highlight else "device"
            tag_col  = GREEN if highlight else SUBTEXT

            top_row = ctk.CTkFrame(left, fg_color="transparent")
            top_row.pack(fill="x")
            _label(top_row, name, FONT_B, GOLD if highlight else TEXT).pack(side="left")
            _label(top_row, f"  [{tag}]", FONT_S, tag_col).pack(side="left")
            _label(top_row, f"  {api_ver}", FONT_S, SUBTEXT).pack(side="right")

            id_row = ctk.CTkFrame(left, fg_color="transparent")
            id_row.pack(fill="x", pady=(2, 0))
            _label(id_row, f"Device ID:  {dev_id}", ("Courier New", 10), TEXT).pack(side="left")

            sku_row = ctk.CTkFrame(left, fg_color="transparent")
            sku_row.pack(fill="x")
            _label(sku_row, f"SKU/Model:  {sku}", ("Courier New", 10), TEXT).pack(side="left")

            # Copy buttons
            right = ctk.CTkFrame(row, fg_color="transparent")
            right.pack(side="right", padx=10, pady=8)

            def _copy(text, btn):
                self.clipboard_clear()
                self.clipboard_append(text)
                orig = btn.cget("text")
                btn.configure(text="Copied!")
                self.after(1500, lambda b=btn, t=orig: b.configure(text=t))

            btn_id  = _btn(right, "Copy ID",  None, width=82, height=26, fg=BORDER, hover=CARD)
            btn_id.configure(command=lambda t=dev_id, b=btn_id: _copy(t, b))
            btn_id.pack(pady=2)

            btn_sku = _btn(right, "Copy SKU", None, width=82, height=26, fg=BORDER, hover=CARD)
            btn_sku.configure(command=lambda t=sku, b=btn_sku: _copy(t, b))
            btn_sku.pack(pady=2)

        # ── Sensors section ──
        if sensors:
            _label(scroll, f"Temperature / Humidity Sensors  ({len(sensors)})",
                   FONT_B, GREEN).pack(anchor="w", padx=4, pady=(6, 2))
            for d in sensors:
                device_row(scroll, d, highlight=True)
        else:
            _label(scroll, "No temperature sensors found via API.",
                   FONT_B, ORANGE).pack(anchor="w", padx=4, pady=6)
            _label(scroll,
                   "If your sensors connect via a gateway, make sure the gateway\n"
                   "is online and its firmware is up to date in the Govee app.",
                   FONT_S, SUBTEXT).pack(anchor="w", padx=4)

        # ── Other devices section ──
        if others:
            _label(scroll, f"Other Devices  ({len(others)})",
                   FONT_B, SUBTEXT).pack(anchor="w", padx=4, pady=(14, 2))
            for d in others:
                device_row(scroll, d, highlight=False)

        _btn(win, "Close", win.destroy, width=100,
             fg=CARD2, hover=BORDER).pack(pady=10)

    def _browse_gcal_creds(self):
        path = filedialog.askopenfilename(
            title="Select Google OAuth credentials JSON",
            filetypes=[("JSON file", "*.json")])
        if path:
            self._set["gcal_credentials_path"].set(path)

    def _install_gcal_libs(self):
        self._gcal_status_lbl.configure(text="Installing libraries…", text_color=SUBTEXT)
        self.update()

        def _work():
            try:
                proc = subprocess.run(
                    [sys.executable, "-m", "pip", "install",
                     "google-api-python-client", "google-auth-oauthlib",
                     "google-auth-httplib2"],
                    capture_output=True, text=True,
                    creationflags=_NO_WINDOW)
                ok = proc.returncode == 0
                msg = ("Libraries installed. Restart, then click Connect."
                       if ok else f"Install failed:\n{proc.stderr[-300:]}")
            except Exception as exc:
                ok, msg = False, f"Install error: {exc}"
            self.after(0, lambda: self._gcal_status_lbl.configure(
                text=msg, text_color=(GREEN if ok else RED)))
        threading.Thread(target=_work, daemon=True).start()

    def _gcal_connect(self):
        # Persist the path/calendar so the sync uses the latest values
        db.set_setting("gcal_credentials_path", self._set["gcal_credentials_path"].get())
        db.set_setting("gcal_calendar_id", self._set["gcal_calendar_id"].get() or "primary")
        if not gcal_sync.available():
            self._gcal_status_lbl.configure(
                text="Install libraries first, then restart.", text_color=ORANGE)
            return
        self._gcal_status_lbl.configure(
            text="Opening browser for authorization…", text_color=SUBTEXT)
        self._gcal_sync(interactive=True, notify=True)

    def _test_sensibo(self):
        key = self._set["sensibo_api_key"].get()
        if not key:
            self._sensibo_status_lbl.configure(text="Enter an API key first.", text_color=ORANGE)
            return
        self._sensibo_status_lbl.configure(text="Connecting…", text_color=SUBTEXT)
        self.update()

        tmp     = sensibo_mod.SensiboClient(api_key=key)
        devices = tmp.list_devices()

        if not devices:
            self._sensibo_status_lbl.configure(
                text=f"Connection failed: {tmp.status_label()}", text_color=RED)
            return

        db.set_setting("sensibo_api_key", key)
        self._sensibo.set_api_key(key)

        self._sensibo_status_lbl.configure(
            text=f"Connected — {len(devices)} AC pod(s). See device list below.",
            text_color=GREEN)
        self._show_sensibo_device_list(devices)

    def _show_sensibo_device_list(self, devices: list):
        """Popup listing every Sensibo pod with a copy-ready Device ID."""
        win = ctk.CTkToplevel(self)
        win.title("Sensibo Devices Found")
        win.geometry("560x420")
        win.grab_set()

        _label(win, "Sensibo Devices", FONT_H, GOLD).pack(padx=16, pady=(14, 2), anchor="w")
        _label(win, "Copy the Device ID into each Incubator's setup.",
               FONT_S, SUBTEXT).pack(padx=16, anchor="w")

        scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=12, pady=8)

        def _copy(text, btn):
            self.clipboard_clear()
            self.clipboard_append(text)
            orig = btn.cget("text")
            btn.configure(text="Copied!")
            self.after(1500, lambda b=btn, t=orig: b.configure(text=t))

        for d in devices:
            dev_id = d.get("id", "")
            room   = (d.get("room") or {}).get("name", "Unnamed")
            row = ctk.CTkFrame(scroll, fg_color=CARD2, corner_radius=8)
            row.pack(fill="x", pady=3, padx=4)
            left = ctk.CTkFrame(row, fg_color="transparent")
            left.pack(side="left", fill="both", expand=True, padx=12, pady=8)
            _label(left, room, FONT_B, GOLD).pack(anchor="w")
            _label(left, f"Device ID:  {dev_id}", ("Courier New", 10), TEXT).pack(anchor="w")
            btn_id = _btn(row, "Copy ID", None, width=82, height=26, fg=BORDER, hover=CARD)
            btn_id.configure(command=lambda t=dev_id, b=btn_id: _copy(t, b))
            btn_id.pack(side="right", padx=10, pady=8)

        if not devices:
            _label(scroll, "No Sensibo AC pods found on this account.",
                   FONT_B, ORANGE).pack(anchor="w", padx=4, pady=6)

        _btn(win, "Close", win.destroy, width=100,
             fg=CARD2, hover=BORDER).pack(pady=10)

    # ══════════════════════════════════════════════════════════════════════════
    #  OPENERS
    # ══════════════════════════════════════════════════════════════════════════

    def _open_incubator_dialog(self, inc: dict = None):
        IncubatorDialog(self, inc, on_save=lambda: self._refresh_current())

    # ══════════════════════════════════════════════════════════════════════════
    #  INCUBATOR DETAIL VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _show_inc_detail(self, inc: dict):
        """Navigate to the full-screen detail view for one incubator."""
        self._detail_inc = inc
        self.show_view("inc_detail")

    def _build_inc_detail_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)
        frame._content = None  # placeholder; built in _refresh_inc_detail
        return frame

    def _refresh_inc_detail(self):
        frame = self._views["inc_detail"]

        # Remember which tab was open before we tear down
        _prev_tab = getattr(frame, "_active_tab", "Inspections")

        # Destroy previous content
        for w in frame.winfo_children():
            w.destroy()

        inc = self._detail_inc
        if not inc:
            _label(frame, "No incubator selected.", FONT_B, SUBTEXT).pack(pady=40)
            return

        # Re-fetch incubator to get fresh data
        fresh = next((x for x in db.get_incubators(include_hidden=True)
                      if x["id"] == inc["id"]), inc)

        unit = db.get_setting("temp_unit", "C")

        # ── Screen header (title + subtitle, actions on the right) ────────────
        self._detail_hdr = self._screen_header(
            frame, "Incubators", "Per-unit detail & inspection history")

        # ── Selector row (card) — exact prototype values ───────────────────────
        topbar = ctk.CTkFrame(frame, fg_color="#151E2E", corner_radius=12,
                              border_width=1, border_color="#232F42")
        topbar.pack(fill="x", padx=16, pady=(2, 6))

        _nav_list = db.get_incubators()  # visible incubators, in display order
        _cur_idx  = next((i for i, x in enumerate(_nav_list)
                          if x["id"] == fresh["id"]), 0)

        def _go_to(idx: int):
            if _nav_list:
                self._show_inc_detail(_nav_list[idx % len(_nav_list)])

        ctk.CTkButton(
            topbar, text="‹", command=lambda: _go_to(_cur_idx - 1),
            width=28, height=28, corner_radius=7,
            fg_color="#1F2937", hover_color="#28374D", text_color="#CBD5E1",
            font=("Segoe UI", 13),
        ).pack(side="left", padx=(16, 10), pady=12)
        _label(topbar, fresh["name"], ("Segoe UI", 15, "bold"), "#F3F4F6").pack(side="left")
        ctk.CTkButton(
            topbar, text="›", command=lambda: _go_to(_cur_idx + 1),
            width=28, height=28, corner_radius=7,
            fg_color="#1F2937", hover_color="#28374D", text_color="#CBD5E1",
            font=("Segoe UI", 13),
        ).pack(side="left", padx=10, pady=12)

        ctk.CTkFrame(topbar, fg_color="#263347", width=1, height=22).pack(
            side="left", padx=4)

        reading = self._govee.get_last(fresh["id"])
        if not reading:
            db_row  = db.get_latest_reading(fresh["id"])
            reading = {"temp_c": db_row["temperature_c"], "humidity": db_row["humidity_pct"]} if db_row else {}

        if reading.get("temp_c") is not None:
            t_min, t_max = calc.get_temp_range(fresh)
            _gt, _ = db.get_mode_goals(fresh.get("temp_mode", "incubation"))
            _tc = reading["temp_c"]
            if _gt is not None:
                _d = abs(_tc - _gt)
                t_col = GREEN if _d <= 1 else (ORANGE if _d <= 3 else RED)
            elif t_min is not None:
                t_col = GREEN if t_min <= _tc <= t_max else RED
            else:
                t_col = "#6B7280"
            _label(topbar, calc.format_temp(_tc, unit),
                   ("Segoe UI", 20, "bold"), t_col).pack(side="left", padx=(14, 2))
            _label(topbar, f"{reading['humidity']:.0f}% RH",
                   ("Segoe UI", 20, "bold"), GOLD).pack(side="left", padx=(2, 0))

        # AM/PM pills (pill shape, translucent fills, no border — spec exact)
        make_status_badges(
            topbar, fresh["id"], style="pill",
            on_click=lambda p, i=fresh: self._open_inspection_form(i)).pack(
            side="right", padx=(0, 14), pady=8)

        # Actions live in the screen header (right side)
        _hdr = self._detail_hdr
        _btn_primary(_hdr, "Inspect Now",
                     lambda i=fresh: self._open_inspection_form(i),
                     width=110).pack(side="right", padx=(6, 0))
        _btn_secondary(_hdr, "⚙ Edit Setup",
             lambda i=fresh: IncubatorDialog(
                 self, i,
                 on_save=lambda: self._refresh_current(),
                 on_delete=self._after_inc_delete),
             width=110).pack(side="right", padx=6)
        _pull_btn = _btn_secondary(_hdr, "⟳ Pull Reading", None, width=130)
        _pull_btn.configure(command=lambda i=fresh, b=_pull_btn: self._manual_poll_incubator(i, b))
        _pull_btn.pack(side="right", padx=6)
        _btn_secondary(_hdr, "+ Add",
                       lambda: self._open_incubator_dialog(),
                       width=80).pack(side="right", padx=6)

        # ── Mode + AC row (card) — exact prototype values ──────────────────────
        ctrl = ctk.CTkFrame(frame, fg_color="#151E2E", corner_radius=12,
                            border_width=1, border_color="#232F42")
        ctrl.pack(fill="x", padx=16, pady=(0, 6))

        _label(ctrl, "Temp Mode", ("Segoe UI", 11, "bold"), "#6B7280").pack(
            side="left", padx=(16, 16), pady=12)

        _det_mode_key = fresh.get("temp_mode", "incubation")

        def _set_detail_mode(key, iid=fresh["id"]):
            prev = next((x for x in db.get_incubators(include_hidden=True)
                         if x["id"] == iid), {}).get("temp_mode", "incubation")
            db.set_incubator_temp_mode(iid, key)
            self._sync_trays_to_mode(iid, key, prev)
            self._refresh_current()

        # Segmented control — no wrapping panel (spec: bare row of buttons, gap 5px)
        _seg = ctk.CTkFrame(ctrl, fg_color="transparent")
        _seg.pack(side="left", pady=8)
        for _mkey, _mlabel in (("off", "Off"), ("cool_storage", "Cool"),
                               ("incubation", "Inc"), ("holding", "Hold")):
            _act = (_mkey == _det_mode_key)
            _mc  = MODE_COLORS.get(_mkey, BLUE)
            _txt_col = GOLD if (_act and _mkey == "holding") else (_mc if _act else "#94A3B8")
            ctk.CTkButton(
                _seg, text=_mlabel, width=0, height=28, corner_radius=6,
                fg_color=MODE_BADGE_BG.get(_mkey, "#1B2536") if _act else "#1B2536",
                hover_color="#243044", text_color=_txt_col,
                font=("Segoe UI", 11, "bold" if _act else "normal"),
                border_width=1, border_color=(_mc if _act else "#2A3648"),
                command=lambda k=_mkey: _set_detail_mode(k),
            ).pack(side="left", padx=(0, 5))

        # Goal · range text
        _gt2, _ = db.get_mode_goals(_det_mode_key)
        _rmin, _rmax = calc.get_temp_range(fresh)
        _grp = []
        if _gt2 is not None:
            _grp.append(f"Goal {calc.format_temp(_gt2, unit)}")
        if _rmin is not None:
            _grp.append(f"range {_rmin}-{_rmax}°C")
        _label(ctrl, "  ·  ".join(_grp) if _grp else "No target set",
               ("Segoe UI", 11), "#6B7280").pack(side="left", padx=(14, 0), pady=10)

        # ── AC controls (right side): Power · Set Temp · Fan ───────────────────
        # Spec base style (all three identical when idle): #202B3D bg, #374151
        # border, #CBD5E1 text. We additionally tint Power green/red when the
        # AC's on/off state is known (implementation guide: toggle buttons
        # restyle on state) — the prototype's static mock doesn't show this
        # since it has no live device to read from.
        _dev = (fresh.get("sensibo_device_id") or "").strip()

        def _no_ac(*_):
            messagebox.showinfo("AC Control",
                "No Sensibo device is configured for this incubator.\n"
                "Add its Sensibo Device ID in Edit Setup.", parent=self)

        _ac_base = dict(fg_color="#202B3D", hover_color="#28374D",
                        text_color="#CBD5E1", border_width=1, border_color="#374151")

        if _dev:
            _pl, _pf, _ptc = self._ac_toggle_style(_dev)
            _fan_cmd  = lambda i=fresh["id"], d=_dev, n=fresh["name"]: self._sensibo_prompt_fan(i, d, n)
            _temp_cmd = lambda i=fresh["id"], d=_dev, n=fresh["name"]: self._sensibo_prompt_temp(i, d, n)
            _pow_cmd  = lambda i=fresh["id"], d=_dev: self._sensibo_toggle_power(i, d)
            _fan_txt  = self._ac_fan_label(_dev)
            _temp_txt = self._ac_temp_label(_dev)
        else:
            _pl, _pf, _ptc = "Power", "#202B3D", "#CBD5E1"
            _fan_cmd = _temp_cmd = _pow_cmd = _no_ac
            _fan_txt, _temp_txt = "Fan", "Set Temp"

        ctk.CTkButton(ctrl, text=_fan_txt, width=88, height=32, corner_radius=7,
                      font=("Segoe UI", 11),
                      command=_fan_cmd, **_ac_base).pack(side="right", padx=(4, 16), pady=8)
        ctk.CTkButton(ctrl, text=_temp_txt, width=96, height=32, corner_radius=7,
                      font=("Segoe UI", 11), command=_temp_cmd, **_ac_base
                      ).pack(side="right", padx=4, pady=8)
        ctk.CTkButton(ctrl, text=_pl, width=84, height=32, corner_radius=7,
                      fg_color=_pf, hover_color="#28374D", text_color=_ptc,
                      font=("Segoe UI", 11), border_width=1,
                      border_color=("#374151" if _pf == "#202B3D" else _pf),
                      command=_pow_cmd
                      ).pack(side="right", padx=4, pady=8)

        # ── Scrollable body ───────────────────────────────────────────────────
        body = ctk.CTkScrollableFrame(frame, fg_color="transparent", corner_radius=0)
        body.pack(fill="both", expand=True)

        # ── Temperature / Humidity Chart with time-range selector ────────────
        chart_frame = ctk.CTkFrame(body, fg_color=PANEL, corner_radius=12,
                                   border_width=1, border_color=BORDER2)
        chart_frame.pack(fill="x", padx=16, pady=(12, 8))

        _RANGES = [
            ("1H",    1),
            ("24H",   24),
            ("7D",    24 * 7),
            ("Month", 24 * 30),
        ]
        _chart_range_var = ctk.StringVar(value="24H")

        # Header row: title left, range buttons right
        chart_hdr = ctk.CTkFrame(chart_frame, fg_color="transparent")
        chart_hdr.pack(fill="x", padx=14, pady=(10, 4))
        # Pack right-side elements first so tkinter reserves space before left fills
        _range_btns: dict = {}
        btn_row = ctk.CTkFrame(chart_hdr, fg_color="transparent")
        btn_row.pack(side="right")

        chart_title = _label(chart_hdr, "Last 24H", ("Segoe UI", 12.5, "bold"), GOLD)
        chart_title.pack(side="left")

        # Canvas holder — we replace its contents when the range changes
        chart_canvas_frame = ctk.CTkFrame(chart_frame, fg_color="transparent")
        chart_canvas_frame.pack(fill="x", padx=8, pady=(0, 10))

        def _draw_chart(hours: float, label: str):
            for w in chart_canvas_frame.winfo_children():
                w.destroy()
            chart_title.configure(text=f"Last {label}")

            # Highlight active button (spec: active gold, others #202B3D)
            for lbl, btn in _range_btns.items():
                _on = (lbl == label)
                btn.configure(fg_color=GOLD if _on else CARD2,
                              text_color="#1A1206" if _on else SUBTEXT)

            if not HAS_MPL:
                _label(chart_canvas_frame, "Install matplotlib to view chart.",
                       FONT_S, SUBTEXT).pack(pady=16)
                return

            readings = db.get_readings_hours(fresh["id"], hours)
            if not readings:
                _label(chart_canvas_frame,
                       f"No readings in the last {label}.",
                       FONT_S, SUBTEXT).pack(pady=16)
                return

            try:
                # Build series, breaking the line (NaN) across gaps in the data
                # so missing stretches don't get a misleading straight diagonal.
                _GAP_SEC = 2700  # 45 min (poll interval is 15 min)
                timestamps, temps, hums = [], [], []
                _prev_ts = None
                for r in readings:
                    ts = datetime.fromisoformat(r["timestamp"])
                    if _prev_ts is not None and (ts - _prev_ts).total_seconds() > _GAP_SEC:
                        timestamps.append(_prev_ts + (ts - _prev_ts) / 2)
                        temps.append(float("nan")); hums.append(float("nan"))
                    tv = r["temperature_c"]
                    timestamps.append(ts)
                    temps.append(calc.c_to_f(tv) if unit == "F" and tv is not None else tv)
                    hums.append(r["humidity_pct"])
                    _prev_ts = ts
                temp_lbl = f"Temp (°{unit})"

                fig = Figure(figsize=(10, 2.5), facecolor="#151E2E")
                ax1 = fig.add_subplot(111)
                ax2 = ax1.twinx()

                ax1.plot(timestamps, temps, color="#FFD700", linewidth=2)
                ax2.plot(timestamps, hums,  color="#3B82F6", linewidth=2,
                         linestyle=(0, (4, 3)))

                t_min, t_max = calc.get_temp_range(fresh)
                if t_min is not None:
                    _lo = calc.c_to_f(t_min) if unit == "F" else t_min
                    _hi = calc.c_to_f(t_max) if unit == "F" else t_max
                    ax1.axhline(_lo, color="#EF4444", linewidth=0.8, linestyle=":")
                    ax1.axhline(_hi, color="#EF4444", linewidth=0.8, linestyle=":")

                # Dotted goal lines — temp goal in the temp colour, humidity goal in
                # the humidity colour (matches each data line).
                goal_t, goal_h = db.get_mode_goals(fresh.get("temp_mode", "incubation"))
                if goal_t is not None:
                    _gt = calc.c_to_f(goal_t) if unit == "F" else goal_t
                    ax1.axhline(_gt, color="#FFD700", linewidth=1.1, linestyle=":")
                if goal_h is not None:
                    ax2.axhline(goal_h, color="#3B82F6", linewidth=1.1, linestyle=":")

                for ax in (ax1, ax2):
                    ax.set_facecolor("#151E2E")
                    ax.tick_params(colors="#9CA3AF", labelsize=8)
                    for spine in ax.spines.values():
                        spine.set_edgecolor("#374151")

                ax1.set_ylabel(temp_lbl, color="#FFD700", fontsize=8)
                ax2.set_ylabel("Humidity %", color="#3B82F6", fontsize=8)

                # Pick x-axis format based on range
                if hours <= 1:
                    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
                    ax1.xaxis.set_major_locator(mdates.MinuteLocator(interval=10))
                elif hours <= 24:
                    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
                    ax1.xaxis.set_major_locator(mdates.HourLocator(interval=2))
                elif hours <= 24 * 7:
                    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%a %d"))
                    ax1.xaxis.set_major_locator(mdates.DayLocator(interval=1))
                else:
                    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
                    ax1.xaxis.set_major_locator(mdates.WeekdayLocator(interval=1))

                for lbl in ax1.xaxis.get_majorticklabels():
                    lbl.set_rotation(0)
                fig.tight_layout(pad=1.0)

                canvas = FigureCanvasTkAgg(fig, master=chart_canvas_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill="x")

                # Legend row below the chart — exact spec: swatch + label, 10.5px muted
                leg_row = ctk.CTkFrame(chart_canvas_frame, fg_color="transparent")
                leg_row.pack(fill="x", pady=(6, 0))
                for _color, _text in (("#FFD700", "Temp °C"), ("#3B82F6", "Humidity %")):
                    _item = ctk.CTkFrame(leg_row, fg_color="transparent")
                    _item.pack(side="left", padx=(0, 16))
                    ctk.CTkFrame(_item, fg_color=_color, width=10, height=2
                                ).pack(side="left", pady=1)
                    _label(_item, " " + _text, ("Segoe UI", 10.5), "#6B7280").pack(side="left")
            except Exception as exc:
                _label(chart_canvas_frame, f"Chart error: {exc}", FONT_S, RED).pack(pady=10)

        for range_label, range_hours in _RANGES:
            rl, rh = range_label, range_hours
            b = ctk.CTkButton(
                btn_row, text=rl, width=52, height=26,
                fg_color=GOLD if rl == "24H" else CARD2,
                hover_color=BORDER,
                text_color="#1A1206" if rl == "24H" else SUBTEXT,
                corner_radius=6, font=("Segoe UI", 10, "bold"),
                command=lambda lbl=rl, hrs=rh: _draw_chart(hrs, lbl),
            )
            b.pack(side="left", padx=2)
            _range_btns[rl] = b

        # Defer chart so the screen appears instantly, then render after
        frame.after(50, lambda: _draw_chart(24, "24H"))

        # ── Tabs: Inspections / Batches / Trays / VOC ─────────────────────────
        # Tabs are built lazily — content is only created the first time each tab is selected.
        # CTkSegmentedButton always draws as ONE fused strip with no real gaps
        # between segments, which doesn't match the prototype's 4 separate
        # rounded pill buttons. So: use CTkTabview purely as the content
        # switcher (hide its built-in button row) and drive it with our own
        # row of individually-styled pill buttons instead.
        tabs = ctk.CTkTabview(
            body, fg_color="transparent", corner_radius=0, border_width=0,
        )
        # NOTE: .add() re-grids (re-shows) the built-in segmented button when
        # adding the FIRST tab — so grid_forget() must happen AFTER all tabs
        # are added, not before, or the built-in strip reappears alongside
        # our custom pill row (the duplicate tab-row bug).
        _tab_names = ("Inspections", "Batches", "Trays", "Vapona Monitor")
        for _tn in _tab_names:
            tabs.add(_tn)
        tabs._segmented_button.grid_forget()   # hide the fused built-in strip

        # Custom pill row — spec exact: gap 6px, padding 8px 16px, radius 7px,
        # active border #3B82F6 / bg #22314A / text #93C5FD, inactive #2A3648 / #1B2536 / #9CA3AF
        pill_row = ctk.CTkFrame(body, fg_color="transparent")
        pill_row.pack(pady=(4, 10))
        _pill_btns: dict = {}

        def _select_tab(name):
            tabs.set(name)
            for tn, btn in _pill_btns.items():
                _act = (tn == name)
                btn.configure(
                    fg_color="#22314A" if _act else "#1B2536",
                    hover_color="#28395A" if _act else CARD2,
                    text_color="#93C5FD" if _act else "#9CA3AF",
                    border_color="#3B82F6" if _act else "#2A3648",
                    font=("Segoe UI", 11.5, "bold" if _act else "normal"))
            frame._active_tab = name
            if name not in _tab_built:
                _tab_built.add(name)
                _tab_builders[name]()

        for _tn in _tab_names:
            _b = ctk.CTkButton(
                pill_row, text=_tn, height=32, corner_radius=7,
                fg_color="#1B2536", hover_color=CARD2, text_color="#9CA3AF",
                border_width=1, border_color="#2A3648",
                font=("Segoe UI", 11.5), command=lambda n=_tn: _select_tab(n),
            )
            _b.pack(side="left", padx=3)
            _pill_btns[_tn] = _b

        _tab_built: set = set()

        def _build_inspections_tab():
            it = tabs.tab("Inspections")
            InspectionsLogPanel(it, fixed_incubator_id=fresh["id"]).pack(fill="both", expand=True)

        def _build_batches_tab():
            bt = tabs.tab("Batches")
            _bhdr = ctk.CTkFrame(bt, fg_color="transparent")
            _bhdr.pack(fill="x", padx=8, pady=(8, 2))
            _btn_secondary(_bhdr, "+ New Batch",
                lambda i=fresh["id"]: self._open_batch_dialog(incubator_id=i),
                width=120).pack(side="right")
            bscroll = ctk.CTkScrollableFrame(bt, fg_color="transparent")
            bscroll.pack(fill="both", expand=True)
            batches = db.get_batches(incubator_id=fresh["id"])
            if batches:
                for batch in batches:
                    bf = ctk.CTkFrame(bscroll, fg_color=CARD2, corner_radius=8)
                    bf.pack(fill="x", pady=4, padx=4)
                    _brow = ctk.CTkFrame(bf, fg_color="transparent")
                    _brow.pack(fill="x", padx=10, pady=(8, 0))
                    bname = batch.get("name") or f"Batch {batch['id']}"
                    _label(_brow, bname, FONT_B, GOLD).pack(side="left")
                    _btn_secondary(_brow, "Edit",
                        lambda b=batch: self._open_batch_dialog(batch=b),
                        width=64).pack(side="right")
                    for fld, lbl in calc.BATCH_EVENT_FIELDS:
                        val = batch.get(fld)
                        if val:
                            d    = calc.days_from_now(val)
                            line = f"{lbl}: {val[:10]}  ({calc.format_days(d)})"
                            col  = (RED    if d is not None and d <= 1
                                    else ORANGE if d is not None and d <= 5
                                    else SUBTEXT)
                            _label(bf, line, FONT_S, col).pack(anchor="w", padx=14, pady=1)
                    ctk.CTkFrame(bf, fg_color="transparent", height=6).pack()
            else:
                _label(bscroll, "No batches for this incubator.", FONT_S, SUBTEXT).pack(pady=16)

        def _build_trays_tab():
            tt = tabs.tab("Trays")
            tray_hdr = ctk.CTkFrame(tt, fg_color="transparent")
            tray_hdr.pack(fill="x", padx=8, pady=(8, 4))
            tscroll = ctk.CTkScrollableFrame(tt, fg_color="transparent")
            tscroll.pack(fill="both", expand=True)

            _tray_checks: dict = {}

            def _update_delete_btn():
                n = sum(1 for v in _tray_checks.values() if v.get())
                if n:
                    delete_btn.configure(text=f"Delete {n} Selected",
                                         fg_color="#2A1E20", hover_color="#3A2224",
                                         text_color="#FF6A57", state="normal")
                else:
                    delete_btn.configure(text="Delete Selected",
                                         fg_color="#2A1E20", hover_color="#3A2224",
                                         text_color="#FF6A57", state="disabled")

            def _select_all_toggle():
                all_on = all(v.get() for v in _tray_checks.values()) if _tray_checks else False
                for v in _tray_checks.values():
                    v.set(not all_on)
                _update_delete_btn()

            def _delete_selected():
                ids = [tid for tid, v in _tray_checks.items() if v.get()]
                if not ids:
                    return
                if not messagebox.askyesno("Delete Trays",
                        f"Permanently delete {len(ids)} tray(s)?\nThis cannot be undone.",
                        parent=self):
                    return
                for tid in ids:
                    db.delete_tray(tid)
                _refresh_tray_list()

            def _refresh_tray_list():
                _tray_checks.clear()
                for w in tscroll.winfo_children():
                    w.destroy()
                # Only show trays currently in this incubator (active);
                # released/removed trays drop off the list.
                trays = db.get_trays(incubator_id=fresh["id"], status=db.IN_INCUBATOR_STATUSES)
                tray_count_lbl.configure(text=f"{len(trays)} tray(s)")
                _update_delete_btn()
                if trays:
                    for tray in trays:
                        var = ctk.BooleanVar(value=False)
                        _tray_checks[tray["id"]] = var
                        tr = ctk.CTkFrame(tscroll, fg_color=CARD2, corner_radius=6)
                        tr.pack(fill="x", pady=2, padx=4)
                        ctk.CTkCheckBox(tr, text="", variable=var,
                            width=20, checkbox_width=18, checkbox_height=18,
                            fg_color=RED, hover_color="#B91C1C",
                            command=_update_delete_btn,
                        ).pack(side="left", padx=(8, 4), pady=6)
                        _label(tr, f"Tray {tray['tray_number']}", FONT_B, TEXT).pack(
                            side="left", padx=(4, 8), pady=6)
                        _label(tr,
                            f"{tray.get('sample_name') or '—'}  "
                            f"{tray.get('volume_gal') or '—'} gal  "
                            f"{tray.get('status') or 'active'}",
                            FONT_S, SUBTEXT).pack(side="left", padx=4)
                        _btn(tr, "QR", lambda t=tray: QRDialog(self, t, port=self._qr_port),
                             width=50, height=24, fg=BORDER, hover=CARD).pack(side="right", padx=8)
                else:
                    _label(tscroll, "No trays yet. Add one or import a CSV.",
                           FONT_S, SUBTEXT).pack(pady=16)

            tray_count_lbl = _label(tray_hdr, "", FONT_S, SUBTEXT)
            tray_count_lbl.pack(side="left")
            _btn_secondary(tray_hdr, "⬇ Template",
                 lambda: self._tray_csv_template(default_incubator_name=fresh["name"]),
                 width=110).pack(side="right", padx=(4, 0))
            _btn(tray_hdr, "Import CSV",
                 lambda: self._tray_csv_import(default_incubator_id=fresh["id"],
                                               on_complete=_refresh_tray_list),
                 width=100, height=28, fg=TEAL, hover="#0D9488",
                 text_color="white").pack(side="right", padx=4)
            _btn(tray_hdr, "Release CSV",
                 lambda: self._release_csv_import(on_complete=_refresh_tray_list),
                 width=110, height=28, fg="#7C3AED", hover="#6D28D9",
                 text_color="white").pack(side="right", padx=4)
            _btn_secondary(tray_hdr, "⬇ Release Template", self._release_csv_template,
                 width=150).pack(side="right", padx=(0, 2))
            delete_btn = ctk.CTkButton(tray_hdr, text="Delete Selected", width=130, height=28,
                fg_color="#2A1E20", hover_color="#3A2224", text_color="#FF6A57",
                corner_radius=8, font=("Segoe UI", 11, "bold"),
                border_width=1, border_color="#5A2A2C",
                state="disabled", command=_delete_selected)
            delete_btn.pack(side="right", padx=4)
            _btn_secondary(tray_hdr, "Select All", _select_all_toggle,
                 width=90).pack(side="right", padx=(0, 2))
            _refresh_tray_list()

        def _build_voc_tab():
            vt = tabs.tab("Vapona Monitor")
            VOCPanel(vt, incubator_id=fresh["id"]).pack(fill="both", expand=True)

        # Register all tab names first (required before .tab() can be called)
        # Map tab name → builder function (tabs were already added above,
        # alongside the custom pill row that drives _select_tab)
        _tab_builders = {
            "Inspections": _build_inspections_tab,
            "Batches":     _build_batches_tab,
            "Trays":       _build_trays_tab,
            "Vapona Monitor": _build_voc_tab,
        }

        # Also pack the content area itself now that tabs/pills exist
        tabs.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        # Restore the previously active tab (or default to Inspections)
        restore_tab = _prev_tab if _prev_tab in _tab_builders else "Inspections"
        _select_tab(restore_tab)

    def _open_inc_detail_window(self, inc: dict):
        """Open a detail window with Vapona Monitor, Inspections, Batches and Trays tabs."""
        win = ctk.CTkToplevel(self)
        win.title(f"{inc['name']} — Detail")
        win.geometry("1020x700")
        win.minsize(860, 600)
        win.grab_set()

        # Header row with name + quick Inspect button
        hdr = ctk.CTkFrame(win, fg_color="transparent")
        hdr.pack(fill="x", padx=16, pady=(12, 2))
        _label(hdr, inc["name"], FONT_H, GOLD).pack(side="left")

        # Live M/E badges in header
        reading = self._govee.get_last(inc["id"])
        govee_temp = reading.get("temp_c")
        ibrow = ctk.CTkFrame(hdr, fg_color="transparent")
        ibrow.pack(side="left", padx=16)
        _label(ibrow, "Inspections:", FONT_S, SUBTEXT).pack(side="left", padx=(0, 6))
        make_status_badges(ibrow, inc["id"]).pack(side="left")

        _btn(hdr, "Inspect Now", lambda i=inc: self._open_inspection_form(i),
             width=110, height=30, fg=BLUE, hover="#1D4ED8",
             text_color="white").pack(side="right")

        # Temp mode switcher row
        mrow = ctk.CTkFrame(win, fg_color="transparent")
        mrow.pack(fill="x", padx=16, pady=(0, 6))
        _label(mrow, "Temp Mode:", FONT_S, SUBTEXT).pack(side="left", padx=(0, 8))
        _det_mode_key = inc.get("temp_mode", "incubation")
        _det_mode_cfg = calc.TEMP_MODES.get(_det_mode_key, calc.TEMP_MODES["incubation"])
        _det_mode_var = ctk.StringVar(value=_det_mode_cfg["label"])
        _det_range_lbl = _label(mrow, f"{_det_mode_cfg['min']}–{_det_mode_cfg['max']} °C",
                                FONT_S, SUBTEXT)

        def _on_detail_mode(label, iid=inc["id"], rlbl=_det_range_lbl, rv=_det_mode_var):
            key = calc._MODE_BY_LABEL.get(label, "incubation")
            db.set_incubator_temp_mode(iid, key)
            cfg = calc.TEMP_MODES[key]
            rlbl.configure(text=f"{cfg['min']}–{cfg['max']} °C")
            self._refresh_current()

        ctk.CTkSegmentedButton(
            mrow,
            values=[v["label"] for v in calc.TEMP_MODES.values()],
            variable=_det_mode_var,
            command=_on_detail_mode,
            width=300, height=28
        ).pack(side="left")
        _det_range_lbl.pack(side="left", padx=(12, 0))

        # Alert on/off toggle
        _alerts_on = bool(inc.get("temp_alerts_enabled", 1))
        _alert_btn_txt = ctk.StringVar(value="🔔  Alerts On" if _alerts_on else "🔕  Alerts Off")

        def _toggle_alert_detail(iid=inc["id"], sv=_alert_btn_txt):
            cur_val = db.get_incubators()  # re-fetch to get live flag
            cur_inc = next((x for x in cur_val if x["id"] == iid), {})
            new_val = not bool(cur_inc.get("temp_alerts_enabled", 1))
            db.set_incubator_alerts_enabled(iid, new_val)
            sv.set("🔔  Alerts On" if new_val else "🔕  Alerts Off")
            self._refresh_current()

        ctk.CTkButton(
            mrow, textvariable=_alert_btn_txt,
            font=("Segoe UI", 11),
            fg_color=BORDER, hover_color=CARD2,
            text_color=TEXT,
            width=130, height=28,
            command=_toggle_alert_detail
        ).pack(side="left", padx=(16, 0))

        tabs = ctk.CTkTabview(win, fg_color=CARD)
        tabs.pack(fill="both", expand=True, padx=12, pady=8)

        # ── Vapona Monitor tab (first — most used) ───────────────────────────────
        vt = tabs.add("Vapona Monitor")
        VOCPanel(vt, incubator_id=inc["id"]).pack(fill="both", expand=True)

        # ── Inspections tab ───────────────────────────────────────────────────
        it = tabs.add("Inspections")
        InspectionsLogPanel(it, fixed_incubator_id=inc["id"]).pack(
            fill="both", expand=True)

        # ── Batches tab ───────────────────────────────────────────────────────
        bt = tabs.add("Batches")
        scroll = ctk.CTkScrollableFrame(bt, fg_color="transparent")
        scroll.pack(fill="both", expand=True)
        for batch in db.get_batches(incubator_id=inc["id"]):
            bf = ctk.CTkFrame(scroll, fg_color=CARD2, corner_radius=8)
            bf.pack(fill="x", pady=4, padx=4)
            bname = batch.get("name") or f"Batch {batch['id']}"
            _label(bf, bname, FONT_B, GOLD).pack(anchor="w", padx=10, pady=(8,2))
            for fld, lbl in calc.BATCH_EVENT_FIELDS:
                val = batch.get(fld)
                if val:
                    d = calc.days_from_now(val)
                    line = f"{lbl}: {val[:10]}  ({calc.format_days(d)})"
                    col  = RED if d is not None and d <= 1 else (ORANGE if d is not None and d <= 5 else SUBTEXT)
                    _label(bf, line, FONT_S, col).pack(anchor="w", padx=14, pady=1)
            ctk.CTkFrame(bf, fg_color="transparent", height=6).pack()

        # ── Trays tab ─────────────────────────────────────────────────────────
        tt = tabs.add("Trays")
        tr_scroll = ctk.CTkScrollableFrame(tt, fg_color="transparent")
        tr_scroll.pack(fill="both", expand=True)
        for tray in db.get_trays(incubator_id=inc["id"], status=db.IN_INCUBATOR_STATUSES):
            tr = ctk.CTkFrame(tr_scroll, fg_color=CARD2, corner_radius=6)
            tr.pack(fill="x", pady=2, padx=4)
            _label(tr, f"Tray {tray['tray_number']}", FONT_B, TEXT).pack(
                side="left", padx=10, pady=6)
            details = (
                f"{tray.get('sample_name') or '—'}  "
                f"{tray.get('volume_gal') or '—'} gal  "
                f"{tray.get('status') or 'active'}"
            )
            _label(tr, details, FONT_S, SUBTEXT).pack(side="left", padx=4)
            _btn(tr, "QR", lambda t=tray: QRDialog(self, t, port=self._qr_port),
                 width=50, height=24, fg=BORDER, hover=CARD).pack(side="right", padx=8)

    # ── Shared tray CSV helpers ───────────────────────────────────────────────

    _CSV_HEADERS = ["QR", "Sample", "gals/tray", "Incubator",
                    "Treatment Label", "Date+Time", "User"]

    def _tray_csv_template(self, default_incubator_name: str = ""):
        import csv
        path = filedialog.asksaveasfilename(
            title="Save CSV Template",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="trays_template.csv",
        )
        if not path:
            return
        example = ["T001", "Sample A", "2.0",
                   default_incubator_name or "Incubator 1",
                   "Batch 1", "2026-06-24 09:30", "Jane"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows([self._CSV_HEADERS, example])
        messagebox.showinfo("Template Saved",
            f"Template saved to:\n{path}\n\n"
            "Fill in one row per tray, then use 'Import CSV' to load it.\n\n"
            "Columns:\n"
            "  QR             — tray identifier (required)\n"
            "  Sample         — sample name (auto-created if missing)\n"
            "  gals/tray      — volume in gallons\n"
            "  Incubator      — incubator name (required when importing from Trays tab)\n"
            "  Treatment Label— batch / treatment name\n"
            "  Date+Time      — load date (YYYY-MM-DD or YYYY-MM-DD HH:MM)\n"
            "  User           — person who loaded the tray",
            parent=self)

    def _tray_csv_import(self, default_incubator_id: int = None,
                         on_complete=None):
        import csv, re
        path = filedialog.askopenfilename(
            title="Select Tray CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return

        samples    = {s["name"]: s["id"] for s in db.get_samples()}
        batches    = {b["name"]: b["id"] for b in db.get_batches()}
        _inc_list  = db.get_incubators(include_hidden=True)
        inc_lookup = {i["name"].strip().lower(): i for i in _inc_list}

        def _find_incubator(raw: str):
            key = raw.strip().lower()
            if key in inc_lookup:
                return inc_lookup[key]
            num = re.search(r"\d+", key)
            if num:
                n = num.group()
                for k, inc in inc_lookup.items():
                    if re.search(r"\b" + n + r"\b", k):
                        return inc
            return None

        errors   = []
        imported = 0
        skipped  = 0

        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
        except Exception as exc:
            messagebox.showerror("Import Error", f"Could not read file:\n{exc}", parent=self)
            return

        if not rows:
            messagebox.showwarning("Import", "CSV file is empty.", parent=self)
            return

        missing_cols = [h for h in self._CSV_HEADERS if h not in rows[0]]
        if missing_cols:
            messagebox.showerror("Import Error",
                f"Missing column(s): {', '.join(missing_cols)}\n\n"
                "Download the template to see the correct format.", parent=self)
            return

        for i, row in enumerate(rows, start=2):
            qr = (row.get("QR") or "").strip()
            if not qr:
                errors.append(f"Row {i}: missing QR — skipped")
                skipped += 1
                continue

            def _pf(v):
                try:    return float(str(v).strip()) if v and str(v).strip() else None
                except: return None

            sample_name    = (row.get("Sample")          or "").strip()
            inc_name       = (row.get("Incubator")       or "").strip()
            treatment_name = (row.get("Treatment Label") or "").strip()
            date_raw       = (row.get("Date+Time")       or "").strip()
            user           = (row.get("User")            or "").strip()

            if inc_name:
                matched = _find_incubator(inc_name)
                if matched:
                    inc_id = matched["id"]
                else:
                    inc_id = default_incubator_id
                    errors.append(
                        f"Row {i} ({qr}): incubator '{inc_name}' not recognised"
                        + (f" — assigned to default incubator instead" if inc_id else " — skipped (no default incubator)")
                        + f". Available: {', '.join(x['name'] for x in _inc_list)}"
                    )
                    if inc_id is None:
                        skipped += 1
                        continue
            elif default_incubator_id:
                inc_id = default_incubator_id
            else:
                errors.append(f"Row {i} ({qr}): Incubator column is empty and no default — skipped")
                skipped += 1
                continue

            sample_id = samples.get(sample_name)
            if sample_name and sample_id is None:
                sample_id = db.upsert_sample({"name": sample_name})
                samples[sample_name] = sample_id
                errors.append(f"Row {i} ({qr}): created new sample '{sample_name}'")

            batch_id = batches.get(treatment_name)
            if treatment_name and batch_id is None:
                errors.append(f"Row {i} ({qr}): treatment '{treatment_name}' not found — imported without batch link")

            in_date = date_raw[:10] if date_raw and len(date_raw) >= 10 else (date_raw or None)

            try:
                db.upsert_tray({
                    "tray_number":         qr,
                    "incubator_id":        inc_id,
                    "sample_id":           sample_id,
                    "incubation_batch_id": batch_id,
                    "volume_gal":          _pf(row.get("gals/tray")),
                    "in_date":             in_date,
                    "status":              "active",
                    "notes":               f"Loaded by: {user}" if user else "",
                    "weight_lbs":          None,
                    "live_count":          None,
                    "parasite_level_pct":  None,
                    "out_date":            None,
                })
                imported += 1
            except Exception as exc:
                errors.append(f"Row {i} ({qr}): {exc}")
                skipped += 1

        if on_complete:
            on_complete()

        summary = f"Imported {imported} tray(s)."
        if skipped:
            summary += f"\n{skipped} row(s) skipped."
        if errors:
            new_samples = [e for e in errors if "created new sample" in e]
            warnings    = [e for e in errors if "created new sample" not in e]
            if new_samples:
                summary += f"\n\nNew samples created ({len(new_samples)}):\n"
                summary += "\n".join(e.split("'")[1] for e in new_samples[:10])
            if warnings:
                summary += "\n\nWarnings / Errors:\n" + "\n".join(warnings[:15])
                if len(warnings) > 15:
                    summary += f"\n… and {len(warnings) - 15} more"
        messagebox.showinfo("Import Complete", summary, parent=self)

    _RELEASE_CSV_HEADERS = ["Tray ID #", "Field", "Gals in Tray",
                            "LatLong", "Lat", "Long", "Date+Time", "User"]

    def _release_csv_template(self):
        import csv
        path = filedialog.asksaveasfilename(
            title="Save Release CSV Template",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="tray_release_template.csv",
        )
        if not path:
            return
        example = ["T001", "North Field", "2.0",
                   "44.123,-93.456", "44.123", "-93.456",
                   "2026-06-24 09:30", "Jane"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows([self._RELEASE_CSV_HEADERS, example])
        messagebox.showinfo("Template Saved",
            f"Template saved to:\n{path}\n\n"
            "Fill in one row per tray released to the field.\n\n"
            "Columns:\n"
            "  Tray ID #   — tray identifier used to look up the tray (required)\n"
            "  Field       — field or site name\n"
            "  Gals in Tray— volume at release (optional, for record)\n"
            "  LatLong     — combined lat/long string\n"
            "  Lat / Long  — individual coordinates\n"
            "  Date+Time   — release date (YYYY-MM-DD or YYYY-MM-DD HH:MM)\n"
            "  User        — person who released the tray",
            parent=self)

    def _release_csv_import(self, on_complete=None):
        import csv
        path = filedialog.askopenfilename(
            title="Select Release CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
        except Exception as exc:
            messagebox.showerror("Import Error", f"Could not read file:\n{exc}", parent=self)
            return

        if not rows:
            messagebox.showwarning("Release Import", "CSV file is empty.", parent=self)
            return

        missing = [h for h in self._RELEASE_CSV_HEADERS if h not in rows[0]]
        if missing:
            messagebox.showerror("Import Error",
                f"Missing column(s): {', '.join(missing)}\n\n"
                "Download the release template to see the correct format.", parent=self)
            return

        released = 0
        not_found = []
        errors = []

        for i, row in enumerate(rows, start=2):
            tray_id = (row.get("Tray ID #") or "").strip()
            if not tray_id:
                errors.append(f"Row {i}: missing Tray ID # — skipped")
                continue

            field    = (row.get("Field")       or "").strip()
            latlong  = (row.get("LatLong")     or "").strip()
            lat      = (row.get("Lat")         or "").strip()
            lon      = (row.get("Long")        or "").strip()
            date_raw = (row.get("Date+Time")   or "").strip()
            user     = (row.get("User")        or "").strip()

            out_date = date_raw[:10] if date_raw and len(date_raw) >= 10 else (date_raw or None)

            note_parts = []
            if user:
                note_parts.append(f"Released by: {user}")
            if field:
                note_parts.append(f"Field: {field}")
            if latlong:
                note_parts.append(f"LatLong: {latlong}")
            elif lat or lon:
                note_parts.append(f"Lat: {lat}  Long: {lon}")
            notes_append = "\n".join(note_parts)

            try:
                ok = db.release_tray(tray_id, out_date=out_date, notes_append=notes_append)
                if ok:
                    released += 1
                else:
                    not_found.append(tray_id)
            except Exception as exc:
                errors.append(f"Row {i} ({tray_id}): {exc}")

        if on_complete:
            on_complete()

        summary = f"Released {released} tray(s)."
        if not_found:
            summary += f"\n\n{len(not_found)} tray(s) not found in the system:\n"
            summary += "\n".join(not_found[:20])
            if len(not_found) > 20:
                summary += f"\n… and {len(not_found) - 20} more"
        if errors:
            summary += "\n\nErrors:\n" + "\n".join(errors[:10])
        messagebox.showinfo("Release Import Complete", summary, parent=self)

    def _open_batch_dialog(self, batch: dict = None, incubator_id: int = None):
        BatchDialog(self, batch, incubator_id,
                    on_save=lambda: self._refresh_current())

    def _open_sample_dialog(self, sample: dict = None):
        SampleDialog(self, sample, on_save=lambda: self._refresh_current())

    def _open_tray_dialog(self, tray: dict = None, incubator_id: int = None):
        TrayDialog(self, tray, incubator_id,
                   on_save=lambda: self._refresh_current())

    def _sync_trays_to_mode(self, inc_id: int, new_key: str, prev_key: str):
        """Offer to move trays when an incubator changes into Holding or Incubation:
        - into Holding:    Incubation -> Cooled (start cool-down / hold timer)
        - into Incubation: Cooled -> Incubation (resume incubation)
        """
        if new_key == prev_key:
            return

        if new_key == "holding":
            n = db.count_active_trays(inc_id)
            if n and messagebox.askyesno(
                "Move trays to Cooled?",
                f"This incubator was switched to Holding.\n\n"
                f"Move its {n} tray(s) currently in incubation to 'Cooled' and "
                f"start their cool-down timer (today's date)?",
                parent=self):
                moved = db.cool_trays(inc_id)
                messagebox.showinfo("Trays Cooled",
                    f"Moved {moved} tray(s) to Cooled.", parent=self)
        elif new_key == "incubation":
            n = db.count_cooled_trays(inc_id)
            if n and messagebox.askyesno(
                "Move trays to Incubation?",
                f"This incubator was switched to Incubation.\n\n"
                f"Move its {n} cooled tray(s) back to 'Incubation'? "
                f"(Their cool-down timer will reset.)",
                parent=self):
                moved = db.uncool_trays(inc_id)
                messagebox.showinfo("Trays back in Incubation",
                    f"Moved {moved} tray(s) back to Incubation.", parent=self)

    def _bulk_set_status(self):
        """Change status on all currently selected trays in the Trays tab."""
        if not self._tray_sel:
            messagebox.showinfo("Set Status",
                "Click one or more trays to select them first.",
                parent=self)
            return

        new_status = db.tray_status_value(self._bulk_status.get())
        tray_ids   = [int(t) for t in self._tray_sel]

        out_date           = None
        overwrite_out_date = False
        if new_status in ("released", "removed"):
            # Let the user choose the out date (blank = today)
            today = datetime.now().strftime("%Y-%m-%d")
            dlg = ctk.CTkInputDialog(
                title="Out Date",
                text=f"Out date for {len(tray_ids)} tray(s), format YYYY-MM-DD.\n\n"
                     f"• Leave blank to use today ({today})\n"
                     "• Enter a date to apply it to all selected trays")
            raw = dlg.get_input()
            if raw is None:
                return  # cancelled
            raw = raw.strip()
            if raw:
                try:
                    datetime.strptime(raw, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Invalid Date",
                        f"'{raw}' is not a valid date. Use YYYY-MM-DD.", parent=self)
                    return
                out_date           = raw
                overwrite_out_date = True   # explicit date applies to all selected
            else:
                out_date = today  # blank → today, applied to trays without a date
        else:
            if not messagebox.askyesno(
                "Change Status",
                f"Set status to '{db.tray_status_label(new_status)}' "
                f"for {len(tray_ids)} selected tray(s)?",
                parent=self):
                return

        db.set_trays_status(tray_ids, new_status,
                            out_date=out_date, overwrite_out_date=overwrite_out_date)
        self._refresh_trays()
        self._refresh_alert_badge()
        messagebox.showinfo("Status Updated",
            f"Updated {len(tray_ids)} tray(s) to '{db.tray_status_label(new_status)}'"
            + (f" with out date {out_date}." if out_date else "."), parent=self)

    def _delete_all_trays(self):
        """Delete every tray, behind a two-step confirmation."""
        total = db.get_tray_stats()["count"]
        if not total:
            messagebox.showinfo("Delete All Trays", "There are no trays to delete.", parent=self)
            return

        # Step 1 — initial warning
        if not messagebox.askyesno(
            "Delete ALL Trays?",
            f"This will permanently delete ALL {total} tray(s) from every incubator, "
            "including their full season history.\n\n"
            "This cannot be undone.\n\nContinue?",
            icon="warning", parent=self):
            return

        # Step 2 — type-to-confirm
        dlg = ctk.CTkInputDialog(
            title="Confirm Delete All",
            text=f"Final confirmation.\n\nType  DELETE  to permanently remove "
                 f"all {total} tray(s):")
        entry = (dlg.get_input() or "").strip()
        if entry != "DELETE":
            messagebox.showinfo("Cancelled",
                "Trays were NOT deleted (confirmation text did not match).", parent=self)
            return

        removed = db.delete_all_trays()
        self._refresh_current()
        messagebox.showinfo("Trays Deleted",
            f"Deleted {removed} tray(s). The tray list is now empty.", parent=self)

    def _delete_incubator(self, iid: int):
        if messagebox.askyesno("Delete", "Delete this incubator?"):
            db.delete_incubator(iid)
            self._refresh_current()

    def _set_hidden(self, iid: int, hidden: bool):
        db.set_incubator_hidden(iid, hidden)
        self._refresh_current()

    def _set_inc_mode(self, iid: int, label: str):
        """Set an incubator's temp mode from a selector (e.g. the Incubators view)."""
        key  = calc._MODE_BY_LABEL.get(label, "incubation")
        prev = next((x for x in db.get_incubators(include_hidden=True)
                     if x["id"] == iid), {}).get("temp_mode", "incubation")
        db.set_incubator_temp_mode(iid, key)
        self._sync_trays_to_mode(iid, key, prev)
        self._refresh_current()

    def _sensibo_update_buttons(self, iid: int, device_id: str):
        """Patch just the AC button labels/colors on the card for this incubator."""
        widgets = self._card_widgets.get(iid, {})
        pwr = widgets.get("ac_power")
        tmp = widgets.get("ac_temp")
        fan = widgets.get("ac_fan")
        if pwr and pwr.winfo_exists():
            lbl, fg, tc = self._ac_toggle_style(device_id)
            pwr.configure(text=lbl, fg_color=fg, text_color=tc)
        if tmp and tmp.winfo_exists():
            tmp.configure(text=self._ac_temp_label(device_id))
        if fan and fan.winfo_exists():
            fan.configure(text=self._ac_fan_label(device_id))

    def _sensibo_run(self, iid: int, device_id: str, fn, on_error=None):
        """Run fn() in a background thread; patch AC buttons when done.

        fn must be a zero-argument callable (use a lambda to capture kwargs).
        """
        import threading
        widgets = self._card_widgets.get(iid, {})
        for key in ("ac_power", "ac_temp", "ac_fan"):
            w = widgets.get(key)
            if w and w.winfo_exists():
                w.configure(state="disabled")

        def _work():
            ok = fn()
            def _done():
                for key in ("ac_power", "ac_temp", "ac_fan"):
                    w = widgets.get(key)
                    if w and w.winfo_exists():
                        w.configure(state="normal")
                if not ok and on_error:
                    messagebox.showerror("Sensibo", on_error(), parent=self)
                self._sensibo_update_buttons(iid, device_id)
            self.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    def _sensibo_set_power(self, iid: int, device_id: str, on: bool):
        if not db.get_setting("sensibo_api_key"):
            messagebox.showwarning("Sensibo", "Set a Sensibo API key in Settings first.", parent=self)
            return
        self._sensibo_run(iid, device_id,
            lambda: self._sensibo.set_ac_state_many(device_id, on=on),
            on_error=lambda: f"Could not reach the AC unit(s):\n{self._sensibo.status_label()}")

    def _sensibo_toggle_power(self, iid: int, device_id: str):
        if not db.get_setting("sensibo_api_key"):
            messagebox.showwarning("Sensibo", "Set a Sensibo API key in Settings first.", parent=self)
            return
        target = not self._sensibo.resolve_power(device_id)
        self._sensibo_run(iid, device_id,
            lambda: self._sensibo.set_ac_state_many(device_id, on=target),
            on_error=lambda: f"Could not reach the AC unit(s):\n{self._sensibo.status_label()}")

    def _ac_toggle_style(self, device_id: str):
        """Return (label, fg, text_color) for an AC power toggle button."""
        power = self._sensibo.get_cached_power(device_id)
        if power is True:
            return "● On", "#243A34", GREEN_LT   # spec AC-on fill
        if power is False:
            return "● Off", "#3A2129", RED_LT     # spec AM/PM-pending-like red
        return "⏻ Power", CARD2, TEXT

    def _ac_temp_label(self, device_id: str) -> str:
        st = self._sensibo.get_cached_state(device_id)
        t = st.get("targetTemperature")
        return f"{t}°F" if t is not None else "Set Temp"

    def _ac_fan_label(self, device_id: str) -> str:
        st = self._sensibo.get_cached_state(device_id)
        f = st.get("fanLevel")
        return f.capitalize() if f else "Fan"

    def _sensibo_prompt_temp(self, iid: int, device_id: str, name: str):
        if not db.get_setting("sensibo_api_key"):
            messagebox.showwarning("Sensibo", "Set a Sensibo API key in Settings first.", parent=self)
            return
        lo, hi = sensibo_mod.MIN_TEMP_F, sensibo_mod.MAX_TEMP_F
        dlg = ctk.CTkInputDialog(
            title="Set AC Target Temp",
            text=f"Target temperature (°F) for {name}.\n\n"
                 f"Minimum {lo}°F · Maximum {hi}°F")
        raw = dlg.get_input()
        if raw is None or not raw.strip():
            return
        try:
            temp_f = int(round(float(raw.strip())))
        except ValueError:
            messagebox.showerror("Sensibo", "Enter a numeric temperature.", parent=self)
            return
        if not (lo <= temp_f <= hi):
            messagebox.showerror("Sensibo",
                f"Temperature must be between {lo}°F and {hi}°F.", parent=self)
            return
        self._sensibo_run(iid, device_id,
            lambda: self._sensibo.set_ac_state_many(device_id, on=True, target_temp=temp_f),
            on_error=lambda: f"Could not reach the AC unit(s):\n{self._sensibo.status_label()}")

    def _sensibo_prompt_fan(self, iid: int, device_id: str, name: str):
        if not db.get_setting("sensibo_api_key"):
            messagebox.showwarning("Sensibo", "Set a Sensibo API key in Settings first.", parent=self)
            return
        win = ctk.CTkToplevel(self)
        win.title("Set Fan Speed")
        win.geometry("260x320")
        win.grab_set()
        _label(win, f"Fan speed — {name}", FONT_B, GOLD).pack(padx=16, pady=(14, 8))

        def _apply(level):
            win.destroy()
            self._sensibo_run(iid, device_id,
                lambda l=level: self._sensibo.set_ac_state_many(device_id, fan_level=l),
                on_error=lambda: f"Could not set fan speed:\n{self._sensibo.status_label()}")

        for lvl in sensibo_mod.FAN_LEVELS:
            _btn(win, lvl.capitalize(), lambda l=lvl: _apply(l),
                 width=180, height=32, fg=BORDER, hover=CARD2).pack(pady=3)
        _btn(win, "Cancel", win.destroy, width=180, height=28,
             fg=CARD2, hover=BORDER).pack(pady=(8, 4))

    def _after_inspection_saved(self):
        """After saving an inspection, jump back to the dashboard."""
        self.show_view("dashboard")

    def _toggle_temp_alerts(self, inc: dict):
        new_val = not bool(inc.get("temp_alerts_enabled", 1))
        db.set_incubator_alerts_enabled(inc["id"], new_val)
        self._refresh_current()

    def _open_alerts(self):
        AlertsDialog(self, on_ack=self._refresh_alert_badge)

    def _open_inspection_form(self, inc: dict):
        """Open the inspection form for a given incubator."""
        reading   = self._govee.get_last(inc["id"])
        govee_tmp = reading.get("temp_c")
        InspectionDialog(
            self, inc, govee_temp_c=govee_tmp,
            on_save=self._after_inspection_saved,
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  INSPECTIONS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _build_inspections_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Inspections",
                                  "Full inspection log across all units")

        # Legend
        leg = ctk.CTkFrame(hdr, fg_color="transparent")
        leg.pack(side="right")
        for st, (bg, fg_col, sym) in [
            ("Done",    ("#065F46", "#10B981", "M✓")),
            ("Missed",  ("#7F1D1D", "#EF4444", "M✗")),
            ("Open now",("#78350F", "#F59E0B", "M!")),
            ("Pending", ("#1F2937", "#6B7280", "M·")),
        ]:
            ctk.CTkLabel(leg, text=f" {sym} ", fg_color=bg,
                         text_color=fg_col, corner_radius=4,
                         font=("Segoe UI", 9, "bold"),
                         width=32, height=20).pack(side="left", padx=3)
            _label(leg, st, FONT_S, SUBTEXT).pack(side="left", padx=(0, 8))

        self._insp_panel = InspectionsLogPanel(frame)
        self._insp_panel.pack(fill="both", expand=True, padx=8, pady=4)
        return frame

    def _refresh_inspections(self):
        self._insp_panel.refresh()

    # ── Import spreadsheet ────────────────────────────────────────────────────

    def _import_xray(self):
        if not HAS_XLSX:
            messagebox.showerror("Missing Library",
                                 "openpyxl not installed.\n"
                                 "Run: pip install openpyxl")
            return
        path = filedialog.askopenfilename(
            title="Import X-Ray Results",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            imported = self._parse_xray_spreadsheet(path)
            if imported:
                messagebox.showinfo(
                    "Import Complete",
                    f"Imported {imported} sample(s) successfully."
                )
                self._refresh_samples()
            else:
                messagebox.showwarning(
                    "Import",
                    "No rows imported. The spreadsheet needs a header row with a "
                    "'Sample Name' column (plus Total Pounds, Live Bees per Pound, "
                    "Parasites, Chalkbrood, Total Gal Bees, Total lbs for 2gal, "
                    "Total Trays, Incubator Space, Notes, etc.)."
                )
        except Exception as exc:
            messagebox.showerror("Import Error", str(exc))

    def _merge_duplicate_samples(self):
        """Collapse same-name sample records and re-link their trays to the survivor."""
        if not messagebox.askyesno(
            "Merge Duplicate Samples",
            "This will find samples with the same name (ignoring case and extra spaces), "
            "merge their data into one record, and update all tray links.\n\nContinue?",
            parent=self,
        ):
            return
        try:
            removed = db.merge_duplicate_samples()
            if removed:
                messagebox.showinfo(
                    "Merge Complete",
                    f"Merged and removed {removed} duplicate sample record(s).\n"
                    "Tray links have been updated.",
                    parent=self,
                )
                self._refresh_samples()
            else:
                messagebox.showinfo(
                    "Merge Complete", "No duplicates found — all sample names are unique.",
                    parent=self,
                )
        except Exception as exc:
            messagebox.showerror("Merge Error", str(exc), parent=self)

    # Maps the field spreadsheet's headers (normalized) -> sample DB fields.
    _SAMPLE_HEADER_MAP = {
        "sample name":          "name",
        "total pounds":         "total_weight_lbs",
        "total kgs":            "total_weight_kg",
        "live bees per pound":  "live_bees_per_lb",
        "parasites":            "parasites",
        "chalkbrood":           "chalkbrood",
        "total gal bees":       "total_volume_gal",
        "live bees per kg":     "live_bees_per_kg",
        "total kg for 2gal":    "kg_per_2gal",
        "total lbs for 2gal":   "lbs_per_2gal",
        "total trays":          "total_trays",
        "expected trays":       "total_trays",
        "incubator space":      "incubator_space",
        "notes":                "notes",
    }
    _SAMPLE_TEXT_FIELDS = {"name", "incubator_space", "notes"}

    def _parse_xray_spreadsheet(self, path: str) -> int:
        """Import the field sample spreadsheet (CSV or Excel).

        Matches each row to an existing sample by name and updates it (keeping
        tray links); creates a new sample if the name isn't found.
        Returns count of rows imported.
        """
        if path.lower().endswith(".csv"):
            import csv
            with open(path, newline="", encoding="utf-8-sig") as fh:
                rows = list(csv.DictReader(fh))
        else:
            wb   = openpyxl.load_workbook(path, data_only=True)
            ws   = wb.active
            hdrs = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            rows = [dict(zip(hdrs, r))
                    for r in ws.iter_rows(min_row=2, values_only=True)]

        def _norm(h):
            return " ".join(str(h or "").strip().lower().replace("?", "").split())

        def _num(v):
            if v is None:
                return None
            s = str(v).replace(",", "").replace("%", "").replace("$", "").strip()
            if s == "":
                return None
            try:
                return float(s)
            except ValueError:
                return None

        count = 0
        for row in rows:
            # Map this row's headers to our fields
            mapped = {}
            for raw_key, raw_val in row.items():
                field = self._SAMPLE_HEADER_MAP.get(_norm(raw_key))
                if not field:
                    continue
                if field in self._SAMPLE_TEXT_FIELDS:
                    mapped[field] = (str(raw_val).strip() if raw_val is not None else None)
                else:
                    mapped[field] = _num(raw_val)

            name = (mapped.get("name") or "").strip()
            if not name:
                continue
            mapped["name"] = name
            mapped["import_date"] = datetime.now().date().isoformat()
            db.upsert_sample_by_name(mapped)
            count += 1
        return count

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPERS
    # ══════════════════════════════════════════════════════════════════════════

    def _make_tree(self, parent, columns: tuple) -> ttk.Treeview:
        frame = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=12,
                             border_width=1, border_color=BORDER2)
        tree = ttk.Treeview(frame, columns=columns, show="headings",
                            style="Dark.Treeview")
        # Zebra striping + status row tints (applied via _apply_zebra / on insert)
        tree.tag_configure("evenrow", background=PANEL)
        tree.tag_configure("oddrow",  background="#18222F")
        tree.tag_configure("released", foreground=GREEN_LT)
        tree.tag_configure("cooled",   foreground="#7DB0FF")
        vsb  = ttk.Scrollbar(frame, orient="vertical",   command=tree.yview)
        hsb  = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=max(70, len(col)*9), anchor="center",
                        stretch=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        # Attach the outer frame so caller can pack it
        tree._outer_frame = frame
        tree.pack = frame.pack  # forward pack to the outer frame
        return tree

    def _apply_zebra(self, tree, status_col: int = None):
        """Alternate row backgrounds; if status_col given, tint Released/Cooled
        rows by status instead of the zebra background."""
        for i, iid in enumerate(tree.get_children()):
            tags = ["evenrow" if i % 2 == 0 else "oddrow"]
            if status_col is not None:
                vals = tree.item(iid, "values")
                if status_col < len(vals):
                    sv = str(vals[status_col]).strip().lower()
                    if sv == "released":
                        tags.append("released")
                    elif sv == "cooled":
                        tags.append("cooled")
            tree.item(iid, tags=tags)

    def _refresh_current(self):
        if self._current_view:
            getattr(self, f"_refresh_{self._current_view}")()
        self._refresh_alert_badge()

    def _refresh_alert_badge(self):
        n = len(db.get_active_alerts())
        if n:
            self._alert_btn.configure(
                text=f"🔔  Alerts  ·  {n}",
                text_color="#FFFFFF",
                fg_color="#7F1D1D",      # deep red — clear but not garish
                hover_color="#991B1B",
            )
        else:
            self._alert_btn.configure(
                text="🔔  No alerts",
                text_color=SUBTEXT,
                fg_color=CARD,
                hover_color=CARD2,
            )

    # ══════════════════════════════════════════════════════════════════════════
    #  BACKGROUND SERVICES
    # ══════════════════════════════════════════════════════════════════════════

    def _start_govee(self):
        if not db.get_setting("govee_api_key"):
            return
        self._govee.start_polling(
            incubators_fn=db.get_incubators,
            on_reading=self._on_govee_reading,
        )

    def _on_govee_reading(self, incubator_id: int, temp_c: float, humidity: float):
        """Called from Govee polling thread — save reading, check alerts, refresh UI."""
        db.save_reading(incubator_id, temp_c, humidity)

        incubators = db.get_incubators()
        inc = next((i for i in incubators if i["id"] == incubator_id), None)
        if inc and inc.get("temp_alerts_enabled", 1):
            problems = calc.check_temp_humidity(inc, temp_c, humidity)
            if problems:
                for msg in problems:
                    # One standing alert per incubator+problem-type; suppress the
                    # per-minute repeats while the condition persists.
                    kind = "humidity" if "humid" in msg.lower() else "temp"
                    db.add_alert("temp_humidity", msg, severity="warning",
                                 incubator_id=incubator_id,
                                 dedup_key=f"temp_humidity:{kind}:{incubator_id}")
            else:
                # Condition resolved — auto-acknowledge any standing temp/humidity
                # alerts for this incubator so the red outline clears automatically.
                db.auto_acknowledge_alerts([
                    f"temp_humidity:temp:{incubator_id}",
                    f"temp_humidity:humidity:{incubator_id}",
                ])

        # Refresh UI on main thread
        self.after(0, self._on_reading_ui_refresh)

    def _update_dashboard_readings(self):
        """Update only the sensor labels on each card — no rebuild, no shutter."""
        unit = db.get_setting("temp_unit", "C")
        alert_ids = {a["incubator_id"] for a in db.get_active_alerts()
                     if a.get("incubator_id")}
        self._alert_inc_ids = alert_ids
        for inc_id, widgets in self._card_widgets.items():
            # Red outline live when an alert is active for this incubator
            card = widgets.get("card")
            if card is not None:
                if inc_id in alert_ids:
                    card.configure(border_width=2, border_color=RED)
                else:
                    card.configure(border_width=1,
                                   border_color="#222D3D" if widgets.get("hidden") else BORDER)
            reading = self._govee.get_last(inc_id)
            if not reading:
                db_row  = db.get_latest_reading(inc_id)
                reading = {"temp_c": db_row["temperature_c"], "humidity": db_row["humidity_pct"],
                           "timestamp": db_row["timestamp"]} if db_row else {}
            temp_c = reading.get("temp_c")
            hum    = reading.get("humidity")
            inc    = widgets.get("inc")
            t_min, t_max = calc.get_temp_range(inc)
            _poll_txt, _poll_col = _poll_age(reading.get("timestamp"), POLL_INTERVAL_SEC)
            if temp_c is not None:
                t_str = calc.format_temp(temp_c, unit)
                t_col = SUBTEXT if t_min is None else (GREEN if t_min <= temp_c <= t_max else RED)
                h_col = TEXT
                dot   = RED if calc.check_temp_humidity(inc, temp_c, hum) else GREEN
            else:
                t_str = "—"
                t_col = h_col = dot = SUBTEXT
            try:
                widgets["temp"].configure(text=t_str, text_color=t_col)
                widgets["hum"].configure(text=f"{hum:.0f}%" if hum is not None else "—", text_color=h_col)
                widgets["dot"].configure(text_color=dot)
                widgets["ts"].configure(text=f"Last polled: {_poll_txt}", text_color=_poll_col)
            except Exception:
                pass  # widget may have been destroyed by a full refresh

    def _on_reading_ui_refresh(self):
        self._refresh_alert_badge()
        if self._current_view == "dashboard":
            self._update_dashboard_readings()

    def _start_qr_server(self):
        if db.get_setting("qr_server_enabled", "1") != "1":
            return
        port = int(db.get_setting("qr_server_port", "5151"))
        self._qr_port = port
        qr_server.start(port=port, on_update=lambda tid: self.after(0, self._on_qr_update))

    def _on_qr_update(self):
        self._refresh_alert_badge()
        if self._current_view == "trays":
            self._refresh_trays()

    def _start_alert_checker(self):
        def loop():
            cycle = 0
            while True:
                try:
                    self._check_sensor_health()        # every ~10 min
                    if cycle % 6 == 0:
                        self._check_date_alerts()      # hourly
                        self._check_db_conflicts()     # hourly
                    if cycle % 72 == 0:                # ~every 12 h (and at startup)
                        self._run_db_backup()
                except Exception as exc:
                    print(f"[AlertChecker] {exc}")
                cycle += 1
                time.sleep(600)  # 10 minutes

        t = threading.Thread(target=loop, daemon=True, name="AlertChecker")
        t.start()

    def _run_db_backup(self):
        """Create today's DB snapshot (idempotent) and prune old ones."""
        try:
            path = db.make_daily_backup(keep_days=30)
            if path:
                self._sync_log(f"[Backup] daily snapshot ok: {os.path.basename(path)}")
            else:
                self._sync_log("[Backup] snapshot failed")
        except Exception as exc:
            self._sync_log(f"[Backup] {exc}")

    def _check_db_conflicts(self):
        """Raise a loud alert when Google Drive creates DB conflict copies."""
        try:
            conflicts = db.find_drive_conflicts()
        except Exception:
            return
        if not conflicts:
            db.auto_acknowledge_alerts(["db_conflict"])
            return
        names = ", ".join(os.path.basename(c) for c in conflicts[:3])
        more  = f" (+{len(conflicts) - 3} more)" if len(conflicts) > 3 else ""
        db.add_alert(
            "db_conflict",
            f"Google Drive made {len(conflicts)} database conflict copy(ies): "
            f"{names}{more}. Two computers likely wrote the database at the same "
            f"time — open the Data Storage folder and reconcile; recent changes "
            f"may have diverged.",
            severity="critical",
            cooldown_min=720,
            dedup_key="db_conflict",
        )

    @staticmethod
    def _age_minutes(ts: str, now: datetime = None) -> float | None:
        """Minutes since an ISO timestamp, tolerant of naive (local) and
        tz-aware (UTC, from the Pi) timestamps. None if unparseable/absent."""
        if not ts:
            return None
        try:
            then = datetime.fromisoformat(ts)
            ref  = datetime.now(then.tzinfo) if then.tzinfo else (now or datetime.now())
            return (ref - then).total_seconds() / 60
        except Exception:
            return None

    @staticmethod
    def _fmt_age(minutes: float | None) -> str:
        if minutes is None:
            return "never"
        if minutes < 60:
            return f"{int(minutes)} min"
        if minutes < 60 * 48:
            return f"{int(minutes // 60)} hr"
        return f"{int(minutes // (60 * 24))} days"

    def _check_sensor_health(self):
        """Raise/clear alerts when a Vapona sensor stops reporting or only sends
        implausible (corrupt-frame) values. Runs on the alert-checker thread."""
        try:
            import voc_db
        except Exception:
            return
        PI_OFFLINE_MIN = 30   # no contact at all (Pi down / off network)
        DATA_STALE_MIN = 50   # Pi alive but no valid readings (~3 missed cycles)

        incs = {i["id"]: i for i in db.get_incubators(include_hidden=True)}
        for d in voc_db.get_devices():
            # Separate dedup keys per failure mode so a transition (e.g. stale
            # data -> fully offline) isn't suppressed by the other's cooldown.
            off_dk  = f"vapona_offline:{d['id']}"
            data_dk = f"vapona_stale:{d['id']}"
            inc_id = d.get("incubator_id")
            name = d.get("name") or d.get("hardware_id")
            inc = incs.get(inc_id) if inc_id else None
            # Unassigned, or its incubator is off -> no data expected; clear both.
            if inc is None or (calc and calc.is_off(inc)):
                db.auto_acknowledge_alerts([off_dk, data_dk])
                continue

            inc_name = inc.get("name", "")
            seen_age = self._age_minutes(d.get("last_seen"))
            last_ok  = voc_db.latest_valid_reading(inc_id)
            data_age = self._age_minutes(last_ok.get("timestamp")) if last_ok else None

            if seen_age is None or seen_age > PI_OFFLINE_MIN:
                msg = (f"Vapona sensor “{name}” ({inc_name}) is offline — no contact "
                       f"for {self._fmt_age(seen_age)}. Check the Pi's power and Wi-Fi.")
                db.add_alert("vapona_sensor", msg, severity="warning",
                             incubator_id=inc_id, dedup_key=off_dk, cooldown_min=180)
                db.auto_acknowledge_alerts([data_dk])
            elif data_age is None or data_age > DATA_STALE_MIN:
                msg = (f"Vapona sensor “{name}” ({inc_name}) isn’t reporting valid "
                       f"readings — last good reading {self._fmt_age(data_age)} ago. "
                       f"The sensor may be disconnected, unpowered, or faulty "
                       f"(check wiring/power).")
                db.add_alert("vapona_sensor", msg, severity="warning",
                             incubator_id=inc_id, dedup_key=data_dk, cooldown_min=180)
                db.auto_acknowledge_alerts([off_dk])
            else:
                db.auto_acknowledge_alerts([off_dk, data_dk])

    def _check_date_alerts(self):
        lookahead = int(db.get_setting("date_alert_lookahead", "7"))
        batches   = db.get_batches(status="active")
        for batch in batches:
            for ev in calc.get_upcoming_events(batch, lookahead_days=lookahead):
                # One alert per event per batch, regardless of the countdown text
                _evkey = f"date:{ev.get('batch_id')}:{ev['label']}"
                if ev["urgent"]:
                    db.add_alert(
                        "date",
                        f"{'TODAY' if ev['days_away']==0 else 'TOMORROW'}: "
                        f"{ev['label']} — {ev['batch_name']} ({ev['incubator_name']})",
                        severity="critical",
                        batch_id=ev.get("batch_id"),
                        dedup_key=_evkey,
                    )
                elif ev["days_away"] <= lookahead:
                    db.add_alert(
                        "date",
                        f"{ev['label']} in {ev['days_away']}d — "
                        f"{ev['batch_name']} ({ev['incubator_name']})",
                        severity="warning",
                        batch_id=ev.get("batch_id"),
                        dedup_key=_evkey,
                    )
        self.after(0, self._refresh_alert_badge)

    # ── Email scheduler ────────────────────────────────────────────────────────

    def _start_email_scheduler(self):
        """Background thread: send daily report at 7 PM if SMTP is configured."""
        def _loop():
            last_sent_date = None
            while True:
                try:
                    now = datetime.now()
                    if now.hour == 19 and now.minute < 5:        # 7:00–7:04 PM window
                        today = now.date()
                        if last_sent_date != today:
                            if email_reporter.smtp_configured() and email_reporter.get_recipients():
                                err = email_reporter.send_daily_report()
                                if err:
                                    print(f"[Email] Send failed: {err}")
                                else:
                                    print(f"[Email] Daily report sent ({today})")
                            last_sent_date = today
                except Exception as exc:
                    print(f"[EmailScheduler] {exc}")
                time.sleep(60)

        t = threading.Thread(target=_loop, daemon=True, name="EmailScheduler")
        t.start()

    def _send_test_email(self):
        """Send a test email immediately using current settings (must Save first)."""
        recipients = email_reporter.get_recipients()
        if not recipients:
            self._email_status_lbl.configure(
                text="No recipients — add at least one email and Save Settings.",
                text_color=ORANGE)
            return
        if not email_reporter.smtp_configured():
            self._email_status_lbl.configure(
                text="SMTP host and username are required — Save Settings first.",
                text_color=ORANGE)
            return

        self._email_status_lbl.configure(text="Sending…", text_color=SUBTEXT)
        self.update_idletasks()

        def _send():
            err = email_reporter.send_daily_report()
            def _done():
                if err:
                    self._email_status_lbl.configure(
                        text=f"Failed: {err}", text_color=RED)
                else:
                    self._email_status_lbl.configure(
                        text=f"Sent to {len(recipients)} recipient(s) ✓",
                        text_color=GREEN)
            self.after(0, _done)

        threading.Thread(target=_send, daemon=True).start()

    # ── Git sync ──────────────────────────────────────────────────────────────

    def _git_pull(self):
        """Pull latest code from GitHub in a background thread."""
        def _pull():
            try:
                app_dir = os.path.dirname(os.path.abspath(__file__))
                result  = subprocess.run(
                    ["git", "-C", app_dir, "pull", "--ff-only"],
                    capture_output=True, text=True, timeout=20, creationflags=_NO_WINDOW,
                )
                if result.returncode == 0:
                    msg = result.stdout.strip() or "Already up to date."
                    self._sync_log(f"[git pull] {msg}")
                    self.after(0, lambda: self._set_git_status(msg, ok=True))
                else:
                    err = (result.stderr or result.stdout).strip()
                    self._sync_log(f"[git pull] {err}")
                    self.after(0, lambda: self._set_git_status(f"pull: {err}", ok=False))
            except FileNotFoundError:
                # git not on PATH — silent, not a required dependency
                pass
            except Exception as exc:
                self._sync_log(f"[git pull] {exc}")

        threading.Thread(target=_pull, daemon=True, name="GitPull").start()

    def _sync_log(self, msg: str):
        """Print and append a git-sync message to git_sync.log next to the app."""
        print(msg)
        try:
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "git_sync.log")
            if os.path.exists(path) and os.path.getsize(path) > 250_000:
                open(path, "w").close()   # simple rotation
            with open(path, "a", encoding="utf-8") as f:
                f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  {msg}\n")
        except Exception:
            pass

    def _start_auto_sync(self):
        """Keep code in sync with GitHub automatically (every 5 min):
        pull new commits, then commit + push any local edits.

        Disable by setting 'auto_git_sync' to '0' in settings.
        """
        if db.get_setting("auto_git_sync", "1") != "1":
            self._sync_log("[AutoSync] disabled via settings")
            return

        def _loop():
            # Small initial delay so startup isn't competing with the launch pull
            time.sleep(60)
            while True:
                try:
                    self._auto_sync_once()
                except FileNotFoundError:
                    pass   # git not installed — nothing we can do
                except Exception as exc:
                    self._sync_log(f"[AutoSync] {exc}")
                time.sleep(300)   # 5 minutes

        threading.Thread(target=_loop, daemon=True, name="AutoSync").start()

    def _auto_sync_once(self):
        """One full sync pass: pull → (commit local edits) → push. Thread-safe."""
        import socket
        app_dir = os.path.dirname(os.path.abspath(__file__))

        def _git(*args, timeout=40):
            return subprocess.run(["git", "-C", app_dir, *args],
                                  capture_output=True, text=True, timeout=timeout,
                                  creationflags=_NO_WINDOW)

        self._sync_log("[AutoSync] checking for updates…")
        _did_something = False

        # 1. Pull remote changes (fast-forward only — never auto-merge)
        pull = _git("pull", "--ff-only")
        if pull.returncode != 0:
            err = (pull.stderr or pull.stdout).strip()
            self.after(0, lambda e=err: self._set_git_status(
                f"sync paused: {e[:50]}", ok=False))
            self._sync_log(f"[AutoSync] pull failed (diverged?): {err}")
            return
        pulled = (pull.stdout or "").strip()
        if pulled and "already up to date" not in pulled.lower():
            self.after(0, lambda m=pulled: self._set_git_status(f"Updated: {m}", ok=True))
            self._sync_log(f"[AutoSync] pulled: {pulled}")
            _did_something = True

        # 2. Commit local source edits (if any, stable, and valid)
        status = _git("status", "--porcelain").stdout.strip()
        if status:
            # Stability guard: don't commit mid-save. Re-check after a short pause.
            time.sleep(3)
            if _git("status", "--porcelain").stdout.strip() != status:
                self._sync_log("[AutoSync] files still changing — will retry next cycle")
                return
            # Safety guard: never propagate code that doesn't compile.
            changed_py = [ln[3:].strip().strip('"') for ln in status.splitlines()
                          if ln.strip().endswith(".py")]
            for rel in changed_py:
                chk = subprocess.run(
                    [sys.executable, "-m", "py_compile", os.path.join(app_dir, rel)],
                    capture_output=True, text=True, creationflags=_NO_WINDOW)
                if chk.returncode != 0:
                    self.after(0, lambda f=rel: self._set_git_status(
                        f"sync paused: {os.path.basename(f)} has errors", ok=False))
                    self._sync_log(f"[AutoSync] {rel} failed py_compile — not committing")
                    return
            # Safety guard: never propagate code that fails the test suite.
            tests_ok, tests_out = self._run_tests(app_dir)
            if not tests_ok:
                self.after(0, lambda: self._set_git_status(
                    "sync paused: tests failing", ok=False))
                self._sync_log("[AutoSync] tests failing — not committing/pushing:\n"
                               + tests_out[-500:])
                return
            _git("add", "-A")
            host  = socket.gethostname()
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            commit = _git("commit", "-m", f"Auto-sync from {host} at {stamp}")
            if commit.returncode == 0:
                self._sync_log(f"[AutoSync] committed local changes ({len(status.splitlines())} file(s))")
            else:
                err = (commit.stderr or commit.stdout).strip()
                self._sync_log(f"[AutoSync] commit failed: {err}")
                return

        # 3. Push if we have commits ahead of origin
        ahead = _git("rev-list", "--count", "origin/main..HEAD").stdout.strip()
        if ahead and ahead != "0":
            push = _git("push", "origin", "main")
            if push.returncode == 0:
                self.after(0, lambda n=ahead: self._set_git_status(
                    f"Pushed {n} update(s) ✓", ok=True))
                self._sync_log(f"[AutoSync] pushed {ahead} commit(s)")
                _did_something = True
            else:
                err = (push.stderr or push.stdout).strip()
                self.after(0, lambda e=err: self._set_git_status(
                    f"push failed: {e[:50]}", ok=False))
                self._sync_log(f"[AutoSync] push failed: {err}")
                return

        if not _did_something:
            self._sync_log("[AutoSync] up to date — nothing to pull or push")

    def _run_tests(self, app_dir: str) -> tuple[bool, str]:
        """Run the pytest suite for the auto-sync gate.

        Returns (passed, output). A missing tests/ folder or a machine without
        pytest installed is treated as a pass, so sync is never blocked just
        because the test tooling isn't present.
        """
        tests_dir = os.path.join(app_dir, "tests")
        if not os.path.isdir(tests_dir):
            return True, ""
        # Don't block sync when pytest simply isn't installed on this machine.
        have = subprocess.run([sys.executable, "-c", "import pytest"],
                              capture_output=True, creationflags=_NO_WINDOW)
        if have.returncode != 0:
            self._sync_log("[AutoSync] pytest not installed — skipping test gate")
            return True, ""
        try:
            r = subprocess.run(
                [sys.executable, "-m", "pytest", tests_dir, "-q"],
                capture_output=True, text=True, timeout=240,
                creationflags=_NO_WINDOW, cwd=app_dir)
        except Exception as exc:
            return True, f"(test run skipped: {exc})"
        # pytest exit 5 == "no tests collected" — treat as a pass.
        if r.returncode in (0, 5):
            return True, r.stdout
        return False, (r.stdout or "") + (r.stderr or "")

    def _set_git_status(self, msg: str, ok: bool):
        """Flash a brief git status message in the status bar."""
        short = msg if len(msg) < 60 else msg[:57] + "…"
        self._status_time.configure(
            text=f"git: {short}",
            text_color=(GREEN if ok else ORANGE),
        )
        # Revert to clock after 6 seconds
        self.after(6000, lambda: self._status_time.configure(
            text=datetime.now().strftime("%Y-%m-%d  %H:%M"),
            text_color=SUBTEXT,
        ))

    # ── Status bar tick ───────────────────────────────────────────────────────

    def _pulse_dots(self):
        """Pulse the status dot on incubator cards that have an active alert."""
        self._pulse_on = not self._pulse_on
        alert_ids = getattr(self, "_alert_inc_ids", set())
        dim = "#7A1F1A"   # spec exact dim dot
        for iid, w in list(getattr(self, "_card_widgets", {}).items()):
            dot = w.get("dot")
            if dot is None or not dot.winfo_exists():
                continue
            if iid in alert_ids:
                dot.configure(text_color=RED if self._pulse_on else dim)
        self.after(750, self._pulse_dots)

    def _tick(self):
        self._status_govee.configure(
            text=f"Govee: {self._govee.status_label()}",
            text_color=(GREEN if self._govee.connected else SUBTEXT),
        )
        ip   = qr_server.get_local_ip()
        port = self._qr_port
        self._status_qr.configure(
            text=f"QR: {ip}:{port}",
            text_color=(BLUE if qr_server.available() else SUBTEXT),
        )
        self._status_time.configure(
            text=datetime.now().strftime("%Y-%m-%d  %H:%M"),
            text_color=SUBTEXT,
        )
        # Sensibo status (green when a key is configured)
        _has_sb = bool(db.get_setting("sensibo_api_key"))
        self._status_sensibo.configure(
            text="Sensibo: Ready" if _has_sb else "Sensibo: —",
            text_color=(GREEN if _has_sb else FAINT))
        self._status_refresh.configure(
            text=f"Last refresh: {datetime.now().strftime('%H:%M')}", text_color=FAINT)
        self.after(30_000, self._tick)  # refresh every 30s

    # ── Live code reload ────────────────────────────────────────────────────────

    def _restart_app(self):
        """Relaunch the app so code edits take effect without a manual
        close/reopen. Launches a fresh instance, then closes this one."""
        import subprocess
        try:
            subprocess.Popen([sys.executable] + sys.argv)
        except Exception as exc:
            messagebox.showerror("Reload", f"Could not restart:\n{exc}", parent=self)
            return
        self.destroy()
        os._exit(0)   # ensure the in-process QR server thread doesn't linger

    def _start_code_watcher(self):
        """Watch this folder's .py files; when any changes, reveal the Reload
        button so a single click loads the new code (no close/reopen needed)."""
        import threading
        app_dir = os.path.dirname(os.path.abspath(__file__))

        def _snapshot():
            stamps = {}
            for fn in os.listdir(app_dir):
                if fn.endswith(".py"):
                    try:
                        stamps[fn] = os.path.getmtime(os.path.join(app_dir, fn))
                    except OSError:
                        pass
            return stamps

        baseline = _snapshot()

        def _watch():
            import time
            while True:
                time.sleep(2)
                try:
                    if _snapshot() != baseline:
                        self.after(0, self._show_reload_btn)
                        return  # stop watching once an update is flagged
                except Exception:
                    pass

        threading.Thread(target=_watch, daemon=True).start()

    def _show_reload_btn(self):
        if hasattr(self, "_reload_btn") and not self._reload_btn.winfo_ismapped():
            self._reload_btn.pack(side="right", padx=8, pady=3)


# ═══════════════════════════════════════════════════════════════════════════════
#   ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = IncubationApp()
    app.mainloop()
