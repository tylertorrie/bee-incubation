"""
inspection_form.py — Inspection dialog and log panel.

InspectionDialog     — form for recording one inspection
InspectionsLogPanel  — sortable/filterable treeview of all records
make_status_badges   — returns a small CTkFrame with M/E colored chips
"""
import csv
from datetime import datetime
from tkinter import filedialog, messagebox
from tkinter import ttk
import customtkinter as ctk

import inspection_db as idb

# ── Palette (matches main app) ────────────────────────────────────────────────
GOLD    = "#FFD700"
DK_GOLD = "#B8860B"
GREEN   = "#4CAF50"
TEAL    = "#10B981"
AMBER   = "#F59E0B"
RED     = "#EF4444"
BLUE    = "#3B82F6"
CARD    = "#1F2937"
CARD2   = "#263347"
BORDER  = "#374151"
TEXT    = "#F3F4F6"
SUBTEXT = "#9CA3AF"
SIDEBAR = "#111827"
ORANGE  = "#FF9800"

FONT_H = ("Segoe UI", 13, "bold")
FONT_B = ("Segoe UI", 11)
FONT_S = ("Segoe UI", 10)
FONT_XS = ("Segoe UI", 9, "bold")

# ── Badge styles: (bg, text, symbol) ─────────────────────────────────────────
BADGE_STYLE = {
    "done":    ("#065F46", "#10B981", "✓"),
    "missed":  ("#7F1D1D", "#EF4444", "✗"),
    "open":    ("#78350F", "#F59E0B", "!"),
    "pending": ("#1F2937", "#6B7280", "·"),
}
PERIOD_LABEL = {
    "morning": "Morning (6–10 am)",
    "evening": "Evening (4–10 pm)",
    "manual":  "Manual Entry",
}


# ── Widget helpers ────────────────────────────────────────────────────────────

def _lbl(parent, text, font=FONT_B, color=TEXT, **kw):
    return ctk.CTkLabel(parent, text=text, font=font, text_color=color, **kw)


def _btn(parent, text, cmd, width=110, height=30, fg=CARD2,
         hover=BORDER, tc=TEXT, **kw):
    return ctk.CTkButton(parent, text=text, command=cmd, width=width,
                         height=height, fg_color=fg, hover_color=hover,
                         text_color=tc, corner_radius=6, **kw)


def _check(parent, text: str, initial: bool = False) -> ctk.CTkCheckBox:
    var = ctk.BooleanVar(value=initial)
    cb = ctk.CTkCheckBox(parent, text=text, variable=var,
                         fg_color=DK_GOLD, hover_color=GOLD,
                         checkmark_color="black",
                         text_color=TEXT, font=FONT_B)
    cb._var = var   # direct access without calling .cget
    return cb


# ── Public badge helper ───────────────────────────────────────────────────────

def make_status_badges(parent, incubator_id: int, on_click=None) -> ctk.CTkFrame:
    """
    Return a small frame with Morning / Evening inspection pills.

    Colours:
      green ✓  — that inspection has been done today
      red   •  — not done yet (default)

    If `on_click` is given, the pills become buttons; clicking one calls
    on_click(period) where period is "morning" or "evening".
    """
    status = idb.get_inspection_status(incubator_id)
    row    = ctk.CTkFrame(parent, fg_color="transparent")
    pill_font = ("Segoe UI", 11, "bold")

    for period, label, icon in (("morning", "AM", "🌅"), ("evening", "PM", "🌙")):
        done = status.get(period) == "done"
        bg   = "#15803D" if done else "#B91C1C"      # green / red
        hov  = "#16A34A" if done else "#DC2626"
        fg   = "#FFFFFF"
        sym  = "✓" if done else "•"
        text = f"{icon} {label} {sym}"

        if on_click:
            ctk.CTkButton(
                row, text=text, width=70, height=28,
                fg_color=bg, hover_color=hov, text_color=fg,
                corner_radius=14, font=pill_font,
                command=lambda p=period: on_click(p),
            ).pack(side="left", padx=3)
        else:
            ctk.CTkLabel(
                row, text=text, width=70, height=28,
                fg_color=bg, text_color=fg,
                corner_radius=14, font=pill_font,
            ).pack(side="left", padx=3)
    return row


# ══════════════════════════════════════════════════════════════════════════════
#  InspectionDialog
# ══════════════════════════════════════════════════════════════════════════════

class InspectionDialog(ctk.CTkToplevel):
    """
    Form for recording a single incubator inspection.

    Parameters
    ----------
    master        : parent widget
    inc           : incubator dict  — must have 'id' and 'name'
    govee_temp_c  : latest Govee temperature (float) or None
    on_save       : callable() invoked after a successful save
    """

    def __init__(self, master, inc: dict,
                 govee_temp_c: float = None, on_save=None):
        super().__init__(master)
        self._inc        = inc
        self._govee_temp = govee_temp_c
        self._on_save    = on_save
        self._period     = idb.get_current_period()

        self.title(f"Inspection — {inc['name']}")
        self.geometry("500x580")
        self.resizable(False, False)
        self.grab_set()
        self._build()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=(4, 0))

        self._build_header(scroll)
        self._build_temp_section(scroll)
        self._build_checklist(scroll)
        self._build_notes(scroll)

        bf = ctk.CTkFrame(self, fg_color="transparent")
        bf.pack(fill="x", padx=16, pady=12)
        _btn(bf, "Save Inspection", self._save,
             fg=DK_GOLD, hover=GOLD, tc="black", width=155).pack(side="right", padx=4)
        _btn(bf, "Cancel", self.destroy, width=100).pack(side="right")

    def _build_header(self, parent):
        hdr = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        hdr.pack(fill="x", padx=8, pady=(8, 4))

        now_str     = datetime.now().strftime("%a %b %d, %Y  ·  %I:%M %p")
        period_full = PERIOD_LABEL.get(self._period, "Manual Entry")
        period_col  = AMBER if self._period in ("morning", "evening") else SUBTEXT

        _lbl(hdr, self._inc["name"], FONT_H, GOLD).pack(
            anchor="w", padx=14, pady=(10, 2))
        _lbl(hdr, now_str, FONT_S, SUBTEXT).pack(anchor="w", padx=14)
        _lbl(hdr, f"● {period_full}", FONT_S, period_col).pack(
            anchor="w", padx=14, pady=(2, 10))

    def _build_temp_section(self, parent):
        sec = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        sec.pack(fill="x", padx=8, pady=4)
        _lbl(sec, "Temperature", FONT_H, GOLD).pack(
            anchor="w", padx=14, pady=(10, 4))

        g = ctk.CTkFrame(sec, fg_color="transparent")
        g.pack(fill="x", padx=14, pady=(0, 10))
        g.columnconfigure(1, weight=1)

        # Thermometer entry
        _lbl(g, "Thermometer (°C)", FONT_S, SUBTEXT).grid(
            row=0, column=0, sticky="w", padx=(0, 12), pady=5)
        self._temp_entry = ctk.CTkEntry(
            g, placeholder_text="e.g. 27.5", width=130,
            fg_color=CARD2, border_color=BORDER, text_color=TEXT)
        self._temp_entry.grid(row=0, column=1, sticky="w", pady=5)
        self._temp_entry.bind("<KeyRelease>", lambda _e: self._update_temp_feedback())

        # Govee reading
        _lbl(g, "Govee Reading", FONT_S, SUBTEXT).grid(
            row=1, column=0, sticky="w", padx=(0, 12), pady=5)
        if self._govee_temp is not None:
            govee_txt = f"{self._govee_temp:.1f} °C"
            govee_col = TEXT
        else:
            govee_txt = "No reading available"
            govee_col = SUBTEXT
        _lbl(g, govee_txt, FONT_B, govee_col).grid(
            row=1, column=1, sticky="w", pady=5)

        # Feedback label
        self._temp_feedback = _lbl(g, "", FONT_S, SUBTEXT)
        self._temp_feedback.grid(row=2, column=0, columnspan=2, sticky="w", pady=(2, 0))
        self._update_temp_feedback()

    def _build_checklist(self, parent):
        sec = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        sec.pack(fill="x", padx=8, pady=4)
        _lbl(sec, "Checklist", FONT_H, GOLD).pack(
            anchor="w", padx=14, pady=(10, 6))

        items = [
            ("heat_pumps_ok",      "Heat pumps on and working (both)",  True),
            ("parasites_emerging", "Parasites emerging",                False),
            ("bees_emerging",      "Bees emerging",                     False),
            ("fans_ok",            "Fans on and working",               True),
            ("black_lights_ok",    "All black lights working",          True),
        ]
        self._checks = {}
        for key, label, default in items:
            cb = _check(sec, label, default)
            cb.pack(anchor="w", padx=20, pady=4)
            self._checks[key] = cb

        ctk.CTkFrame(sec, fg_color="transparent", height=6).pack()

    def _build_notes(self, parent):
        sec = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=10)
        sec.pack(fill="x", padx=8, pady=4)
        _lbl(sec, "Notes", FONT_H, GOLD).pack(
            anchor="w", padx=14, pady=(10, 4))
        self._notes = ctk.CTkTextbox(
            sec, height=75, fg_color=CARD2,
            border_color=BORDER, text_color=TEXT)
        self._notes.pack(fill="x", padx=14, pady=(0, 12))

    # ── Temp feedback ─────────────────────────────────────────────────────────

    def _update_temp_feedback(self):
        raw = self._temp_entry.get().strip()
        if not raw:
            self._temp_feedback.configure(text="", text_color=SUBTEXT)
            return
        try:
            thermo = float(raw)
        except ValueError:
            self._temp_feedback.configure(
                text="Enter a valid number", text_color=RED)
            return

        if self._govee_temp is None:
            self._temp_feedback.configure(
                text="No Govee reading to compare against", text_color=SUBTEXT)
            return

        diff = abs(thermo - self._govee_temp)
        if diff > idb.TEMP_ALERT_THRESHOLD:
            self._temp_feedback.configure(
                text=f"⚠  Δ {diff:.1f}°C from Govee — OUT OF RANGE (>{idb.TEMP_ALERT_THRESHOLD:.0f}°C)",
                text_color=RED)
        else:
            self._temp_feedback.configure(
                text=f"✓  Δ {diff:.1f}°C from Govee — OK",
                text_color=TEAL)

    # ── Save ──────────────────────────────────────────────────────────────────

    def _save(self):
        thermo_c   = None
        temp_alert = False
        temp_diff  = None
        raw = self._temp_entry.get().strip()
        if raw:
            try:
                thermo_c = float(raw)
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "Temperature must be a number.", parent=self)
                return
            if self._govee_temp is not None:
                temp_diff  = abs(thermo_c - self._govee_temp)
                temp_alert = temp_diff > idb.TEMP_ALERT_THRESHOLD

        data = {
            "incubator_id":       self._inc["id"],
            "period":             self._period,
            "thermometer_temp_c": thermo_c,
            "govee_temp_c":       self._govee_temp,
            "temp_diff_c":        temp_diff,
            "temp_alert":         temp_alert,
            "notes":              self._notes.get("1.0", "end").strip(),
        }
        data.update({k: cb._var.get() for k, cb in self._checks.items()})

        idb.save_inspection(data)

        # Log an app alert if temp is out of range
        if temp_alert:
            try:
                import incubation_db as _mdb
                _mdb.add_alert(
                    "inspection_temp",
                    (f"Inspection temp alert — {self._inc['name']}: "
                     f"Thermometer {thermo_c:.1f}°C vs "
                     f"Govee {self._govee_temp:.1f}°C "
                     f"(Δ {temp_diff:.1f}°C)"),
                    severity="warning",
                    incubator_id=self._inc["id"],
                    dedup_key=f"inspection_temp:{self._inc['id']}",
                )
            except Exception:
                pass

        if self._on_save:
            self._on_save()
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  InspectionsLogPanel
# ══════════════════════════════════════════════════════════════════════════════

class InspectionsLogPanel(ctk.CTkFrame):
    """
    Sortable, filterable spreadsheet-style log of inspection records.

    Parameters
    ----------
    fixed_incubator_id : int or None
        If set, locks the panel to one incubator and hides the incubator
        filter dropdown.  Used when embedded inside an incubator detail tab.
    """

    COLS = (
        "Date / Time", "Incubator", "Period",
        "Thermo °C", "Govee °C", "Δ Temp",
        "Heat Pumps", "Parasites", "Bees", "Fans", "Lights",
        "Notes",
    )
    _COL_KEY = {
        "Date / Time": "timestamp",
        "Incubator":   "incubator_name",
        "Period":      "period",
        "Thermo °C":   "thermometer_temp_c",
        "Govee °C":    "govee_temp_c",
        "Δ Temp":      "temp_diff_c",
        "Heat Pumps":  "heat_pumps_ok",
        "Parasites":   "parasites_emerging",
        "Bees":        "bees_emerging",
        "Fans":        "fans_ok",
        "Lights":      "black_lights_ok",
        "Notes":       "notes",
    }
    _COL_W = {
        "Date / Time": 145, "Incubator": 120, "Period": 75,
        "Thermo °C": 85, "Govee °C": 85, "Δ Temp": 68,
        "Heat Pumps": 92, "Parasites": 82, "Bees": 62,
        "Fans": 62, "Lights": 62, "Notes": 220,
    }

    def __init__(self, master, fixed_incubator_id: int = None, **kwargs):
        kwargs.setdefault("fg_color", "transparent")
        super().__init__(master, **kwargs)
        self._fixed_id  = fixed_incubator_id
        self._sort_col  = None
        self._sort_rev  = False
        self._cache     = []
        self._build()
        self.refresh()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        # Filter bar
        fbar = ctk.CTkFrame(self, fg_color=CARD, corner_radius=8)
        fbar.pack(fill="x", padx=8, pady=(8, 4))

        _lbl(fbar, "Filter:", FONT_S, SUBTEXT).pack(side="left", padx=(10, 4), pady=7)

        # Incubator dropdown (hidden in fixed mode)
        self._flt_inc = None
        self._inc_map = None
        if self._fixed_id is None:
            import incubation_db as _db
            incubators     = _db.get_incubators()
            self._inc_map  = {"All Incubators": None}
            self._inc_map.update({i["name"]: i["id"] for i in incubators})
            self._flt_inc  = ctk.CTkComboBox(
                fbar, values=list(self._inc_map.keys()), width=180,
                fg_color=CARD2, border_color=BORDER, text_color=TEXT,
                command=lambda _: self.refresh())
            self._flt_inc.set("All Incubators")
            self._flt_inc.pack(side="left", padx=4, pady=7)

        # Period dropdown
        self._flt_period = ctk.CTkComboBox(
            fbar,
            values=["All Periods", "morning", "evening", "manual"],
            width=130, fg_color=CARD2, border_color=BORDER, text_color=TEXT,
            command=lambda _: self.refresh())
        self._flt_period.set("All Periods")
        self._flt_period.pack(side="left", padx=4, pady=7)

        # Record count
        self._count_lbl = _lbl(fbar, "", FONT_S, SUBTEXT)
        self._count_lbl.pack(side="left", padx=10)

        # Export buttons
        _btn(fbar, "Export CSV", self._export_csv,
             fg=BORDER, hover=CARD2, width=105, height=26).pack(
             side="right", padx=4, pady=7)
        try:
            import openpyxl as _  # noqa
            _btn(fbar, "Export Excel", self._export_xlsx,
                 fg=BORDER, hover=CARD2, width=115, height=26).pack(
                 side="right", padx=4, pady=7)
        except ImportError:
            pass

        # Treeview
        self._build_tree()

        # Bottom action row
        btm = ctk.CTkFrame(self, fg_color="transparent")
        btm.pack(fill="x", padx=8, pady=(4, 8))
        _btn(btm, "Delete Selected", self._delete_selected,
             fg="#4B0000", hover=RED, width=140, height=26).pack(side="right", padx=4)
        _btn(btm, "Refresh", self.refresh,
             fg=BORDER, hover=CARD2, width=100, height=26).pack(side="right", padx=4)

    def _build_tree(self):
        outer = ctk.CTkFrame(self, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=8, pady=2)

        # Ensure treeview style exists
        style = ttk.Style()
        try:
            style.configure("Dark.Treeview",
                background=CARD, foreground=TEXT,
                fieldbackground=CARD, borderwidth=0,
                rowheight=26, font=("Segoe UI", 10))
            style.configure("Dark.Treeview.Heading",
                background=SIDEBAR, foreground=GOLD,
                relief="flat", font=("Segoe UI", 10, "bold"))
            style.map("Dark.Treeview",
                background=[("selected", "#3B4F6B")],
                foreground=[("selected", TEXT)])
        except Exception:
            pass

        self._tree = ttk.Treeview(
            outer, columns=self.COLS, show="headings",
            style="Dark.Treeview", selectmode="extended")
        vsb = ttk.Scrollbar(outer, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(outer, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        for col in self.COLS:
            w = self._COL_W.get(col, 80)
            self._tree.heading(col, text=col,
                               command=lambda c=col: self._sort_by(c))
            self._tree.column(col, width=w, anchor="center",
                              stretch=(col == "Notes"))

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        # Row tags
        self._tree.tag_configure("alert",
            background="#2D1515", foreground="#FCA5A5")
        self._tree.tag_configure("parasites",
            background="#1A2E1A", foreground="#86EFAC")
        self._tree.tag_configure("bees",
            background="#1A2E2E", foreground="#67E8F9")

    # ── Data ─────────────────────────────────────────────────────────────────

    def refresh(self):
        if self._fixed_id is not None:
            inc_id = self._fixed_id
        elif self._inc_map and self._flt_inc:
            inc_id = self._inc_map.get(self._flt_inc.get())
        else:
            inc_id = None

        rows   = idb.get_inspections(incubator_id=inc_id)
        period = self._flt_period.get()
        if period != "All Periods":
            rows = [r for r in rows if r.get("period") == period]

        self._cache = rows
        self._populate(rows)

    def _sort_by(self, col: str):
        self._sort_rev = (not self._sort_rev) if self._sort_col == col else False
        self._sort_col = col
        self._populate(self._cache)

    def _populate(self, rows: list):
        if self._sort_col:
            db_key = self._COL_KEY.get(self._sort_col, self._sort_col)
            rows   = sorted(rows,
                            key=lambda r: (r.get(db_key) is None,
                                           r.get(db_key) or ""),
                            reverse=self._sort_rev)

        self._tree.delete(*self._tree.get_children())
        for r in rows:
            ts  = (r.get("timestamp") or "")[:16].replace("T", " ")
            td  = r.get("temp_diff_c")

            if r.get("temp_alert"):
                tag = ("alert",)
            elif r.get("parasites_emerging"):
                tag = ("parasites",)
            elif r.get("bees_emerging"):
                tag = ("bees",)
            else:
                tag = ()

            def _yn(v):  return "Yes" if v else "No"
            def _f1(v):  return f"{v:.1f}" if v is not None else "—"

            delta_str = "—"
            if td is not None:
                delta_str = f"⚠ {td:.1f}" if r.get("temp_alert") else f"{td:.1f}"

            self._tree.insert("", "end", iid=str(r["id"]), tags=tag, values=(
                ts,
                r.get("incubator_name") or "—",
                (r.get("period") or "manual").capitalize(),
                _f1(r.get("thermometer_temp_c")),
                _f1(r.get("govee_temp_c")),
                delta_str,
                _yn(r.get("heat_pumps_ok")),
                _yn(r.get("parasites_emerging")),
                _yn(r.get("bees_emerging")),
                _yn(r.get("fans_ok")),
                _yn(r.get("black_lights_ok")),
                r.get("notes") or "",
            ))

        n = len(rows)
        self._count_lbl.configure(
            text=f"{n} record{'s' if n != 1 else ''}")

    # ── Actions ───────────────────────────────────────────────────────────────

    def _delete_selected(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Delete",
                "Select one or more records first.", parent=self)
            return
        if not messagebox.askyesno(
                "Delete",
                f"Delete {len(sel)} inspection record{'s' if len(sel)>1 else ''}?",
                parent=self):
            return
        for iid in sel:
            try:
                idb.delete_inspection(int(iid))
            except Exception:
                pass
        self.refresh()

    # ── Export helpers ────────────────────────────────────────────────────────

    _EXPORT_HEADERS = [
        "Date/Time", "Incubator", "Period",
        "Thermometer_C", "Govee_C", "Temp_Diff_C", "Temp_Alert",
        "Heat_Pumps_OK", "Parasites_Emerging", "Bees_Emerging",
        "Fans_OK", "Black_Lights_OK", "Notes",
    ]

    def _row_to_list(self, r: dict) -> list:
        def yn(v): return "Yes" if v else "No"
        return [
            (r.get("timestamp") or "")[:16].replace("T", " "),
            r.get("incubator_name") or "",
            (r.get("period") or "").capitalize(),
            r.get("thermometer_temp_c") or "",
            r.get("govee_temp_c") or "",
            r.get("temp_diff_c") or "",
            yn(r.get("temp_alert")),
            yn(r.get("heat_pumps_ok")),
            yn(r.get("parasites_emerging")),
            yn(r.get("bees_emerging")),
            yn(r.get("fans_ok")),
            yn(r.get("black_lights_ok")),
            r.get("notes") or "",
        ]

    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            title="Export Inspections",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="inspections.csv",
        )
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(self._EXPORT_HEADERS)
            for r in self._cache:
                w.writerow(self._row_to_list(r))
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=self)

    def _export_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl not installed.\npip install openpyxl", parent=self)
            return

        friendly = [
            "Date / Time", "Incubator", "Period",
            "Thermometer (°C)", "Govee (°C)", "Temp Diff (°C)", "Temp Alert",
            "Heat Pumps OK", "Parasites Emerging", "Bees Emerging",
            "Fans OK", "Black Lights OK", "Notes",
        ]
        path = filedialog.asksaveasfilename(
            title="Export Inspections",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="inspections.xlsx",
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Inspections"

        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFD700")
        for ci, h in enumerate(friendly, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        alert_fill = PatternFill("solid", fgColor="7F1D1D")
        for ri, r in enumerate(self._cache, 2):
            ws.append(self._row_to_list(r))
            if r.get("temp_alert"):
                for ci in range(1, len(friendly) + 1):
                    ws.cell(row=ri, column=ci).fill = alert_fill

        col_widths = [18, 16, 10, 15, 12, 12, 10, 12, 17, 14, 10, 14, 40]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(ci)].width = w

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=self)
