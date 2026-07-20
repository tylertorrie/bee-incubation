"""
views/samples_view.py  —  the Samples screen plus x-ray spreadsheet import.

Extracted from incubation_app.py as a mixin (pure relocation).
"""
import csv
import math
from datetime import datetime, date
from tkinter import filedialog, messagebox

import customtkinter as ctk

import incubation_db as db

from ui_theme import (
    GOLD, DK_GOLD, GREEN, GREEN_LT, TEAL, ORANGE, RED, RED_LT, BLUE, LINK,
    BG, BARBG, SIDEBAR, RIGHTPANE, CARD, PANEL, NESTED, CARD2,
    BORDER, BORDER2, SUBBORDER, TEXT, TEXT2, SUBTEXT, FAINT,
    FONT_H, FONT_B, FONT_S, MODE_COLORS, MODE_BADGE_BG,
    _treeview_style, _label, _btn, _btn_primary, _btn_secondary,
    _entry, _combo, _mix, _poll_age, _FormRow,
)

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


class SamplesViewMixin:
    """The samples screen plus x-ray spreadsheet import."""

    def _build_samples_view(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._main, fg_color="transparent", corner_radius=0)

        hdr = self._screen_header(frame, "Samples",
                                  "X-Ray results & sample records")
        _btn_primary(hdr, "+ Add Sample", lambda: self._open_sample_dialog(),
                     width=130).pack(side="right")
        _btn_secondary(hdr, "Merge Duplicates", self._merge_duplicate_samples,
                       width=140).pack(side="right", padx=8)
        _btn_secondary(hdr, "Import Spreadsheet", self._import_xray,
                       width=150).pack(side="right")

        # Search box — filter the sample list by name as you type
        sbar = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=10,
                            border_width=1, border_color=BORDER2)
        sbar.pack(fill="x", padx=12, pady=(0, 6))
        _label(sbar, "Search:", FONT_S, SUBTEXT).pack(side="left", padx=10, pady=6)
        self._smp_search = ctk.CTkEntry(
            sbar, placeholder_text="Sample name…", width=260,
            fg_color=CARD2, border_color=BORDER, text_color=TEXT)
        self._smp_search.pack(side="left", padx=6, pady=6)
        self._smp_search.bind("<KeyRelease>", lambda e: self._refresh_samples())

        # Year filter — samples used by trays started in the chosen year
        _smp_years = self._tray_years()
        self._smp_year = _combo(sbar, ["All Years"] + _smp_years, 120)
        _cur_year = str(datetime.now().year)
        self._smp_year.set(_cur_year if _cur_year in _smp_years else "All Years")
        self._smp_year.pack(side="left", padx=6, pady=6)
        self._smp_year.configure(command=lambda _: self._refresh_samples())

        self._smp_count_lbl = _label(sbar, "", FONT_S, SUBTEXT)
        self._smp_count_lbl.pack(side="right", padx=12, pady=6)

        # Label-grid table (per the CTk guide) — link-blue names, styled cells
        tbl = ctk.CTkFrame(frame, fg_color=PANEL, corner_radius=12,
                           border_width=1, border_color=BORDER2)
        tbl.pack(fill="both", expand=True, padx=12, pady=4)
        self._smp_scroll = ctk.CTkScrollableFrame(tbl, fg_color="transparent")
        self._smp_scroll.pack(fill="both", expand=True, padx=2, pady=2)
        self._smp_sort_col = 0
        self._smp_sort_asc = True
        return frame

    # (label, grid weight, align 'w'|'e', kind)
    _SMP_COLS = [
        ("Name", 3, "w", "link"), ("Total Kg", 1, "e", "num"),
        ("Live Bees/Kg", 2, "e", "num"), ("Parasites", 1, "e", "num"),
        ("Chalkbrood", 1, "e", "num"), ("Total Gal", 1, "e", "num"),
        ("Kg for 2gal", 1, "e", "num"), ("Expected", 1, "e", "num"),
        ("Actual", 1, "e", "num"), ("Inc. Space", 1, "w", "text"),
        ("Notes", 3, "w", "text"),
    ]

    def _sort_sample_tree(self, col_idx: int):
        if self._smp_sort_col == col_idx:
            self._smp_sort_asc = not self._smp_sort_asc
        else:
            self._smp_sort_col = col_idx
            self._smp_sort_asc = True
        self._refresh_samples()

    def _refresh_samples(self):
        scroll = self._smp_scroll
        for w in scroll.winfo_children():
            w.destroy()

        def _n(v, dec=1):
            return f"{v:,.{dec}f}" if isinstance(v, (int, float)) else "—"

        def _kg(lbs_val, kg_val, dec=1):
            if isinstance(kg_val, (int, float)):
                return f"{kg_val:,.{dec}f}"
            if isinstance(lbs_val, (int, float)):
                return f"{lbs_val * 0.45359237:,.{dec}f}"
            return "—"

        def _per_kg(per_lb_val, per_kg_val, dec=0):
            if isinstance(per_kg_val, (int, float)):
                return f"{per_kg_val:,.{dec}f}"
            if isinstance(per_lb_val, (int, float)):
                return f"{per_lb_val / 0.45359237:,.{dec}f}"
            return "—"

        q = ""
        if getattr(self, "_smp_search", None) is not None:
            q = self._smp_search.get().strip().lower()

        samples = db.get_samples()
        actual_counts = db.get_tray_counts_by_sample(statuses=None)

        yr = getattr(self, "_smp_year", None)
        yr = yr.get() if yr else "All Years"
        if yr and yr != "All Years":
            year_ids = db.current_year_sample_ids(int(yr))
            samples = [s for s in samples if s["id"] in year_ids]

        # Build (id, [display values]) rows
        data = []
        for s in samples:
            if q and q not in (s["name"] or "").lower():
                continue
            data.append((s["id"], [
                s["name"],
                _kg(s.get("total_weight_lbs"), s.get("total_weight_kg")),
                _per_kg(s.get("live_bees_per_lb"), s.get("live_bees_per_kg")),
                _n(s.get("parasites")),
                _n(s.get("chalkbrood")),
                _n(s.get("total_volume_gal")),
                _kg(s.get("lbs_per_2gal"), s.get("kg_per_2gal"), 2),
                (str(math.ceil(s["total_trays"]))
                 if isinstance(s.get("total_trays"), (int, float)) else "—"),
                str(actual_counts.get(s["id"], 0)),
                s.get("incubator_space") or "—",
                (s.get("notes") or "").replace("\n", " "),
            ]))

        # Sort by the active column
        sc = self._smp_sort_col

        def _key(row):
            v = row[1][sc]
            if v is None or v == "—":
                return (2, 0.0, "")
            try:
                return (0, float(str(v).replace("%", "").replace(",", "")), "")
            except (ValueError, AttributeError):
                return (1, 0.0, str(v).lower())
        data.sort(key=lambda r: str(r[1][0]).lower())
        data.sort(key=_key, reverse=not self._smp_sort_asc)

        # Header row (clickable to sort)
        hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 2))
        for ci, (lbl, wt, al, kind) in enumerate(self._SMP_COLS):
            hdr.columnconfigure(ci, weight=wt, uniform="smp")
            arrow = (" ↑" if self._smp_sort_asc else " ↓") if ci == sc else ""
            ctk.CTkButton(
                hdr, text=lbl + arrow, height=30, corner_radius=6,
                anchor="w" if al == "w" else "e",
                fg_color="transparent", hover_color=CARD2, text_color=GOLD,
                font=("Segoe UI", 11, "bold"),
                command=lambda c=ci: self._sort_sample_tree(c)
            ).grid(row=0, column=ci, sticky="ew", padx=1)
        ctk.CTkFrame(scroll, fg_color=BORDER2, height=1).pack(fill="x")

        # Data rows
        for i, (sid, vals) in enumerate(data):
            rf = ctk.CTkFrame(scroll, fg_color=PANEL if i % 2 == 0 else "#18222F",
                              corner_radius=0)
            rf.pack(fill="x")
            for ci, (lbl, wt, al, kind) in enumerate(self._SMP_COLS):
                rf.columnconfigure(ci, weight=wt, uniform="smp")
                col = LINK if kind == "link" else (
                    "#E5E7EB" if al == "e" else "#CBD5E1")
                _label(rf, str(vals[ci]), ("Segoe UI", 11), col,
                       anchor="w" if al == "w" else "e", wraplength=200 if kind == "text" else 0
                       ).grid(row=0, column=ci, sticky="ew", padx=8, pady=6)
            rf.bind("<Double-1>", lambda e, s=sid: self._open_sample_by_id(s))
            for ch in rf.winfo_children():
                ch.bind("<Double-1>", lambda e, s=sid: self._open_sample_by_id(s))

        if getattr(self, "_smp_count_lbl", None) is not None:
            self._smp_count_lbl.configure(
                text=f"{len(data)} of {len(samples)} samples" if q
                else f"{len(samples)} samples")

    def _open_sample_by_id(self, sid: int):
        sample = next((s for s in db.get_samples() if s["id"] == sid), None)
        if sample:
            self._open_sample_dialog(sample)

    # ══════════════════════════════════════════════════════════════════════════
    #  ANALYTICS VIEW
    # ══════════════════════════════════════════════════════════════════════════

    def _import_xray(self):
        if not HAS_XLSX:
            messagebox.showerror("Missing Library",
                                 "openpyxl not installed.\n"
                                 "Run: pip install openpyxl")
            return
        path = filedialog.askopenfilename(
            title="Import X-Ray Results",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            imported = self._parse_xray_spreadsheet(path)
            if imported:
                messagebox.showinfo(
                    "Import Complete",
                    f"Imported {imported} sample(s) successfully."
                )
                self._refresh_samples()
            else:
                messagebox.showwarning(
                    "Import",
                    "No rows imported. The spreadsheet needs a header row with a "
                    "'Sample Name' column (plus Total Pounds, Live Bees per Pound, "
                    "Parasites, Chalkbrood, Total Gal Bees, Total lbs for 2gal, "
                    "Total Trays, Incubator Space, Notes, etc.)."
                )
        except Exception as exc:
            messagebox.showerror("Import Error", str(exc))

    def _merge_duplicate_samples(self):
        """Collapse same-name sample records and re-link their trays to the survivor."""
        if not messagebox.askyesno(
            "Merge Duplicate Samples",
            "This will find samples with the same name (ignoring case and extra spaces), "
            "merge their data into one record, and update all tray links.\n\nContinue?",
            parent=self,
        ):
            return
        try:
            removed = db.merge_duplicate_samples()
            if removed:
                messagebox.showinfo(
                    "Merge Complete",
                    f"Merged and removed {removed} duplicate sample record(s).\n"
                    "Tray links have been updated.",
                    parent=self,
                )
                self._refresh_samples()
            else:
                messagebox.showinfo(
                    "Merge Complete", "No duplicates found — all sample names are unique.",
                    parent=self,
                )
        except Exception as exc:
            messagebox.showerror("Merge Error", str(exc), parent=self)

    # Maps the field spreadsheet's headers (normalized) -> sample DB fields.
    _SAMPLE_HEADER_MAP = {
        "sample name":          "name",
        "total pounds":         "total_weight_lbs",
        "total kgs":            "total_weight_kg",
        "live bees per pound":  "live_bees_per_lb",
        "parasites":            "parasites",
        "chalkbrood":           "chalkbrood",
        "total gal bees":       "total_volume_gal",
        "live bees per kg":     "live_bees_per_kg",
        "total kg for 2gal":    "kg_per_2gal",
        "total lbs for 2gal":   "lbs_per_2gal",
        "total trays":          "total_trays",
        "expected trays":       "total_trays",
        "incubator space":      "incubator_space",
        "notes":                "notes",
    }
    _SAMPLE_TEXT_FIELDS = {"name", "incubator_space", "notes"}

    def _parse_xray_spreadsheet(self, path: str) -> int:
        """Import the field sample spreadsheet (CSV or Excel).

        Matches each row to an existing sample by name and updates it (keeping
        tray links); creates a new sample if the name isn't found.
        Returns count of rows imported.
        """
        if path.lower().endswith(".csv"):
            import csv
            with open(path, newline="", encoding="utf-8-sig") as fh:
                rows = list(csv.DictReader(fh))
        else:
            wb   = openpyxl.load_workbook(path, data_only=True)
            ws   = wb.active
            hdrs = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            rows = [dict(zip(hdrs, r))
                    for r in ws.iter_rows(min_row=2, values_only=True)]

        def _norm(h):
            return " ".join(str(h or "").strip().lower().replace("?", "").split())

        def _num(v):
            if v is None:
                return None
            s = str(v).replace(",", "").replace("%", "").replace("$", "").strip()
            if s == "":
                return None
            try:
                return float(s)
            except ValueError:
                return None

        count = 0
        for row in rows:
            # Map this row's headers to our fields
            mapped = {}
            for raw_key, raw_val in row.items():
                field = self._SAMPLE_HEADER_MAP.get(_norm(raw_key))
                if not field:
                    continue
                if field in self._SAMPLE_TEXT_FIELDS:
                    mapped[field] = (str(raw_val).strip() if raw_val is not None else None)
                else:
                    mapped[field] = _num(raw_val)

            name = (mapped.get("name") or "").strip()
            if not name:
                continue
            mapped["name"] = name
            mapped["import_date"] = datetime.now().date().isoformat()
            db.upsert_sample_by_name(mapped)
            count += 1
        return count

    # ══════════════════════════════════════════════════════════════════════════
    #  HELPERS
    # ══════════════════════════════════════════════════════════════════════════

